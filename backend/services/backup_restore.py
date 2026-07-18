"""
Backup Restore: importa un bundle ZIP prodotto da `build_backup_zip_bytes` o
`build_full_backup_zip_bytes`.

Bundle base (PCM_backup_*.zip):

    ├── schema.xlsx              (4 sheet: Motivations, Parameters, Questions, QAM)
    ├── languages_metadata.xlsx  (1 sheet: Languages)
    ├── glossary.xlsx            (1 sheet: Glossary)
    └── languages/
        ├── <ID>.xlsx            (Database_model esteso, Answers, Examples, Admin Notes)
        └── ...

Bundle full (PCM_full_backup_*.zip): stesso contenuto del bundle base + cartella
`extras/` con aliases, users (senza hash password), site_content, submissions,
parameter_submissions, archived_questions, parameter_change_logs,
parameter_flags, legal_documents (+ i PDF dell'archivio legale sotto
extras/legal_pdfs/) e entity_versions.jsonl (History).

Strategia (sempre upsert, mai delete):
  1. schema.xlsx          -> upsert Motivations / Parameters / Questions / QAM
  2. languages_metadata   -> upsert Language per ID (creando o aggiornando metadata)
  3. glossary.xlsx        -> upsert Glossary per word
  4. languages/<id>.xlsx  -> REPLACE compilazione (Answer/Example/AnswerMotivation)
                            + ripristino admin_note per (lang, param)
  5. extras/* (se presenti):
     - aliases.xlsx            -> upsert per chiave naturale `old_id`
     - users.xlsx              -> upsert per email; MAI toccata la password
       degli esistenti, i nuovi nascono con password inutilizzabile
     - site_content.xlsx       -> upsert per chiave naturale `key`
     - parameter_flags.xlsx    -> upsert per chiave naturale (lang, param):
       i flag is_unsure/needs_review vengono riallineati al bundle
     - legal_documents.xlsx    -> upsert per (type, version) e (user, doc);
       i PDF in extras/legal_pdfs/ vengono riscritti su disco se assenti
     - entity_versions.jsonl   -> insert-if-missing con dedupe su
       (entity_type, entity_id, operation, created_at)
     - submissions/parameter_submissions/archived_questions/
       parameter_change_logs: ripristinati SOLO se wipe=True. Hanno PK
       auto-increment senza chiave naturale: ricreare senza wipe genererebbe
       duplicati di snapshot. Le tabelle snapshot andrebbero rifatte fresche
       o non toccate, mai mergiate.

Le entità non menzionate restano in DB (no delete). Per un wipe-and-restore
totale usare `wipe=True`: tronca le tabelle dati prima di importare. Gli utenti
non vengono toccati.

Fasi tracciate via `services.migration_progress` (riusato — stesso pattern dei
job di Migration Import). Ogni xlsx del bundle è una "tick" individuale per la
fase di compilation; le altre fasi sono single-tick.

Endpoint chiamante: routers/backup_restore.py
"""
from __future__ import annotations
from dataclasses import dataclass, field, asdict
from typing import List, Optional, Dict, Any
import io
import json
import os
import zipfile

from openpyxl import load_workbook
from sqlalchemy.orm import Session

import models
from config import LEGAL_DOCUMENTS_DIR
from services.language_alias import resolve_language
from services.parameter_alias import resolve_parameter
from services.question_alias import resolve_question
from services.excel_import import import_excel, ImportReport
from services.migration_progress import ProgressReporter, NULL_PROGRESS
from services.dag_eval import run_dag_for_language


# ============================================================================
# Report
# ============================================================================

@dataclass
class BackupRestoreReport:
    files_processed: List[str] = field(default_factory=list)
    files_skipped: List[str] = field(default_factory=list)
    errors: List[Dict[str, Any]] = field(default_factory=list)
    by_file: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    languages_restored: List[str] = field(default_factory=list)
    languages_failed: List[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "files_processed": self.files_processed,
            "files_skipped": self.files_skipped,
            "errors": self.errors,
            "by_file": self.by_file,
            "languages_restored": self.languages_restored,
            "languages_failed": self.languages_failed,
            "total_errors": len(self.errors),
        }


def _absorb_sub_report(report: BackupRestoreReport, file_name: str, sub: ImportReport) -> None:
    """Copia gli errori e summary di una `import_excel` interna nel report globale."""
    report.files_processed.append(file_name)
    report.by_file[file_name] = {
        "sheets_processed": sub.sheets_processed,
        "by_sheet": {k: v.to_dict() for k, v in sub.by_sheet.items()},
        "errors_count": len(sub.errors),
    }
    for e in sub.errors:
        d = e.to_dict()
        d["_file"] = file_name
        report.errors.append(d)


# ============================================================================
# Wipe (operazione facoltativa, distruttiva)
# ============================================================================

# Stesse tabelle wipate da migration_import.import_migration_bundle (in ordine
# FK-safe). Le copio qui perché backup_restore non vuole dipendere dall'altro
# servizio: condividono solo lo schema dati.
_WIPE_TABLES_FK_SAFE = [
    "answer_motivations",
    "examples",
    "answers",
    "language_parameter_evals",
    "language_parameters",
    "language_parameter_statuses",
    "submission_answer_motivations",
    "submission_examples",
    "submission_params",
    "submission_answers",
    "submissions",
    # Archive question tree (figli prima dei genitori)
    "archived_answer_motivations",
    "archived_examples",
    "archived_answers",
    "archived_question_motivations",
    "archived_questions",
    # Parameter submissions (snapshot definizioni parametri)
    "parameter_submission_allowed_motivations",
    "parameter_submission_questions",
    "parameter_submissions",
    "question_allowed_motivations",
    "questions",
    "parameter_change_logs",
    "parameter_defs",
    "motivations",
    "languages",
    "groups",
    "families",
    "top_families",
    "glossary",
    # Contenuti dinamici editabili (HowToCite/About)
    "site_contents",
]


def _wipe_data(db: Session) -> None:
    from sqlalchemy import text
    for tbl in _WIPE_TABLES_FK_SAFE:
        try:
            db.execute(text(f"DELETE FROM {tbl}"))
        except Exception:
            # Tabella inesistente nel DB corrente: skip senza errore
            db.rollback()
    db.commit()


# ============================================================================
# Entry point
# ============================================================================

SCHEMA_FILE = "schema.xlsx"
METADATA_FILE = "languages_metadata.xlsx"
GLOSSARY_FILE = "glossary.xlsx"
LANG_DIR = "languages/"
EXTRAS_DIR = "extras/"


def restore_backup_bundle(
    db: Session,
    zip_bytes: bytes,
    current_user_id: int,
    *,
    wipe: bool = False,
    progress: ProgressReporter = NULL_PROGRESS,
) -> BackupRestoreReport:
    report = BackupRestoreReport()

    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes), "r")
    except zipfile.BadZipFile as e:
        report.errors.append({"_file": "(bundle)", "reason": f"Not a valid ZIP file: {e}"})
        return report

    namelist = zf.namelist()

    if wipe:
        progress.phase("wipe", "Wiping data tables…", total=1)
        try:
            _wipe_data(db)
            progress.tick(1)
        except Exception as e:
            report.errors.append({"_file": "(wipe)", "reason": f"Wipe failed: {e}"})
            return report

    # 1. Schema (richiesto)
    if SCHEMA_FILE in namelist:
        progress.phase("schema", "Importing schema…", total=1)
        with zf.open(SCHEMA_FILE) as f:
            sub = import_excel(db, f.read(), current_user_id, create_missing=True)
        _absorb_sub_report(report, SCHEMA_FILE, sub)
        progress.tick(1)
    else:
        report.errors.append({"_file": SCHEMA_FILE, "reason": "Missing schema.xlsx in bundle"})

    # 2. Languages metadata
    if METADATA_FILE in namelist:
        progress.phase("metadata", "Importing languages metadata…", total=1)
        with zf.open(METADATA_FILE) as f:
            sub = import_excel(db, f.read(), current_user_id, create_missing=True)
        _absorb_sub_report(report, METADATA_FILE, sub)
        progress.tick(1)
    else:
        report.files_skipped.append(METADATA_FILE)

    # 3. Glossary
    if GLOSSARY_FILE in namelist:
        progress.phase("glossary", "Importing glossary…", total=1)
        with zf.open(GLOSSARY_FILE) as f:
            sub = import_excel(db, f.read(), current_user_id, create_missing=True)
        _absorb_sub_report(report, GLOSSARY_FILE, sub)
        progress.tick(1)
    else:
        report.files_skipped.append(GLOSSARY_FILE)

    # 4. Per-language compilation
    lang_files = sorted(
        n for n in namelist
        if n.startswith(LANG_DIR) and n.endswith(".xlsx") and not n.endswith("/")
    )
    total = len(lang_files)
    progress.phase("compilation", f"Restoring {total} language(s)…", total=total)
    for i, name in enumerate(lang_files, start=1):
        lang_id = os.path.splitext(os.path.basename(name))[0]
        progress.tick(current=i, label=f"Restoring {lang_id} ({i}/{total})")
        try:
            with zf.open(name) as f:
                sub = import_excel(db, f.read(), current_user_id, create_missing=True)
        except Exception as e:
            report.errors.append({"_file": name, "reason": f"Cannot read entry: {e}"})
            report.languages_failed.append(lang_id)
            continue
        _absorb_sub_report(report, name, sub)
        if sub.errors:
            report.languages_failed.append(lang_id)
        else:
            report.languages_restored.append(lang_id)

    # 5. Extras (bundle full): site_content sempre upsertato, le tabelle
    # snapshot (submissions/parameter_submissions/archived_questions) solo se
    # wipe=True per evitare duplicati su PK auto-increment.
    extras_files = [
        n for n in namelist
        if n.startswith(EXTRAS_DIR) and (n.endswith(".xlsx") or n.endswith(".jsonl"))
    ]
    if extras_files:
        progress.phase("extras", f"Restoring {len(extras_files)} extra file(s)…", total=len(extras_files))
        for i, name in enumerate(extras_files, start=1):
            base = os.path.basename(name)
            progress.tick(current=i, label=f"Restoring {base} ({i}/{len(extras_files)})")
            try:
                with zf.open(name) as f:
                    data = f.read()
                handler = _EXTRAS_HANDLERS.get(base)
                if handler is None:
                    report.files_skipped.append(name)
                    continue
                handler(db, data, name, report, wipe=wipe, current_user_id=current_user_id)
            except Exception as e:
                report.errors.append({"_file": name, "reason": f"Cannot restore extras: {e}"})
                db.rollback()

    # 5.bis PDF dell'archivio legale (bundle full): rimaterializzati su
    # filesystem se assenti. MAI sovrascritti: l'archivio e' immutabile e
    # un file gia' presente e' per definizione quello giusto (sha256 in DB
    # permette comunque una verifica manuale in caso di dubbio).
    pdf_entries = [
        n for n in namelist
        if n.startswith("extras/legal_pdfs/") and not n.endswith("/")
    ]
    if pdf_entries:
        restored = already = 0
        try:
            os.makedirs(LEGAL_DOCUMENTS_DIR, exist_ok=True)
        except Exception as e:
            report.errors.append({
                "_file": "extras/legal_pdfs/",
                "reason": f"Cannot create legal documents dir: {e}",
            })
        else:
            for entry in pdf_entries:
                fname = os.path.basename(entry)
                if not fname:
                    continue
                target = os.path.join(LEGAL_DOCUMENTS_DIR, fname)
                if os.path.exists(target):
                    already += 1
                    continue
                try:
                    with zf.open(entry) as src, open(target, "wb") as dst:
                        dst.write(src.read())
                    restored += 1
                except Exception as e:
                    report.errors.append({"_file": entry, "reason": f"Cannot restore PDF: {e}"})
            report.by_file["extras/legal_pdfs/"] = {
                "restored": restored, "already_present": already,
            }

    # 6. Recompute final values: il bundle contiene risposte/esempi/motivazioni
    # ma NON i value_orig/value_eval calcolati. Senza questo step, dopo un
    # wipe+restore la tabella `language_parameters` resta vuota e TableA /
    # dashboard / debug parametri appaiono "vuote". Eseguiamo il DAG per
    # ciascuna lingua restorata.
    if report.languages_restored:
        n_lang = len(report.languages_restored)
        progress.phase("recompute", f"Recomputing final values for {n_lang} language(s)…", total=n_lang)
        for i, lang_id in enumerate(report.languages_restored, start=1):
            progress.tick(current=i, label=f"Recomputing {lang_id} ({i}/{n_lang})")
            try:
                run_dag_for_language(lang_id, db)
                db.commit()
            except Exception as e:
                db.rollback()
                report.errors.append({"_file": f"recompute/{lang_id}", "reason": f"Recompute failed: {e}"})

    return report


# ============================================================================
# Extras handlers
# ============================================================================
#
# Ogni handler legge un xlsx in formato noto (vedi services/excel_export.py per
# la definizione dei sheet) e ripristina le righe nelle tabelle DB. Il pattern
# generale per le tabelle gerarchiche è: inserisci il parent senza id esplicito
# (così la PK auto-increment non collide con il valore originario), tieni una
# mappa `old_id -> new_id`, poi inserisci i child rimappando l'FK.
# ============================================================================


def _read_sheet_rows(data: bytes, sheet_name: str):
    """Restituisce (headers, rows) o (None, None) se lo sheet non esiste.

    `rows` è un generatore di dict header→value (esclusa l'intestazione).
    Le righe completamente vuote vengono filtrate."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return None, None
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        return [], iter([])

    def _generator():
        for row in rows_iter:
            if not any(c is not None and c != "" for c in row):
                continue
            yield dict(zip(headers, row))

    return headers, _generator()


def _user_id_by_email(db: Session) -> Dict[str, int]:
    return {u.email: u.id for u in db.query(models.User).all() if u.email}


def _yn_to_bool(v) -> bool:
    s = (str(v) if v is not None else "").strip().lower()
    return s in ("yes", "true", "1", "y")


def _restore_site_content(
    db: Session, data: bytes, name: str, report: "BackupRestoreReport",
    *, wipe: bool, current_user_id: int,
) -> None:
    """Upsert per chiave naturale (`key`). Funziona sia con wipe sia senza."""
    headers, rows = _read_sheet_rows(data, "SiteContents")
    if rows is None:
        report.files_skipped.append(name)
        return

    user_id_by_email = _user_id_by_email(db)
    inserted = updated = 0
    for d in rows:
        key = d.get("Key")
        if not key:
            continue
        existing = db.query(models.SiteContent).filter(models.SiteContent.key == key).first()
        updated_by_id = user_id_by_email.get(d.get("Updated By Email")) if d.get("Updated By Email") else None
        if existing:
            existing.content = d.get("Content") or ""
            existing.page = d.get("Page")
            existing.updated_by_id = updated_by_id
            updated += 1
        else:
            db.add(models.SiteContent(
                key=key,
                content=d.get("Content") or "",
                page=d.get("Page"),
                updated_by_id=updated_by_id,
            ))
            inserted += 1
    db.commit()
    report.files_processed.append(name)
    report.by_file[name] = {"inserted": inserted, "updated": updated}


def _restore_submissions(
    db: Session, data: bytes, name: str, report: "BackupRestoreReport",
    *, wipe: bool, current_user_id: int,
) -> None:
    """Snapshot di lingue inviate per approvazione. Restorato SOLO con wipe=True."""
    if not wipe:
        report.files_skipped.append(name)
        report.by_file[name] = {"reason": "skipped (wipe=False)"}
        return

    user_id_by_email = _user_id_by_email(db)
    id_map: Dict[int, int] = {}

    # Master
    _, rows = _read_sheet_rows(data, "Submissions")
    if rows is None:
        report.files_skipped.append(name)
        return
    n_master = 0
    for d in rows:
        old_id = d.get("ID")
        if old_id is None:
            continue
        # Risoluzione del Language ID dal file: prima cerca l'id corrente,
        # poi gli alias storici (lingua rinominata dopo l'export). Se non
        # trovato, segnala l'errore e salta la submission.
        file_lang_id = d.get("Language ID") or ""
        resolved = resolve_language(db, file_lang_id)
        if resolved.language is None:
            report.errors.append({
                "_file": name, "_row_id": int(old_id),
                "reason": f"Language '{file_lang_id}' not found (no current id, no historical alias).",
            })
            continue
        s = models.Submission(
            language_id=resolved.language.id,
            submitted_by_id=user_id_by_email.get(d.get("Submitted By Email")) if d.get("Submitted By Email") else None,
            submitted_at=d.get("Submitted At"),
            note=d.get("Note") or "",
        )
        db.add(s)
        db.flush()
        id_map[int(old_id)] = s.id
        n_master += 1

    # Children: SubmissionAnswers
    n_ans = n_ex = n_amot = n_par = 0
    _, rows = _read_sheet_rows(data, "SubmissionAnswers")
    if rows is not None:
        for d in rows:
            new_sid = id_map.get(int(d["Submission ID"])) if d.get("Submission ID") is not None else None
            if new_sid is None:
                continue
            db.add(models.SubmissionAnswer(
                submission_id=new_sid,
                question_code=d.get("Question Code") or "",
                response_text=d.get("Response Text") or None,
                comments=d.get("Comments"),
            ))
            n_ans += 1

    _, rows = _read_sheet_rows(data, "SubmissionExamples")
    if rows is not None:
        for d in rows:
            new_sid = id_map.get(int(d["Submission ID"])) if d.get("Submission ID") is not None else None
            if new_sid is None:
                continue
            db.add(models.SubmissionExample(
                submission_id=new_sid,
                question_code=d.get("Question Code") or "",
                textarea=d.get("Textarea"),
                transliteration=d.get("Transliteration"),
                gloss=d.get("Gloss"),
                translation=d.get("Translation"),
                reference=d.get("Reference"),
                is_test=str(d.get("Is Test") or "").strip().upper() in ("TEST", "YES", "Y", "TRUE", "1", "X"),
            ))
            n_ex += 1

    _, rows = _read_sheet_rows(data, "SubmissionAnswerMotivations")
    if rows is not None:
        for d in rows:
            new_sid = id_map.get(int(d["Submission ID"])) if d.get("Submission ID") is not None else None
            if new_sid is None:
                continue
            db.add(models.SubmissionAnswerMotivation(
                submission_id=new_sid,
                question_code=d.get("Question Code") or "",
                motivation_code=d.get("Motivation Code") or "",
                motivation_label=d.get("Motivation Label"),
            ))
            n_amot += 1

    _, rows = _read_sheet_rows(data, "SubmissionParams")
    if rows is not None:
        for d in rows:
            new_sid = id_map.get(int(d["Submission ID"])) if d.get("Submission ID") is not None else None
            if new_sid is None:
                continue
            db.add(models.SubmissionParam(
                submission_id=new_sid,
                parameter_id=d.get("Parameter ID") or "",
                value_orig=d.get("Value Orig") or None,
                warning_orig=_yn_to_bool(d.get("Warning Orig")),
                value_eval=d.get("Value Eval") or None,
                warning_eval=_yn_to_bool(d.get("Warning Eval")),
                evaluated_at=d.get("Evaluated At"),
            ))
            n_par += 1

    db.commit()
    report.files_processed.append(name)
    report.by_file[name] = {
        "submissions": n_master, "answers": n_ans,
        "examples": n_ex, "answer_motivations": n_amot, "params": n_par,
    }


def _restore_parameter_submissions(
    db: Session, data: bytes, name: str, report: "BackupRestoreReport",
    *, wipe: bool, current_user_id: int,
) -> None:
    """Snapshot delle definizioni di parametri. Restorato SOLO con wipe=True."""
    if not wipe:
        report.files_skipped.append(name)
        report.by_file[name] = {"reason": "skipped (wipe=False)"}
        return

    user_id_by_email = _user_id_by_email(db)
    sub_id_map: Dict[int, int] = {}
    q_id_map: Dict[int, int] = {}

    _, rows = _read_sheet_rows(data, "ParameterSubmissions")
    if rows is None:
        report.files_skipped.append(name)
        return
    n_master = 0
    for d in rows:
        old_id = d.get("ID")
        if old_id is None:
            continue
        ps = models.ParameterSubmission(
            parameter_id=d.get("Parameter ID") or "",
            parameter_name=d.get("Parameter Name") or "",
            submitted_by_id=user_id_by_email.get(d.get("Submitted By Email")) if d.get("Submitted By Email") else None,
            submitted_at=d.get("Submitted At"),
            note=d.get("Note") or "",
            short_description=d.get("Short Description") or "",
            long_description=d.get("Long Description") or "",
            implicational_condition=d.get("Implicational Condition") or None,
            description_of_the_implicational_condition=d.get("Description Of Implicational Condition") or "",
            is_active=_yn_to_bool(d.get("Is Active")),
            position=int(d["Position"]) if d.get("Position") not in (None, "") else None,
            schema=d.get("Schema") or "",
            param_type=d.get("Param Type") or "",
            level_of_comparison=d.get("Level Of Comparison") or "",
        )
        db.add(ps)
        db.flush()
        sub_id_map[int(old_id)] = ps.id
        n_master += 1

    n_q = n_am = 0
    _, rows = _read_sheet_rows(data, "Questions")
    if rows is not None:
        for d in rows:
            old_id = d.get("ID")
            new_sid = sub_id_map.get(int(d["Submission ID"])) if d.get("Submission ID") is not None else None
            if old_id is None or new_sid is None:
                continue
            psq = models.ParameterSubmissionQuestion(
                submission_id=new_sid,
                question_code=d.get("Question Code") or "",
                text=d.get("Text") or "",
                template_type=d.get("Template Type") or "",
                instruction=d.get("Instruction"),
                instruction_yes=d.get("Instruction YES"),
                instruction_no=d.get("Instruction NO"),
                example_yes=d.get("Example YES"),
                help_info=d.get("Help Info"),
                is_stop_question=_yn_to_bool(d.get("Is Stop Question")),
                is_active=_yn_to_bool(d.get("Is Active")),
            )
            db.add(psq)
            db.flush()
            q_id_map[int(old_id)] = psq.id
            n_q += 1

    _, rows = _read_sheet_rows(data, "AllowedMotivations")
    if rows is not None:
        for d in rows:
            new_qid = q_id_map.get(int(d["Question ID"])) if d.get("Question ID") is not None else None
            if new_qid is None:
                continue
            db.add(models.ParameterSubmissionAllowedMotivation(
                question_id=new_qid,
                motivation_code=d.get("Motivation Code") or "",
                motivation_label=d.get("Motivation Label") or "",
            ))
            n_am += 1

    db.commit()
    report.files_processed.append(name)
    report.by_file[name] = {
        "parameter_submissions": n_master, "questions": n_q, "allowed_motivations": n_am,
    }


def _restore_archived_questions(
    db: Session, data: bytes, name: str, report: "BackupRestoreReport",
    *, wipe: bool, current_user_id: int,
) -> None:
    """Archivio di domande dismesse + answer/example/motivation collegate.
    Restorato SOLO con wipe=True (PK auto-increment, niente chiave naturale)."""
    if not wipe:
        report.files_skipped.append(name)
        report.by_file[name] = {"reason": "skipped (wipe=False)"}
        return

    user_id_by_email = _user_id_by_email(db)
    aq_map: Dict[int, int] = {}
    aa_map: Dict[int, int] = {}

    _, rows = _read_sheet_rows(data, "ArchivedQuestions")
    if rows is None:
        report.files_skipped.append(name)
        return
    n_q = 0
    for d in rows:
        old_id = d.get("ID")
        if old_id is None:
            continue
        aq = models.ArchivedQuestion(
            original_question_id=d.get("Original Question ID") or "",
            parameter_id=d.get("Parameter ID") or "",
            parameter_name=d.get("Parameter Name") or "",
            text=d.get("Text") or "",
            template_type=d.get("Template Type") or "",
            instruction=d.get("Instruction"),
            instruction_yes=d.get("Instruction YES"),
            instruction_no=d.get("Instruction NO"),
            example_yes=d.get("Example YES"),
            help_info=d.get("Help Info"),
            is_stop_question=_yn_to_bool(d.get("Is Stop Question")),
            is_active=_yn_to_bool(d.get("Is Active")),
            archived_at=d.get("Archived At"),
            archived_by_id=user_id_by_email.get(d.get("Archived By Email")) if d.get("Archived By Email") else None,
            archive_note=d.get("Archive Note") or "",
            answers_count=int(d.get("Answers Count") or 0),
            examples_count=int(d.get("Examples Count") or 0),
        )
        db.add(aq)
        db.flush()
        aq_map[int(old_id)] = aq.id
        n_q += 1

    n_qm = n_a = n_ex = n_am = 0

    _, rows = _read_sheet_rows(data, "ArchivedQuestionMotivations")
    if rows is not None:
        for d in rows:
            new_aqid = aq_map.get(int(d["Archived Question ID"])) if d.get("Archived Question ID") is not None else None
            if new_aqid is None:
                continue
            db.add(models.ArchivedQuestionMotivation(
                archived_question_id=new_aqid,
                motivation_code=d.get("Motivation Code") or "",
                motivation_label=d.get("Motivation Label") or "",
            ))
            n_qm += 1

    _, rows = _read_sheet_rows(data, "ArchivedAnswers")
    if rows is not None:
        for d in rows:
            old_id = d.get("ID")
            new_aqid = aq_map.get(int(d["Archived Question ID"])) if d.get("Archived Question ID") is not None else None
            if old_id is None or new_aqid is None:
                continue
            aa = models.ArchivedAnswer(
                archived_question_id=new_aqid,
                language_id=d.get("Language ID") or "",
                language_name_full=d.get("Language Name Full") or "",
                status=d.get("Status") or None,
                response_text=d.get("Response Text") or None,
                comments=d.get("Comments"),
                original_updated_at=d.get("Original Updated At"),
            )
            db.add(aa)
            db.flush()
            aa_map[int(old_id)] = aa.id
            n_a += 1

    _, rows = _read_sheet_rows(data, "ArchivedExamples")
    if rows is not None:
        for d in rows:
            new_aaid = aa_map.get(int(d["Archived Answer ID"])) if d.get("Archived Answer ID") is not None else None
            if new_aaid is None:
                continue
            db.add(models.ArchivedExample(
                archived_answer_id=new_aaid,
                number=d.get("Number") or "",
                textarea=d.get("Textarea"),
                transliteration=d.get("Transliteration"),
                gloss=d.get("Gloss"),
                translation=d.get("Translation"),
                reference=d.get("Reference"),
            ))
            n_ex += 1

    _, rows = _read_sheet_rows(data, "ArchivedAnswerMotivations")
    if rows is not None:
        for d in rows:
            new_aaid = aa_map.get(int(d["Archived Answer ID"])) if d.get("Archived Answer ID") is not None else None
            if new_aaid is None:
                continue
            db.add(models.ArchivedAnswerMotivation(
                archived_answer_id=new_aaid,
                motivation_code=d.get("Motivation Code") or "",
                motivation_label=d.get("Motivation Label") or "",
            ))
            n_am += 1

    db.commit()
    report.files_processed.append(name)
    report.by_file[name] = {
        "archived_questions": n_q, "question_motivations": n_qm,
        "archived_answers": n_a, "examples": n_ex, "answer_motivations": n_am,
    }


def _restore_parameter_change_logs(
    db: Session, data: bytes, name: str, report: "BackupRestoreReport",
    *, wipe: bool, current_user_id: int,
) -> None:
    """Log 'ultima modifica' dei parametri. Restorato SOLO con wipe=True
    (PK auto-increment, nessuna chiave naturale). `user_id` e' NOT NULL:
    se l'email salvata non esiste nel DB target il log viene attribuito
    all'admin che sta eseguendo il restore (conteggiato nel report)."""
    if not wipe:
        report.files_skipped.append(name)
        report.by_file[name] = {"reason": "skipped (wipe=False)"}
        return

    _, rows = _read_sheet_rows(data, "ParameterChangeLogs")
    if rows is None:
        report.files_skipped.append(name)
        return

    user_id_by_email = _user_id_by_email(db)
    inserted = skipped = user_fallbacks = 0
    for d in rows:
        pid = d.get("Parameter ID") or ""
        resolved = resolve_parameter(db, pid)
        if resolved.parameter is None:
            report.errors.append({
                "_file": name,
                "reason": f"Parameter '{pid}' not found (no current id, no historical alias).",
            })
            skipped += 1
            continue
        uid = user_id_by_email.get(d.get("User Email")) if d.get("User Email") else None
        if uid is None:
            uid = current_user_id
            user_fallbacks += 1
        db.add(models.ParameterChangeLog(
            parameter_id=resolved.parameter.id,
            user_id=uid,
            change_note=d.get("Change Note") or "",
            created_at=d.get("Created At"),
        ))
        inserted += 1
    db.commit()
    report.files_processed.append(name)
    report.by_file[name] = {
        "inserted": inserted, "skipped_rows": skipped, "user_fallbacks": user_fallbacks,
    }


def _restore_parameter_flags(
    db: Session, data: bytes, name: str, report: "BackupRestoreReport",
    *, wipe: bool, current_user_id: int,
) -> None:
    """Flag per (lingua, parametro): is_unsure e needs_review. Upsert per
    chiave naturale (language_id, parameter_id): funziona sia con wipe sia
    senza. Gira DOPO la fase languages/, quindi le righe di status con le
    admin note esistono gia': qui vengono toccati SOLO i due flag."""
    _, rows = _read_sheet_rows(data, "ParameterFlags")
    if rows is None:
        report.files_skipped.append(name)
        return

    updated = created = skipped = 0
    for d in rows:
        lid_raw = d.get("Language ID") or ""
        pid_raw = d.get("Parameter ID") or ""
        lang = resolve_language(db, lid_raw).language
        param = resolve_parameter(db, pid_raw).parameter
        if lang is None or param is None:
            miss = f"language '{lid_raw}'" if lang is None else f"parameter '{pid_raw}'"
            report.errors.append({"_file": name, "reason": f"Cannot restore flags: {miss} not found."})
            skipped += 1
            continue
        is_unsure = _yn_to_bool(d.get("Is Unsure"))
        needs_review = _yn_to_bool(d.get("Needs Review"))
        existing = db.query(models.LanguageParameterStatus).filter(
            models.LanguageParameterStatus.language_id == lang.id,
            models.LanguageParameterStatus.parameter_id == param.id,
        ).first()
        if existing:
            existing.is_unsure = is_unsure
            existing.needs_review = needs_review
            updated += 1
        else:
            db.add(models.LanguageParameterStatus(
                language_id=lang.id, parameter_id=param.id,
                is_unsure=is_unsure, needs_review=needs_review,
            ))
            created += 1
    db.commit()
    report.files_processed.append(name)
    report.by_file[name] = {"updated": updated, "created": created, "skipped_rows": skipped}


def _restore_aliases(
    db: Session, data: bytes, name: str, report: "BackupRestoreReport",
    *, wipe: bool, current_user_id: int,
) -> None:
    """Alias storici di lingue/parametri/question. Upsert per chiave naturale
    `old_id` (UNIQUE): funziona con e senza wipe. Righe il cui old_id coincide
    con un id corrente vengono saltate: l'alias sarebbe morto (i resolver
    trovano prima l'id corrente) e ambiguo."""
    specs = [
        ("LanguageAliases", "Language ID", models.LanguageAlias, "language_id",
         models.Language, lambda v: resolve_language(db, v).language),
        ("ParameterAliases", "Parameter ID", models.ParameterAlias, "parameter_id",
         models.ParameterDef, lambda v: resolve_parameter(db, v).parameter),
        ("QuestionAliases", "Question ID", models.QuestionAlias, "question_id",
         models.Question, lambda v: resolve_question(db, v).question),
    ]

    any_sheet = False
    summary: Dict[str, Any] = {}
    for sheet, id_header, alias_model, fk_attr, entity_model, resolve in specs:
        _, rows = _read_sheet_rows(data, sheet)
        if rows is None:
            continue
        any_sheet = True
        inserted = updated = skipped = 0
        for d in rows:
            old_id = (str(d.get("Old ID")) if d.get("Old ID") is not None else "").strip()
            target_raw = (str(d.get(id_header)) if d.get(id_header) is not None else "").strip()
            if not old_id or not target_raw:
                continue
            entity = resolve(target_raw)
            if entity is None:
                report.errors.append({
                    "_file": name,
                    "reason": f"{sheet}: target '{target_raw}' for alias '{old_id}' not found.",
                })
                skipped += 1
                continue
            # old_id che coincide con un id corrente: alias morto, skip.
            if db.get(entity_model, old_id) is not None:
                skipped += 1
                continue
            existing = db.query(alias_model).filter(alias_model.old_id == old_id).first()
            if existing:
                setattr(existing, fk_attr, entity.id)
                if d.get("Created At"):
                    existing.created_at = d.get("Created At")
                updated += 1
            else:
                kwargs = {fk_attr: entity.id, "old_id": old_id}
                if d.get("Created At"):
                    kwargs["created_at"] = d.get("Created At")
                db.add(alias_model(**kwargs))
                inserted += 1
        summary[sheet] = {"inserted": inserted, "updated": updated, "skipped_rows": skipped}

    if not any_sheet:
        report.files_skipped.append(name)
        return
    db.commit()
    report.files_processed.append(name)
    report.by_file[name] = summary


def _restore_users(
    db: Session, data: bytes, name: str, report: "BackupRestoreReport",
    *, wipe: bool, current_user_id: int,
) -> None:
    """Utenti. Upsert per email (chiave naturale UNIQUE). Il bundle NON
    contiene gli hash password (scelta deliberata): i nuovi utenti nascono
    con una password random inutilizzabile e rientrano via "password
    dimenticata"; agli utenti esistenti la password NON viene mai toccata.
    Safety: l'utente che sta eseguendo il restore non puo' essere
    disattivato o degradato dal bundle. Ripristina anche l'assegnazione
    lingua->utente (colonna Assigned Languages), che l'import dei metadati
    lingue lascia deliberatamente da parte."""
    _, rows = _read_sheet_rows(data, "Users")
    if rows is None:
        report.files_skipped.append(name)
        return

    import secrets
    from auth import get_password_hash

    valid_roles = {"admin", "user", "public"}
    inserted = updated = assigned = skipped_assign = 0
    for d in rows:
        email = (str(d.get("Email")) if d.get("Email") is not None else "").strip()
        if not email:
            continue
        role = (str(d.get("Role")) if d.get("Role") is not None else "").strip() or "public"
        if role not in valid_roles:
            role = "public"
        is_active = _yn_to_bool(d.get("Is Active"))
        existing = db.query(models.User).filter(models.User.email == email).first()
        if existing:
            existing.name = d.get("Name") or None
            existing.surname = d.get("Surname") or None
            if existing.id != current_user_id:
                existing.role = role
                existing.is_active = is_active
            existing.terms_accepted = _yn_to_bool(d.get("Terms Accepted"))
            if d.get("Terms Accepted At"):
                existing.terms_accepted_at = d.get("Terms Accepted At")
            if d.get("Date Joined"):
                existing.date_joined = d.get("Date Joined")
            user = existing
            updated += 1
        else:
            user = models.User(
                email=email,
                hashed_password=get_password_hash(secrets.token_urlsafe(32)),
                name=d.get("Name") or None,
                surname=d.get("Surname") or None,
                role=role,
                is_active=is_active,
                terms_accepted=_yn_to_bool(d.get("Terms Accepted")),
                terms_accepted_at=d.get("Terms Accepted At"),
                date_joined=d.get("Date Joined"),
            )
            db.add(user)
            db.flush()
            inserted += 1

        raw = str(d.get("Assigned Languages") or "")
        for lid in [x.strip() for x in raw.split(",") if x.strip()]:
            resolved = resolve_language(db, lid)
            if resolved.language is None:
                report.errors.append({
                    "_file": name,
                    "reason": f"Assigned language '{lid}' for user '{email}' not found.",
                })
                skipped_assign += 1
                continue
            resolved.language.assigned_user_id = user.id
            assigned += 1

    db.commit()
    report.files_processed.append(name)
    report.by_file[name] = {
        "inserted": inserted, "updated": updated,
        "languages_assigned": assigned, "assignments_skipped": skipped_assign,
    }


def _restore_legal_documents(
    db: Session, data: bytes, name: str, report: "BackupRestoreReport",
    *, wipe: bool, current_user_id: int,
) -> None:
    """Archivio documenti legali + consensi. Upsert per chiavi naturali:
    (type, version) per i documenti, (user, documento) per i consensi.
    Funziona con e senza wipe — queste tabelle non sono MAI wipate (valore
    di prova legale). Dopo l'upsert, is_current viene rinormalizzato: per
    ciascun type e' current solo l'ultima versione pubblicata (stessa
    semantica del flusso di publish), cosi' un bundle vecchio non puo'
    retrocedere il documento corrente."""
    _, rows = _read_sheet_rows(data, "LegalDocuments")
    if rows is None:
        report.files_skipped.append(name)
        return

    docs_ins = docs_upd = 0
    for d in rows:
        dtype = (str(d.get("Type")) if d.get("Type") is not None else "").strip()
        version = (str(d.get("Version")) if d.get("Version") is not None else "").strip()
        if not dtype or not version:
            continue
        vex = None
        if d.get("Vexatious Clauses"):
            try:
                vex = json.loads(d["Vexatious Clauses"])
            except (ValueError, TypeError):
                vex = None
        existing = db.query(models.LegalDocument).filter(
            models.LegalDocument.type == dtype,
            models.LegalDocument.version == version,
        ).first()
        if existing:
            # L'archivio e' immutabile: file_path/sha256 non dovrebbero mai
            # cambiare per una (type, version) data. Il bundle e' la fonte
            # di verita' per nota e clausole.
            existing.vexatious_clauses = vex
            existing.note = d.get("Note") or None
            docs_upd += 1
        else:
            db.add(models.LegalDocument(
                type=dtype,
                version=version,
                file_path=os.path.basename(str(d.get("File Path") or "")),
                sha256=str(d.get("SHA256") or ""),
                published_at=d.get("Published At"),
                is_current=_yn_to_bool(d.get("Is Current")),
                vexatious_clauses=vex,
                note=d.get("Note") or None,
            ))
            docs_ins += 1
    db.flush()

    # Rinormalizza is_current: per type, current = ultima published_at.
    doc_types = [t[0] for t in db.query(models.LegalDocument.type).distinct().all()]
    for dtype in doc_types:
        docs = (
            db.query(models.LegalDocument)
            .filter(models.LegalDocument.type == dtype)
            .order_by(models.LegalDocument.published_at.desc(),
                      models.LegalDocument.id.desc())
            .all()
        )
        for i, doc in enumerate(docs):
            doc.is_current = (i == 0)

    # Consents: upsert per (user, documento). Righe senza utente risolvibile
    # vengono saltate con errore nel report (un consenso senza identita' non
    # e' upsertabile); righe gia' anonimizzate (email vuota per GDPR delete)
    # vengono dedupate su (documento, accepted_at, method).
    cons_ins = cons_upd = cons_skip = 0
    _, rows = _read_sheet_rows(data, "Consents")
    if rows is not None:
        user_id_by_email = _user_id_by_email(db)
        doc_by_key = {
            (x.type, x.version): x for x in db.query(models.LegalDocument).all()
        }
        for d in rows:
            key = (
                (str(d.get("Document Type")) if d.get("Document Type") is not None else "").strip(),
                (str(d.get("Document Version")) if d.get("Document Version") is not None else "").strip(),
            )
            doc = doc_by_key.get(key)
            if doc is None:
                report.errors.append({
                    "_file": name,
                    "reason": f"Consent references unknown document {key}; row skipped.",
                })
                cons_skip += 1
                continue
            email = (str(d.get("User Email")) if d.get("User Email") is not None else "").strip()
            uid = user_id_by_email.get(email) if email else None
            if email and uid is None:
                report.errors.append({
                    "_file": name,
                    "reason": f"Consent user '{email}' not found in target DB; row skipped.",
                })
                cons_skip += 1
                continue
            fields = {
                "ip_address": d.get("IP Address") or None,
                "user_agent": d.get("User Agent") or None,
                "method": str(d.get("Method") or ""),
                "vexatious_clauses_approved": _yn_to_bool(d.get("Vexatious Clauses Approved")),
                "revoked_at": d.get("Revoked At"),
                "revocation_reason": d.get("Revocation Reason") or None,
            }
            if d.get("Accepted At"):
                fields["accepted_at"] = d.get("Accepted At")
            if uid is not None:
                existing = db.query(models.Consent).filter(
                    models.Consent.user_id == uid,
                    models.Consent.legal_document_id == doc.id,
                ).first()
            else:
                existing = db.query(models.Consent).filter(
                    models.Consent.user_id.is_(None),
                    models.Consent.legal_document_id == doc.id,
                    models.Consent.accepted_at == d.get("Accepted At"),
                    models.Consent.method == fields["method"],
                ).first()
            if existing:
                for k, v in fields.items():
                    setattr(existing, k, v)
                cons_upd += 1
            else:
                db.add(models.Consent(user_id=uid, legal_document_id=doc.id, **fields))
                cons_ins += 1

    db.commit()
    report.files_processed.append(name)
    report.by_file[name] = {
        "documents_inserted": docs_ins, "documents_updated": docs_upd,
        "consents_inserted": cons_ins, "consents_updated": cons_upd,
        "consents_skipped": cons_skip,
    }


def _restore_entity_versions(
    db: Session, data: bytes, name: str, report: "BackupRestoreReport",
    *, wipe: bool, current_user_id: int,
) -> None:
    """History (pagina History). File JSON Lines, non xlsx: gli snapshot
    possono superare il limite di 32.767 caratteri di una cella Excel.

    Insert-if-missing con chiave di dedupe (entity_type, entity_id,
    operation, created_at): idempotente, funziona con e senza wipe (la
    tabella non viene mai wipata) e un secondo restore non duplica nulla."""
    from datetime import datetime as _dt

    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError as e:
        report.errors.append({"_file": name, "reason": f"Not valid UTF-8: {e}"})
        report.files_skipped.append(name)
        return

    user_id_by_email = _user_id_by_email(db)
    existing_keys = {
        (v.entity_type, v.entity_id, v.operation, v.created_at)
        for v in db.query(
            models.EntityVersion.entity_type, models.EntityVersion.entity_id,
            models.EntityVersion.operation, models.EntityVersion.created_at,
        ).all()
    }

    inserted = already = bad = 0
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            rec = json.loads(line)
        except ValueError as e:
            bad += 1
            report.errors.append({"_file": name, "reason": f"Bad JSONL line: {e}"})
            continue
        created_at = None
        if rec.get("created_at"):
            try:
                created_at = _dt.fromisoformat(rec["created_at"])
            except (ValueError, TypeError):
                created_at = None
        key = (
            rec.get("entity_type") or "",
            str(rec.get("entity_id") or ""),
            rec.get("operation") or "update",
            created_at,
        )
        if key in existing_keys:
            already += 1
            continue
        existing_keys.add(key)
        db.add(models.EntityVersion(
            entity_type=key[0],
            entity_id=key[1],
            snapshot=rec.get("snapshot") if rec.get("snapshot") is not None else {},
            operation=key[2],
            source=rec.get("source") or "manual",
            note=rec.get("note"),
            user_id=user_id_by_email.get(rec.get("user_email")) if rec.get("user_email") else None,
            created_at=created_at,
        ))
        inserted += 1
    db.commit()
    report.files_processed.append(name)
    report.by_file[name] = {
        "inserted": inserted, "already_present": already, "bad_lines": bad,
    }


_EXTRAS_HANDLERS = {
    "aliases.xlsx": _restore_aliases,
    "users.xlsx": _restore_users,
    "site_content.xlsx": _restore_site_content,
    "submissions.xlsx": _restore_submissions,
    "parameter_submissions.xlsx": _restore_parameter_submissions,
    "archived_questions.xlsx": _restore_archived_questions,
    "parameter_change_logs.xlsx": _restore_parameter_change_logs,
    "parameter_flags.xlsx": _restore_parameter_flags,
    "legal_documents.xlsx": _restore_legal_documents,
    "entity_versions.jsonl": _restore_entity_versions,
}
