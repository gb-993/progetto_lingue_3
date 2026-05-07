from sqlalchemy.orm import Session, joinedload
from datetime import datetime
from time_utils import utc_now
import io

from openpyxl import Workbook

import models
from services.backup_service import (
    _bold_header_row,
    _style_table,
    workbook_to_bytes,
)
from services.citation import apply_excel_citation

# Massimo numero di snapshot mantenuti per ogni parametro
MAX_PER_PARAMETER = 10


def create_parameter_submission(
    db: Session,
    parameter: models.ParameterDef,
    user_id: int,
    note: str = "",
    fixed_time: datetime = None,
):
    """Snapshot 'full' della *definizione* di un parametro.

    Salva: ParameterDef + tutte le Question + motivations ammesse per ciascuna
    Question. Niente dati delle lingue (quelli vivono nei Submission).
    """
    now = fixed_time or utc_now()

    sub = models.ParameterSubmission(
        parameter_id=parameter.id,
        parameter_name=parameter.name or "",
        submitted_by_id=user_id,
        submitted_at=now,
        note=note or "",
        short_description=parameter.short_description or "",
        long_description=parameter.long_description or "",
        implicational_condition=parameter.implicational_condition,
        description_of_the_implicational_condition=
            parameter.description_of_the_implicational_condition or "",
        is_active=parameter.is_active,
        position=parameter.position,
        schema=parameter.schema or "",
        param_type=parameter.param_type or "",
        level_of_comparison=parameter.level_of_comparison or "",
    )
    db.add(sub)
    db.flush()

    questions = (
        db.query(models.Question)
        .options(
            joinedload(models.Question.allowed_motivations)
            .joinedload(models.QuestionAllowedMotivation.motivation)
        )
        .filter(models.Question.parameter_id == parameter.id)
        .all()
    )

    for q in questions:
        sub_q = models.ParameterSubmissionQuestion(
            submission_id=sub.id,
            question_code=q.id,
            text=q.text or "",
            template_type=q.template_type or "",
            instruction=q.instruction,
            instruction_yes=q.instruction_yes,
            instruction_no=q.instruction_no,
            example_yes=q.example_yes,
            help_info=q.help_info,
            is_stop_question=q.is_stop_question,
            is_active=q.is_active,
        )
        db.add(sub_q)
        db.flush()

        for am in q.allowed_motivations:
            mot = am.motivation
            if mot is None:
                continue
            db.add(models.ParameterSubmissionAllowedMotivation(
                question_id=sub_q.id,
                motivation_code=mot.code,
                motivation_label=mot.label or "",
            ))

    db.flush()

    # Pruning: tieni i più recenti N
    subs = (
        db.query(models.ParameterSubmission.id)
        .filter(models.ParameterSubmission.parameter_id == parameter.id)
        .order_by(
            models.ParameterSubmission.submitted_at.desc(),
            models.ParameterSubmission.id.desc(),
        )
        .all()
    )
    pruned_count = 0
    if len(subs) > MAX_PER_PARAMETER:
        ids_to_keep = [s[0] for s in subs[:MAX_PER_PARAMETER]]
        pruned_count = (
            db.query(models.ParameterSubmission)
            .filter(
                models.ParameterSubmission.parameter_id == parameter.id,
                models.ParameterSubmission.id.notin_(ids_to_keep),
            )
            .delete(synchronize_session=False)
        )

    return sub, pruned_count


def create_all_parameters_backup(
    db: Session, user_id: int, note: str = "Global parameters backup"
):
    """Backup globale: uno snapshot per ogni parametro, accomunati dallo stesso
    timestamp (microsecondi azzerati) per formare la stessa "cartella".
    """
    parameters = db.query(models.ParameterDef).all()
    fixed_time = utc_now().replace(microsecond=0)
    total_pruned = 0

    try:
        for p in parameters:
            _, pruned = create_parameter_submission(db, p, user_id, note, fixed_time)
            total_pruned += pruned

        db.commit()

        return {
            "status": "success",
            "parameters_backed_up": len(parameters),
            "pruned": total_pruned,
            "timestamp": fixed_time,
        }
    except Exception:
        db.rollback()
        raise


def create_single_parameter_backup(
    db: Session, parameter: models.ParameterDef, user_id: int, note: str = ""
):
    """Backup di un singolo parametro: cartella dedicata col proprio timestamp."""
    fixed_time = utc_now().replace(microsecond=0)
    try:
        sub, pruned = create_parameter_submission(db, parameter, user_id, note, fixed_time)
        db.commit()
        return {
            "status": "success",
            "submission_id": sub.id,
            "parameter_id": parameter.id,
            "pruned": pruned,
            "timestamp": fixed_time,
        }
    except Exception:
        db.rollback()
        raise


# ============================================================================
# Export di una ParameterSubmission in xlsx (download del backup parametro).
# Tutto deriva dallo snapshot in DB: niente lookup vivi.
# ============================================================================

_PINFO_HEADERS = ["Field", "Value"]
_PQUESTIONS_HEADERS = [
    "Question ID", "Text", "Template type", "Instruction",
    "Instruction YES", "Instruction NO", "Example YES", "Help info",
    "Is stop question", "Is active",
]
_PQAM_HEADERS = ["Question ID", "Allowed motivation"]


def build_parameter_submission_workbook(db: Session, sub: models.ParameterSubmission) -> Workbook:
    """Workbook per una singola ParameterSubmission (backup di un parametro).

    Sheet:
      - Info       : id, name, descrizioni, formula, is_active, position
      - Questions  : 1 riga per question dello snapshot
      - AllowedMot : 1 riga per (question, motivation_label) consentita
    """
    wb = Workbook()

    # === Info ===
    ws_info = wb.active
    ws_info.title = "Info"
    ws_info.append(_PINFO_HEADERS)
    _bold_header_row(ws_info, len(_PINFO_HEADERS))

    submitter = (
        f"{sub.submitted_by.name or ''} {sub.submitted_by.surname or ''}".strip()
        or (sub.submitted_by.email if sub.submitted_by else "")
    ) if sub.submitted_by_id else "System"
    submitted_at_str = sub.submitted_at.strftime("%Y-%m-%d %H:%M UTC") if sub.submitted_at else ""

    rows = [
        ("Parameter ID", sub.parameter_id or ""),
        ("Parameter name", sub.parameter_name or ""),
        ("Schema", sub.schema or ""),
        ("Type", sub.param_type or ""),
        ("Level of comparison", sub.level_of_comparison or ""),
        ("Position", sub.position if sub.position is not None else ""),
        ("Is active", "Yes" if sub.is_active else "No"),
        ("Short description", sub.short_description or ""),
        ("Long description", sub.long_description or ""),
        ("Implicational condition", sub.implicational_condition or ""),
        ("Explanation of the condition", sub.description_of_the_implicational_condition or ""),
        ("Backup date (UTC)", submitted_at_str),
        ("Submitted by", submitter),
        ("Note", sub.note or ""),
    ]
    for k, v in rows:
        ws_info.append([k, v])
    _style_table(ws_info, "ParamBackupInfo", len(_PINFO_HEADERS), [28, 70])

    # === Questions ===
    ws_q = wb.create_sheet("Questions")
    ws_q.append(_PQUESTIONS_HEADERS)
    _bold_header_row(ws_q, len(_PQUESTIONS_HEADERS))
    qs_sorted = sorted(sub.questions, key=lambda q: (q.is_stop_question, q.question_code or ""))
    for q in qs_sorted:
        ws_q.append([
            q.question_code or "",
            q.text or "",
            q.template_type or "",
            q.instruction or "",
            q.instruction_yes or "",
            q.instruction_no or "",
            q.example_yes or "",
            q.help_info or "",
            "Yes" if q.is_stop_question else "No",
            "Yes" if q.is_active else "No",
        ])
    _style_table(ws_q, "ParamBackupQuestions", len(_PQUESTIONS_HEADERS),
                 [16, 36, 16, 24, 24, 24, 24, 22, 12, 10])

    # === Allowed motivations ===
    ws_qam = wb.create_sheet("AllowedMotivations")
    ws_qam.append(_PQAM_HEADERS)
    _bold_header_row(ws_qam, len(_PQAM_HEADERS))
    for q in qs_sorted:
        for m in q.allowed_motivations:
            text = m.motivation_label or m.motivation_code or ""
            if text:
                ws_qam.append([q.question_code or "", text])
    _style_table(ws_qam, "ParamBackupAllowedMot", len(_PQAM_HEADERS), [18, 50])

    apply_excel_citation(wb)
    return wb
