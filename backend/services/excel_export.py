"""
Servizio di export Excel.

Tre workbook builder:
  - build_language_workbook(db, lang, is_admin): per admin produce 7 sheet
    (3 vecchi formato + 4 schema), per user normale solo "Examples".
  - build_language_list_workbook(db, languages): metadati delle lingue selezionate.
  - build_schema_workbook(db): solo i 4 sheet schema (parametri/domande/motivazioni).

I 3 sheet "vecchi" (Database_model, Examples, Answers) hanno header e ordine
colonne IDENTICI al vecchio progetto Django (vedi test_excel_export.py).
Questo garantisce che i file siano scambiabili: download + re-upload = round-trip.
"""
from __future__ import annotations
from typing import Iterable, List, Optional
from datetime import datetime
import io
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import Session

import models
from services.citation import apply_excel_citation


# ============================================================================
# Header costanti (identici al vecchio progetto per i 3 sheet di compatibilità)
# ============================================================================

DATABASE_MODEL_HEADERS = [
    # Identificazione lingua/domanda. Le colonne testuali Question /
    # Question_Examples_YES / Question_Intructions_Comments del vecchio progetto
    # sono state rimosse: ridondanti rispetto al foglio schema globale (deducibili
    # da Question_ID), non lette in import, appesantivano il file inutilmente.
    "Language",
    "Parameter_Label",
    "Question_ID",
    "Language_Answer",
    "Language_Comments",
    "Language_Examples",
    "Language_Example_Gloss",
    "Language_Example_Translation",
    "Language_References",
    # Aggiunte per backup lossless: senza queste, round-trip export → import
    # perdeva motivazioni e admin notes. Codici motivazione separati da "; ".
    # Admin_Note è duplicato su tutte le righe dello stesso parametro perché
    # vive a livello di (lingua, parametro), non di (lingua, question).
    "Motivations",
    "Admin_Note",
]

EXAMPLES_HEADERS = [
    "Language ID",
    "Question ID",
    "Example #",
    "Example text",
    "Transliteration",
    "Gloss",
    "English translation",
    "Reference",
]

ANSWERS_HEADERS = [
    "Language ID",
    "Parameter Label",
    "Question ID",
    "Question",
    "Question status",
    "Answer",
    "Parameter value",
    "Motivation",
    "Comments",
]

# Schema sheets (nuovi, non presenti nel vecchio progetto)
MOTIVATIONS_HEADERS = ["ID", "Code", "Label"]

PARAMETERS_HEADERS = [
    "ID", "Position", "Name", "Schema", "Type", "Level",
    "Short Description", "Long Description",
    "Implicational Condition", "Explanation of Implicational Condition",
    "Is Active",
]

QUESTIONS_HEADERS = [
    "ID", "Parameter ID", "Text", "Template Type",
    "Instruction", "Instruction YES", "Instruction NO",
    "Example YES", "Help Info",
    "Is Stop Question", "Is Active",
]

QUESTION_ALLOWED_MOTIVATIONS_HEADERS = ["Question ID", "Motivation Code"]

# Admin-only sheet: nota libera per (lingua, parametro)
ADMIN_NOTES_HEADERS = ["Parameter ID", "Parameter Name", "Admin Note"]

# Header per la lista lingue
LANGUAGE_LIST_HEADERS = [
    "Name",
    "ID",
    "Top-level family",
    "Family",
    "Group",
    "ISO code",
    "Glottocode",
    "Location",
    "Latitude",
    "Longitude",
    "Supervisor",
    "Informant",
    "Historical",
    "Source",
    "Status",
    "Assigned user",
    "Email",
    "Date last change",
]


_BOLD_WHITE = Font(bold=True, color="FFFFFF")


def _split_lines_to_str(value: Optional[str]) -> str:
    """Normalizza newlines per concatenare in cella multilinea."""
    if not value:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n")


def _style_table(ws, name: str, n_cols: int, widths: List[int]):
    """Applica TableStyle e larghezze colonne."""
    if ws.max_row < 2:
        # comunque applico le larghezze
        for idx, w in enumerate(widths, start=1):
            if idx > n_cols:
                break
            ws.column_dimensions[get_column_letter(idx)].width = w
        return
    ref = f"A1:{get_column_letter(n_cols)}{ws.max_row}"
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)
    ws.freeze_panes = "A2"
    for idx, w in enumerate(widths, start=1):
        if idx > n_cols:
            break
        ws.column_dimensions[get_column_letter(idx)].width = w


def _bold_header_row(ws, n_cols: int):
    for i in range(1, n_cols + 1):
        ws.cell(row=1, column=i).font = _BOLD_WHITE


def _pretty_qc_from_status(status: Optional[str]) -> str:
    """Mappa lo status di un Answer alla label leggibile (replicato dal vecchio progetto)."""
    s = (status or "").lower()
    if s == "approved":
        return "Done"
    if s in ("waiting_for_approval", "waiting"):
        return "Needs review"
    return "Not compiled"


# ============================================================================
# 1. LANGUAGE WORKBOOK (7 sheet per admin, 1 per user)
# ============================================================================

def build_language_workbook(
    db: Session,
    lang: models.Language,
    is_admin: bool = True,
) -> Workbook:
    """
    Genera il workbook di una singola lingua.

    Admin: 7 sheet [Database_model, Answers, Examples, Motivations, Parameters,
                    Questions, QuestionAllowedMotivations]
    User:  1 sheet [Examples]
    """
    # Pre-load: parametri attivi ordinati
    params = (
        db.query(models.ParameterDef)
        .filter(models.ParameterDef.is_active == True)
        .order_by(models.ParameterDef.position, models.ParameterDef.id)
        .all()
    )

    # Domande per parametro (ordinate per id). Solo attive: gli sheet di
    # esempi/risposte non devono mostrare question disattivate.
    # (Lo sheet "Questions" di metadati più sotto fa una query a parte e
    # continua a includere anche le inactive, perché documenta lo schema.)
    questions_by_param: dict[str, list[models.Question]] = {}
    all_questions = (
        db.query(models.Question)
        .filter(models.Question.is_active == True)
        .order_by(models.Question.parameter_id, models.Question.id)
        .all()
    )
    for q in all_questions:
        questions_by_param.setdefault(q.parameter_id, []).append(q)

    # Risposte per la lingua, indicizzate per question_id
    answers = (
        db.query(models.Answer)
        .filter(models.Answer.language_id == lang.id)
        .all()
    )
    answers_by_qid = {a.question_id: a for a in answers}

    # Esempi raggruppati per question_id (passando per la answer)
    examples_by_qid: dict[str, list[models.Example]] = {}
    for a in answers:
        if a.examples:
            examples_by_qid[a.question_id] = sorted(
                list(a.examples),
                key=lambda e: _example_sort_key(e),
            )

    # AnswerMotivation -> mappa motivation_id per veloce lookup
    mot_by_id = {m.id: m for m in db.query(models.Motivation).all()}

    # Admin notes per parametro (lingua corrente). Servono al foglio "Admin Notes"
    # e alla colonna Admin_Note di Database_model. Filtra in Python così
    # l'unico round-trip in DB è lineare nei parametri della lingua.
    notes_by_pid = {
        s.parameter_id: (s.admin_note or "")
        for s in db.query(models.LanguageParameterStatus)
        .filter(models.LanguageParameterStatus.language_id == lang.id)
        .all()
        if s.admin_note
    }

    wb = Workbook()

    # === Sheet Examples (sempre presente) ===
    ws_examples = wb.active
    ws_examples.title = "Examples"
    ws_examples.append(EXAMPLES_HEADERS)
    _bold_header_row(ws_examples, len(EXAMPLES_HEADERS))

    for p in params:
        for q in questions_by_param.get(p.id, []):
            for ex in examples_by_qid.get(q.id, []):
                ws_examples.append([
                    lang.id,
                    q.id,
                    ex.number or "",
                    ex.textarea or "",
                    ex.transliteration or "",
                    ex.gloss or "",
                    ex.translation or "",
                    ex.reference or "",
                ])

    _style_table(ws_examples, "Examples", len(EXAMPLES_HEADERS), [14, 14, 10, 36, 22, 22, 26, 24])

    if not is_admin:
        apply_excel_citation(wb)
        return wb

    # === Sheet Database_model (admin, primo sheet visivo) ===
    ws_db = wb.create_sheet("Database_model", 0)
    ws_db.append(DATABASE_MODEL_HEADERS)
    _bold_header_row(ws_db, len(DATABASE_MODEL_HEADERS))

    for p in params:
        param_admin_note = notes_by_pid.get(p.id, "")
        for q in questions_by_param.get(p.id, []):
            a = answers_by_qid.get(q.id)
            lang_answer = ""
            lang_comments = ""
            mot_codes_str = ""
            if a:
                if a.response_text == "yes":
                    lang_answer = "YES"
                elif a.response_text == "no":
                    lang_answer = "NO"
                lang_comments = a.comments or ""
                # Codici motivazione (non label): identificatore stabile per
                # round-trip. Separatore "; " coerente col foglio Answers.
                codes = []
                for am in a.answer_motivations:
                    m = mot_by_id.get(am.motivation_id)
                    if m and m.code:
                        codes.append(m.code)
                mot_codes_str = "; ".join(codes)

            ex_list = examples_by_qid.get(q.id, [])
            cell_examples = "\n".join((ex.textarea or "") for ex in ex_list) if ex_list else ""
            cell_gloss = "\n".join((ex.gloss or "") for ex in ex_list) if ex_list else ""
            cell_transl = "\n".join((ex.translation or "") for ex in ex_list) if ex_list else ""
            cell_refs = "\n".join((ex.reference or "") for ex in ex_list) if ex_list else ""

            ws_db.append([
                lang.name_full,
                p.id,
                q.id,
                lang_answer,
                lang_comments,
                cell_examples,
                cell_gloss,
                cell_transl,
                cell_refs,
                mot_codes_str,
                param_admin_note,
            ])

    _style_table(
        ws_db, "DatabaseModel", len(DATABASE_MODEL_HEADERS),
        [18, 14, 18, 12, 26, 30, 22, 22, 22, 22, 30],
    )

    # === Sheet Answers (admin) ===
    ws_ans = wb.create_sheet("Answers", 1)
    ws_ans.append(ANSWERS_HEADERS)
    _bold_header_row(ws_ans, len(ANSWERS_HEADERS))

    # Carico anche eval per il "Parameter value"
    lp_qs = (
        db.query(models.LanguageParameter)
        .filter(models.LanguageParameter.language_id == lang.id)
        .all()
    )
    value_orig_by_pid = {lp.parameter_id: (lp.value_orig or "") for lp in lp_qs}
    value_eval_by_pid: dict[str, str] = {}
    for lp in lp_qs:
        if lp.eval and lp.eval.value_eval:
            value_eval_by_pid[lp.parameter_id] = lp.eval.value_eval
        else:
            value_eval_by_pid[lp.parameter_id] = ""

    for p in params:
        for q in questions_by_param.get(p.id, []):
            a = answers_by_qid.get(q.id)
            param_value = value_eval_by_pid.get(q.parameter_id) or value_orig_by_pid.get(q.parameter_id) or ""
            if a:
                mot_codes = []
                for am in a.answer_motivations:
                    m = mot_by_id.get(am.motivation_id)
                    if m:
                        mot_codes.append(m.label or m.code)
                ws_ans.append([
                    lang.id,
                    p.id,
                    q.id,
                    q.text or "",
                    _pretty_qc_from_status(a.status),
                    a.response_text or "",
                    param_value,
                    "; ".join(mot_codes),
                    a.comments or "",
                ])
            else:
                ws_ans.append([
                    lang.id,
                    p.id,
                    q.id,
                    q.text or "",
                    "Not compiled",
                    "",
                    param_value,
                    "",
                    "",
                ])

    _style_table(ws_ans, "Answers", len(ANSWERS_HEADERS), [14, 18, 14, 36, 18, 10, 16, 28, 26])

    # === Sheet Admin Notes (admin) ===
    # Nota libera (testo) per ogni (lingua, parametro). Vivono su
    # LanguageParameterStatus.admin_note. Lo sheet contiene solo i parametri
    # con una nota non vuota, ordinati come gli altri sheet. Riusa
    # `notes_by_pid` già caricato in cima alla funzione per Database_model.
    ws_notes = wb.create_sheet("Admin Notes")
    ws_notes.append(ADMIN_NOTES_HEADERS)
    _bold_header_row(ws_notes, len(ADMIN_NOTES_HEADERS))

    for p in params:
        note = notes_by_pid.get(p.id, "")
        if note:
            ws_notes.append([p.id, p.name or "", note])

    _style_table(ws_notes, "AdminNotes", len(ADMIN_NOTES_HEADERS), [14, 36, 60])

    # I fogli schema (Motivations / Parameters / Questions /
    # QuestionAllowedMotivations) NON sono più replicati in ogni xlsx
    # per-lingua: sono globali, identici tra lingue, e gonfiavano inutilmente
    # ogni file. Vivono ora in `schema.xlsx` alla radice del backup-zip e
    # restano scaricabili via /api/admin/export/schema/xlsx.

    apply_excel_citation(wb)
    return wb


def _example_sort_key(ex: models.Example):
    """Ordina gli esempi per number numerico (se possibile), poi per id."""
    try:
        return (0, int(ex.number or "0"), ex.id or 0)
    except (ValueError, TypeError):
        return (1, str(ex.number or ""), ex.id or 0)


# ============================================================================
# 2. LANGUAGE LIST WORKBOOK (metadati delle lingue selezionate)
# ============================================================================

def _xlsx_sanitize(v):
    """Converte valori non-Excel-friendly in stringhe."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, datetime):
        return v.replace(tzinfo=None) if v.tzinfo else v
    return v


def build_language_list_workbook(
    db: Session,
    languages: Iterable[models.Language],
) -> Workbook:
    """
    Workbook con un solo sheet 'Languages' contenente i metadati delle lingue.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Languages"
    ws.append(LANGUAGE_LIST_HEADERS)
    _bold_header_row(ws, len(LANGUAGE_LIST_HEADERS))

    for L in languages:
        assigned_name = ""
        assigned_email = ""
        if L.assigned_user_id:
            user = db.query(models.User).filter(models.User.id == L.assigned_user_id).first()
            if user:
                full = f"{user.name or ''} {user.surname or ''}".strip()
                assigned_name = full or (user.email or "")
                assigned_email = user.email or ""

        ws.append([
            _xlsx_sanitize(L.name_full),
            _xlsx_sanitize(L.id),
            _xlsx_sanitize(L.top_level_family),
            _xlsx_sanitize(L.family),
            _xlsx_sanitize(L.grp),
            _xlsx_sanitize(L.isocode),
            _xlsx_sanitize(L.glottocode),
            _xlsx_sanitize(L.location),
            _xlsx_sanitize(float(L.latitude) if L.latitude is not None else None),
            _xlsx_sanitize(float(L.longitude) if L.longitude is not None else None),
            _xlsx_sanitize(L.supervisor),
            _xlsx_sanitize(L.informant),
            _xlsx_sanitize(L.historical_language),
            _xlsx_sanitize(L.source),
            _xlsx_sanitize(L.status or "pending"),
            _xlsx_sanitize(assigned_name),
            _xlsx_sanitize(assigned_email),
            L.updated_at if L.updated_at is not None else None,
        ])

    # Stile tabella: header bianco su fondo blu (TableStyleMedium2) +
    # freeze pane + larghezze. Senza il fill colorato del table style
    # l'header (font bianco) sarebbe invisibile su fondo bianco.
    n_cols = len(LANGUAGE_LIST_HEADERS)
    widths = [22, 10, 16, 16, 14, 10, 12, 18, 10, 10, 16, 16, 10, 28, 12, 22, 26, 18]
    _style_table(ws, "Languages", n_cols, widths)

    # Formatta la colonna Date last change (ultima) come datetime locale
    last_col = get_column_letter(n_cols)
    for cell in ws[last_col][1:]:
        cell.number_format = "yyyy-mm-dd hh:mm"

    apply_excel_citation(wb)
    return wb


# ============================================================================
# 3. SCHEMA WORKBOOK (4 sheet: motivations, parameters, questions, qam)
# ============================================================================

def _append_schema_sheets(db: Session, wb: Workbook) -> None:
    """Aggiunge i 4 sheet schema al workbook esistente (in coda)."""
    # Motivations
    ws_mot = wb.create_sheet("Motivations")
    ws_mot.append(MOTIVATIONS_HEADERS)
    _bold_header_row(ws_mot, len(MOTIVATIONS_HEADERS))
    for m in db.query(models.Motivation).order_by(models.Motivation.code).all():
        ws_mot.append([
            m.id,
            m.code or "",
            m.label or "",
        ])
    _style_table(ws_mot, "Motivations", len(MOTIVATIONS_HEADERS), [8, 14, 50])

    # Parameters
    ws_par = wb.create_sheet("Parameters")
    ws_par.append(PARAMETERS_HEADERS)
    _bold_header_row(ws_par, len(PARAMETERS_HEADERS))
    for p in db.query(models.ParameterDef).order_by(models.ParameterDef.position, models.ParameterDef.id).all():
        ws_par.append([
            p.id,
            p.position,
            p.name or "",
            p.schema or "",
            p.param_type or "",
            p.level_of_comparison or "",
            p.short_description or "",
            p.long_description or "",
            p.implicational_condition or "",
            p.description_of_the_implicational_condition or "",
            "Yes" if p.is_active else "No",
        ])
    _style_table(ws_par, "Parameters", len(PARAMETERS_HEADERS), [10, 8, 30, 16, 16, 18, 30, 30, 22, 30, 10])

    # Questions
    ws_q = wb.create_sheet("Questions")
    ws_q.append(QUESTIONS_HEADERS)
    _bold_header_row(ws_q, len(QUESTIONS_HEADERS))
    for q in db.query(models.Question).order_by(models.Question.parameter_id, models.Question.id).all():
        ws_q.append([
            q.id,
            q.parameter_id,
            q.text or "",
            q.template_type or "",
            q.instruction or "",
            q.instruction_yes or "",
            q.instruction_no or "",
            q.example_yes or "",
            q.help_info or "",
            "Yes" if q.is_stop_question else "No",
            "Yes" if q.is_active else "No",
        ])
    _style_table(ws_q, "Questions", len(QUESTIONS_HEADERS), [16, 14, 36, 14, 24, 24, 24, 24, 24, 12, 10])

    # QuestionAllowedMotivations
    ws_qam = wb.create_sheet("QuestionAllowedMotivations")
    ws_qam.append(QUESTION_ALLOWED_MOTIVATIONS_HEADERS)
    _bold_header_row(ws_qam, len(QUESTION_ALLOWED_MOTIVATIONS_HEADERS))
    for qam in (
        db.query(models.QuestionAllowedMotivation)
        .order_by(models.QuestionAllowedMotivation.question_id)
        .all()
    ):
        m = db.query(models.Motivation).filter(models.Motivation.id == qam.motivation_id).first()
        ws_qam.append([
            qam.question_id,
            m.code if m else "",
        ])
    _style_table(ws_qam, "QuestionAllowedMotivations", len(QUESTION_ALLOWED_MOTIVATIONS_HEADERS), [16, 14])


def build_schema_workbook(db: Session) -> Workbook:
    """
    Workbook con SOLO i 4 sheet schema (parametri/domande/motivazioni).
    Esportabile dalla pagina Parameters per editing offline.
    """
    wb = Workbook()
    # Rimuovo lo sheet di default
    default = wb.active
    wb.remove(default)
    _append_schema_sheets(db, wb)
    apply_excel_citation(wb)
    return wb


# ============================================================================
# 4. GLOSSARY WORKBOOK (1 sheet)
# ============================================================================

GLOSSARY_HEADERS = ["Word", "Description"]


def build_glossary_workbook(db: Session) -> Workbook:
    """Workbook con un solo sheet 'Glossary' (Word, Description).

    Usato dal backup-zip per portare in dote anche i termini del glossario:
    sono dati di "contenuto" che dovrebbero sopravvivere a un restore."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(GLOSSARY_HEADERS)
    _bold_header_row(ws, len(GLOSSARY_HEADERS))

    for g in db.query(models.Glossary).order_by(models.Glossary.word).all():
        ws.append([_xlsx_sanitize(g.word), _xlsx_sanitize(g.description)])

    _style_table(ws, "Glossary", len(GLOSSARY_HEADERS), [22, 60])
    apply_excel_citation(wb)
    return wb


# ============================================================================
# 5. EXTRAS WORKBOOKS (per il bundle "full" — vedi sezione 6 più sotto)
# ============================================================================
#
# Tabelle "satellite" non incluse nel bundle base (PCM_backup.zip): contengono
# storico/contenuto editabile che non serve per i dati linguistici ma che è
# utile preservare per un disaster recovery completo.
#   - SiteContent      → testi editabili (HowToCite, About, ecc.)
#   - Submissions      → snapshot di lingue inviate per approvazione
#   - ParameterSubmissions → snapshot delle definizioni di parametri
#   - ArchivedQuestions    → archivio di domande dismesse + dati collegati
# Gli utenti NON sono inclusi: chi restora ricrea/importa users separatamente.

SITE_CONTENT_HEADERS = ["Key", "Content", "Page", "Updated At", "Updated By Email"]

SUBMISSIONS_HEADERS = ["ID", "Language ID", "Submitted By Email", "Submitted At", "Note"]
SUBMISSION_ANSWERS_HEADERS = [
    "ID", "Submission ID", "Question Code", "Response Text", "Comments",
]
SUBMISSION_EXAMPLES_HEADERS = [
    "ID", "Submission ID", "Question Code",
    "Textarea", "Transliteration", "Gloss", "Translation", "Reference",
]
SUBMISSION_ANSWER_MOTIVATIONS_HEADERS = [
    "ID", "Submission ID", "Question Code", "Motivation Code", "Motivation Label",
]
SUBMISSION_PARAMS_HEADERS = [
    "ID", "Submission ID", "Parameter ID",
    "Value Orig", "Warning Orig", "Value Eval", "Warning Eval", "Evaluated At",
]

PARAMETER_SUBMISSIONS_HEADERS = [
    "ID", "Parameter ID", "Parameter Name",
    "Submitted By Email", "Submitted At", "Note",
    "Short Description", "Long Description",
    "Implicational Condition", "Description Of Implicational Condition",
    "Is Active", "Position", "Schema", "Param Type", "Level Of Comparison",
]
PARAMETER_SUBMISSION_QUESTIONS_HEADERS = [
    "ID", "Submission ID", "Question Code", "Text", "Template Type",
    "Instruction", "Instruction YES", "Instruction NO",
    "Example YES", "Help Info", "Is Stop Question", "Is Active",
]
PARAMETER_SUBMISSION_ALLOWED_MOTIVATIONS_HEADERS = [
    "ID", "Question ID", "Motivation Code", "Motivation Label",
]

ARCHIVED_QUESTIONS_HEADERS = [
    "ID", "Original Question ID", "Parameter ID", "Parameter Name",
    "Text", "Template Type",
    "Instruction", "Instruction YES", "Instruction NO",
    "Example YES", "Help Info", "Is Stop Question", "Is Active",
    "Archived At", "Archived By Email", "Archive Note",
    "Answers Count", "Examples Count",
]
ARCHIVED_QUESTION_MOTIVATIONS_HEADERS = [
    "ID", "Archived Question ID", "Motivation Code", "Motivation Label",
]
ARCHIVED_ANSWERS_HEADERS = [
    "ID", "Archived Question ID", "Language ID", "Language Name Full",
    "Status", "Response Text", "Comments", "Original Updated At",
]
ARCHIVED_EXAMPLES_HEADERS = [
    "ID", "Archived Answer ID", "Number",
    "Textarea", "Transliteration", "Gloss", "Translation", "Reference",
]
ARCHIVED_ANSWER_MOTIVATIONS_HEADERS = [
    "ID", "Archived Answer ID", "Motivation Code", "Motivation Label",
]


def _user_email_map(db: Session) -> dict:
    """{user.id: user.email} per denormalizzare gli FK utente nei file extras.

    Salvare l'email invece dell'id rende il bundle indipendente dagli id
    locali del DB sorgente: in fase di restore basta lookup-by-email, e se
    l'utente non c'è l'FK resta NULL (tutte le colonne FK utente in queste
    tabelle sono nullable o ON DELETE SET NULL)."""
    return {u.id: u.email for u in db.query(models.User.id, models.User.email).all()}


def build_site_content_workbook(db: Session) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "SiteContents"
    ws.append(SITE_CONTENT_HEADERS)
    _bold_header_row(ws, len(SITE_CONTENT_HEADERS))

    user_email = _user_email_map(db)
    for r in db.query(models.SiteContent).order_by(models.SiteContent.key).all():
        ws.append([
            _xlsx_sanitize(r.key),
            _xlsx_sanitize(r.content),
            _xlsx_sanitize(r.page),
            _xlsx_sanitize(r.updated_at),
            user_email.get(r.updated_by_id, "") if r.updated_by_id else "",
        ])

    _style_table(ws, "SiteContents", len(SITE_CONTENT_HEADERS), [22, 60, 18, 18, 28])
    apply_excel_citation(wb)
    return wb


def build_submissions_workbook(db: Session) -> Workbook:
    """5 sheet: Submissions (master) + 4 figlie con FK su Submission ID."""
    wb = Workbook()
    wb.remove(wb.active)

    user_email = _user_email_map(db)

    ws = wb.create_sheet("Submissions")
    ws.append(SUBMISSIONS_HEADERS)
    _bold_header_row(ws, len(SUBMISSIONS_HEADERS))
    for s in db.query(models.Submission).order_by(models.Submission.id).all():
        ws.append([
            s.id,
            s.language_id,
            user_email.get(s.submitted_by_id, "") if s.submitted_by_id else "",
            _xlsx_sanitize(s.submitted_at),
            s.note or "",
        ])
    _style_table(ws, "Submissions", len(SUBMISSIONS_HEADERS), [10, 14, 26, 18, 30])

    ws = wb.create_sheet("SubmissionAnswers")
    ws.append(SUBMISSION_ANSWERS_HEADERS)
    _bold_header_row(ws, len(SUBMISSION_ANSWERS_HEADERS))
    for sa in (
        db.query(models.SubmissionAnswer)
        .order_by(models.SubmissionAnswer.submission_id, models.SubmissionAnswer.id)
        .all()
    ):
        ws.append([sa.id, sa.submission_id, sa.question_code, sa.response_text or "", sa.comments or ""])
    _style_table(ws, "SubmissionAnswers", len(SUBMISSION_ANSWERS_HEADERS), [10, 14, 14, 14, 30])

    ws = wb.create_sheet("SubmissionExamples")
    ws.append(SUBMISSION_EXAMPLES_HEADERS)
    _bold_header_row(ws, len(SUBMISSION_EXAMPLES_HEADERS))
    for se in (
        db.query(models.SubmissionExample)
        .order_by(models.SubmissionExample.submission_id, models.SubmissionExample.id)
        .all()
    ):
        ws.append([
            se.id, se.submission_id, se.question_code,
            se.textarea or "", se.transliteration or "",
            se.gloss or "", se.translation or "", se.reference or "",
        ])
    _style_table(ws, "SubmissionExamples", len(SUBMISSION_EXAMPLES_HEADERS), [10, 14, 14, 36, 22, 22, 26, 24])

    ws = wb.create_sheet("SubmissionAnswerMotivations")
    ws.append(SUBMISSION_ANSWER_MOTIVATIONS_HEADERS)
    _bold_header_row(ws, len(SUBMISSION_ANSWER_MOTIVATIONS_HEADERS))
    for sam in (
        db.query(models.SubmissionAnswerMotivation)
        .order_by(models.SubmissionAnswerMotivation.submission_id, models.SubmissionAnswerMotivation.id)
        .all()
    ):
        ws.append([sam.id, sam.submission_id, sam.question_code, sam.motivation_code, sam.motivation_label or ""])
    _style_table(ws, "SubmissionAnswerMotivations", len(SUBMISSION_ANSWER_MOTIVATIONS_HEADERS), [10, 14, 14, 16, 30])

    ws = wb.create_sheet("SubmissionParams")
    ws.append(SUBMISSION_PARAMS_HEADERS)
    _bold_header_row(ws, len(SUBMISSION_PARAMS_HEADERS))
    for sp in (
        db.query(models.SubmissionParam)
        .order_by(models.SubmissionParam.submission_id, models.SubmissionParam.id)
        .all()
    ):
        ws.append([
            sp.id, sp.submission_id, sp.parameter_id,
            sp.value_orig or "", "Yes" if sp.warning_orig else "No",
            sp.value_eval or "", "Yes" if sp.warning_eval else "No",
            _xlsx_sanitize(sp.evaluated_at),
        ])
    _style_table(ws, "SubmissionParams", len(SUBMISSION_PARAMS_HEADERS), [10, 14, 14, 12, 12, 12, 12, 18])

    apply_excel_citation(wb)
    return wb


def build_parameter_submissions_workbook(db: Session) -> Workbook:
    """3 sheet: ParameterSubmissions (master) + Questions + AllowedMotivations."""
    wb = Workbook()
    wb.remove(wb.active)
    user_email = _user_email_map(db)

    ws = wb.create_sheet("ParameterSubmissions")
    ws.append(PARAMETER_SUBMISSIONS_HEADERS)
    _bold_header_row(ws, len(PARAMETER_SUBMISSIONS_HEADERS))
    for ps in db.query(models.ParameterSubmission).order_by(models.ParameterSubmission.id).all():
        ws.append([
            ps.id, ps.parameter_id, ps.parameter_name or "",
            user_email.get(ps.submitted_by_id, "") if ps.submitted_by_id else "",
            _xlsx_sanitize(ps.submitted_at), ps.note or "",
            ps.short_description or "", ps.long_description or "",
            ps.implicational_condition or "",
            ps.description_of_the_implicational_condition or "",
            "Yes" if ps.is_active else "No",
            ps.position if ps.position is not None else "",
            ps.schema or "", ps.param_type or "", ps.level_of_comparison or "",
        ])
    _style_table(ws, "ParameterSubmissions", len(PARAMETER_SUBMISSIONS_HEADERS),
                 [8, 10, 24, 26, 18, 24, 30, 30, 22, 30, 10, 10, 14, 14, 18])

    ws = wb.create_sheet("Questions")
    ws.append(PARAMETER_SUBMISSION_QUESTIONS_HEADERS)
    _bold_header_row(ws, len(PARAMETER_SUBMISSION_QUESTIONS_HEADERS))
    for psq in (
        db.query(models.ParameterSubmissionQuestion)
        .order_by(models.ParameterSubmissionQuestion.submission_id,
                  models.ParameterSubmissionQuestion.id)
        .all()
    ):
        ws.append([
            psq.id, psq.submission_id, psq.question_code, psq.text or "",
            psq.template_type or "", psq.instruction or "",
            psq.instruction_yes or "", psq.instruction_no or "",
            psq.example_yes or "", psq.help_info or "",
            "Yes" if psq.is_stop_question else "No",
            "Yes" if psq.is_active else "No",
        ])
    _style_table(ws, "ParamSubmissionQuestions", len(PARAMETER_SUBMISSION_QUESTIONS_HEADERS),
                 [8, 10, 14, 36, 14, 24, 24, 24, 24, 24, 12, 10])

    ws = wb.create_sheet("AllowedMotivations")
    ws.append(PARAMETER_SUBMISSION_ALLOWED_MOTIVATIONS_HEADERS)
    _bold_header_row(ws, len(PARAMETER_SUBMISSION_ALLOWED_MOTIVATIONS_HEADERS))
    for am in (
        db.query(models.ParameterSubmissionAllowedMotivation)
        .order_by(models.ParameterSubmissionAllowedMotivation.question_id,
                  models.ParameterSubmissionAllowedMotivation.id)
        .all()
    ):
        ws.append([am.id, am.question_id, am.motivation_code, am.motivation_label or ""])
    _style_table(ws, "ParamSubmissionAllowedMotivations",
                 len(PARAMETER_SUBMISSION_ALLOWED_MOTIVATIONS_HEADERS),
                 [8, 12, 16, 30])

    apply_excel_citation(wb)
    return wb


def build_archived_questions_workbook(db: Session) -> Workbook:
    """5 sheet: ArchivedQuestions + 4 figlie/nipoti (3 livelli di gerarchia)."""
    wb = Workbook()
    wb.remove(wb.active)
    user_email = _user_email_map(db)

    ws = wb.create_sheet("ArchivedQuestions")
    ws.append(ARCHIVED_QUESTIONS_HEADERS)
    _bold_header_row(ws, len(ARCHIVED_QUESTIONS_HEADERS))
    for aq in db.query(models.ArchivedQuestion).order_by(models.ArchivedQuestion.id).all():
        ws.append([
            aq.id, aq.original_question_id, aq.parameter_id, aq.parameter_name or "",
            aq.text or "", aq.template_type or "",
            aq.instruction or "", aq.instruction_yes or "", aq.instruction_no or "",
            aq.example_yes or "", aq.help_info or "",
            "Yes" if aq.is_stop_question else "No",
            "Yes" if aq.is_active else "No",
            _xlsx_sanitize(aq.archived_at),
            user_email.get(aq.archived_by_id, "") if aq.archived_by_id else "",
            aq.archive_note or "",
            aq.answers_count or 0, aq.examples_count or 0,
        ])
    _style_table(ws, "ArchivedQuestions", len(ARCHIVED_QUESTIONS_HEADERS),
                 [8, 14, 10, 24, 36, 14, 24, 24, 24, 24, 24, 12, 10, 18, 26, 30, 10, 10])

    ws = wb.create_sheet("ArchivedQuestionMotivations")
    ws.append(ARCHIVED_QUESTION_MOTIVATIONS_HEADERS)
    _bold_header_row(ws, len(ARCHIVED_QUESTION_MOTIVATIONS_HEADERS))
    for aqm in (
        db.query(models.ArchivedQuestionMotivation)
        .order_by(models.ArchivedQuestionMotivation.archived_question_id,
                  models.ArchivedQuestionMotivation.id)
        .all()
    ):
        ws.append([aqm.id, aqm.archived_question_id, aqm.motivation_code, aqm.motivation_label or ""])
    _style_table(ws, "ArchivedQuestionMotivations", len(ARCHIVED_QUESTION_MOTIVATIONS_HEADERS),
                 [8, 14, 16, 30])

    ws = wb.create_sheet("ArchivedAnswers")
    ws.append(ARCHIVED_ANSWERS_HEADERS)
    _bold_header_row(ws, len(ARCHIVED_ANSWERS_HEADERS))
    for aa in (
        db.query(models.ArchivedAnswer)
        .order_by(models.ArchivedAnswer.archived_question_id, models.ArchivedAnswer.id)
        .all()
    ):
        ws.append([
            aa.id, aa.archived_question_id, aa.language_id, aa.language_name_full or "",
            aa.status or "", aa.response_text or "", aa.comments or "",
            _xlsx_sanitize(aa.original_updated_at),
        ])
    _style_table(ws, "ArchivedAnswers", len(ARCHIVED_ANSWERS_HEADERS),
                 [8, 14, 12, 22, 14, 14, 30, 18])

    ws = wb.create_sheet("ArchivedExamples")
    ws.append(ARCHIVED_EXAMPLES_HEADERS)
    _bold_header_row(ws, len(ARCHIVED_EXAMPLES_HEADERS))
    for ae in (
        db.query(models.ArchivedExample)
        .order_by(models.ArchivedExample.archived_answer_id, models.ArchivedExample.id)
        .all()
    ):
        ws.append([
            ae.id, ae.archived_answer_id, ae.number or "",
            ae.textarea or "", ae.transliteration or "",
            ae.gloss or "", ae.translation or "", ae.reference or "",
        ])
    _style_table(ws, "ArchivedExamples", len(ARCHIVED_EXAMPLES_HEADERS),
                 [8, 14, 10, 36, 22, 22, 26, 24])

    ws = wb.create_sheet("ArchivedAnswerMotivations")
    ws.append(ARCHIVED_ANSWER_MOTIVATIONS_HEADERS)
    _bold_header_row(ws, len(ARCHIVED_ANSWER_MOTIVATIONS_HEADERS))
    for aam in (
        db.query(models.ArchivedAnswerMotivation)
        .order_by(models.ArchivedAnswerMotivation.archived_answer_id,
                  models.ArchivedAnswerMotivation.id)
        .all()
    ):
        ws.append([aam.id, aam.archived_answer_id, aam.motivation_code, aam.motivation_label or ""])
    _style_table(ws, "ArchivedAnswerMotivations", len(ARCHIVED_ANSWER_MOTIVATIONS_HEADERS),
                 [8, 14, 16, 30])

    apply_excel_citation(wb)
    return wb


# ============================================================================
# 6. BACKUP ZIP BUILDER
# ============================================================================
#
# Bundle base (`PCM_backup.zip`, prodotto da `build_backup_zip_bytes`):
#
#     PCM_backup_<ts>.zip
#     ├── schema.xlsx              (4 sheet schema globale)
#     ├── languages_metadata.xlsx  (lista lingue con metadati)
#     ├── glossary.xlsx            (Word, Description)
#     └── languages/
#         ├── <ID>.xlsx            (Database_model esteso + Answers + Examples + Admin Notes)
#         └── ...
#
# Bundle full (`PCM_full_backup.zip`, prodotto da `build_full_backup_zip_bytes`):
# stesso contenuto del bundle base + cartella `extras/` con:
#   ├── extras/site_content.xlsx
#   ├── extras/submissions.xlsx
#   ├── extras/parameter_submissions.xlsx
#   └── extras/archived_questions.xlsx
#
# Il restore (services/backup_restore.py) riconosce entrambi i formati: se la
# cartella extras/ è assente processa solo i file base (compat. piena).
# ============================================================================

BACKUP_BUNDLE_VERSION = 1


def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_base_bundle(
    zf: zipfile.ZipFile,
    db: Session,
    languages: list,
    on_language=None,
) -> None:
    """Scrive nel zip i file del bundle base (schema + metadata + glossary +
    languages/). Riusato sia da build_backup_zip_bytes sia dal builder full."""
    zf.writestr("schema.xlsx", _wb_to_bytes(build_schema_workbook(db)))
    zf.writestr(
        "languages_metadata.xlsx",
        _wb_to_bytes(build_language_list_workbook(db, languages)),
    )
    zf.writestr("glossary.xlsx", _wb_to_bytes(build_glossary_workbook(db)))

    total = len(languages)
    for idx, lang in enumerate(languages, start=1):
        if on_language is not None:
            try:
                on_language(idx, total, lang)
            except Exception:
                # Mai bloccare la generazione del backup per un errore di reporting
                pass
        wb = build_language_workbook(db, lang, is_admin=True)
        zf.writestr(f"languages/{lang.id}.xlsx", _wb_to_bytes(wb))


def build_backup_zip_bytes(
    db: Session,
    languages: Iterable[models.Language],
    *,
    on_language=None,
) -> bytes:
    """Costruisce il bundle backup base per la selezione di `languages` data.

    Restituisce i bytes dello zip (NON streaming). Per file grandi conviene
    generarlo in background e servirlo da una tmp directory: vedi Fase 4
    (export asincrono via migration_progress).

    `on_language(idx, total, lang)` è un callback opzionale invocato prima di
    serializzare ogni xlsx per-lingua, utile per il progress reporting.
    """
    languages = list(languages)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        _write_base_bundle(zf, db, languages, on_language=on_language)
    return buf.getvalue()


def build_full_backup_zip_bytes(
    db: Session,
    languages: Iterable[models.Language],
    *,
    on_language=None,
) -> bytes:
    """Bundle full: bundle base + cartella `extras/` con site_content,
    submissions, parameter_submissions, archived_questions.

    Stessa firma di `build_backup_zip_bytes` per facilità di sostituzione.
    Gli utenti NON sono inclusi (scelta deliberata: vanno gestiti con un
    flusso dedicato — pg_dump o reimport manuale)."""
    languages = list(languages)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        _write_base_bundle(zf, db, languages, on_language=on_language)
        zf.writestr("extras/site_content.xlsx", _wb_to_bytes(build_site_content_workbook(db)))
        zf.writestr("extras/submissions.xlsx", _wb_to_bytes(build_submissions_workbook(db)))
        zf.writestr("extras/parameter_submissions.xlsx", _wb_to_bytes(build_parameter_submissions_workbook(db)))
        zf.writestr("extras/archived_questions.xlsx", _wb_to_bytes(build_archived_questions_workbook(db)))
    return buf.getvalue()
