from sqlalchemy.orm import Session, joinedload
from datetime import datetime
from time_utils import utc_now
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import models
from services.citation import apply_excel_citation

# Massimo numero di salvataggi storici mantenuti per ogni lingua
MAX_PER_LANGUAGE = 10

def create_language_submission(db: Session, language: models.Language, user_id: int, note: str = "", fixed_time: datetime = None):
    """
    Crea uno snapshot 'full' per una singola lingua.
    Equivalente al vecchio services.py di Django.
    """
    now = fixed_time or utc_now()

    # 1. Creazione record principale Submission
    sub = models.Submission(
        language_id=language.id,
        submitted_by_id=user_id,
        submitted_at=now,
        note=note or ""
    )
    db.add(sub)
    db.flush() # Fa l'insert nel DB per ottenere l'ID della submission, ma senza committare

    # 2. Estrazione e copia delle Answers (con Motivations e Examples)
    # Usiamo joinedload per evitare il problema query N+1, come faceva select_related/prefetch_related
    answers = db.query(models.Answer).options(
        joinedload(models.Answer.examples),
        joinedload(models.Answer.answer_motivations).joinedload(models.AnswerMotivation.motivation)
    ).filter(models.Answer.language_id == language.id).all()

    sub_answers = []
    sub_mots = []
    sub_ex = []

    for a in answers:
        sub_answers.append(models.SubmissionAnswer(
            submission_id=sub.id,
            question_code=a.question_id,
            response_text=a.response_text,
            comments=a.comments or ""
        ))
        for am in a.answer_motivations:
            sub_mots.append(models.SubmissionAnswerMotivation(
                submission_id=sub.id,
                question_code=a.question_id,
                motivation_code=am.motivation.code,
                motivation_label=am.motivation.label,
            ))
        for ex in a.examples:
            sub_ex.append(models.SubmissionExample(
                submission_id=sub.id,
                question_code=a.question_id,
                textarea=ex.textarea or "",
                transliteration=ex.transliteration or "",
                gloss=ex.gloss or "",
                translation=ex.translation or "",
                reference=ex.reference or ""
            ))

    # 3. Estrazione e copia dei Parametri + Eval (DAG)
    lparams = db.query(models.LanguageParameter).options(
        joinedload(models.LanguageParameter.eval)
    ).filter(models.LanguageParameter.language_id == language.id).all()

    sub_params = []
    for lp in lparams:
        eval_obj = lp.eval
        sub_params.append(models.SubmissionParam(
            submission_id=sub.id,
            parameter_id=lp.parameter_id,
            value_orig=lp.value_orig,
            warning_orig=lp.warning_orig,
            value_eval=eval_obj.value_eval if eval_obj else "0",
            warning_eval=eval_obj.warning_eval if eval_obj else False,
            evaluated_at=now
        ))

    # Inserimento massivo stile bulk_create
    db.add_all(sub_answers)
    db.add_all(sub_mots)
    db.add_all(sub_ex)
    db.add_all(sub_params)
    db.flush()

    # 4. Pruning automatico per limitare lo storage
    subs = db.query(models.Submission.id).filter(
        models.Submission.language_id == language.id
    ).order_by(models.Submission.submitted_at.desc(), models.Submission.id.desc()).all()

    pruned_count = 0
    if len(subs) > MAX_PER_LANGUAGE:
        # Teniamo solo i primi N ID
        ids_to_keep = [s[0] for s in subs[:MAX_PER_LANGUAGE]]
        deleted = db.query(models.Submission).filter(
            models.Submission.language_id == language.id,
            models.Submission.id.notin_(ids_to_keep)
        ).delete(synchronize_session=False)
        pruned_count = deleted

    return sub, pruned_count

def create_all_languages_backup(db: Session, user_id: int, note: str = "Global backup"):
    """
    Forza un backup globale sincronizzato per tutte le lingue.
    Tutte le query condividono una singola transazione.
    """
    languages = db.query(models.Language).all()

    # Trucco fondamentale: azzeriamo i microsecondi per far sì che
    # tutto il backup appartenga alla stessa identica data (la nostra "cartella")
    fixed_time = utc_now().replace(microsecond=0)

    total_pruned = 0

    try:
        for lang in languages:
            _, pruned = create_language_submission(db, lang, user_id, note, fixed_time)
            total_pruned += pruned

        # Confermiamo l'intera transazione solo se tutte le lingue sono state processate con successo
        db.commit()

        return {
            "status": "success",
            "languages_backed_up": len(languages),
            "pruned": total_pruned,
            "timestamp": fixed_time
        }
    except Exception as e:
        # Se anche un solo elemento fallisce, annulliamo tutto (equivalente di with transaction.atomic() in Django)
        db.rollback()
        raise e


# ============================================================================
# Export di una Submission in xlsx (download dei backup, equivalente al
# download "Old questions archive"). Riusa lo snapshot in DB: niente lookup
# sui dati vivi, così il file riflette esattamente lo stato salvato.
# ============================================================================

_BOLD_WHITE = Font(bold=True, color="FFFFFF")


def _bold_header_row(ws, n_cols: int) -> None:
    for i in range(1, n_cols + 1):
        ws.cell(row=1, column=i).font = _BOLD_WHITE


def _style_table(ws, name: str, n_cols: int, widths) -> None:
    """Tabella stile TableStyleMedium2 con header su fondo blu (richiesto
    perché _bold_header_row mette font bianco): senza il fill l'header sarebbe
    invisibile su fondo bianco."""
    if ws.max_row >= 2:
        ref = f"A1:{get_column_letter(n_cols)}{ws.max_row}"
        tbl = Table(displayName=name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tbl)
        ws.freeze_panes = "A2"
    for idx, w in enumerate(widths, start=1):
        if idx > n_cols:
            break
        ws.column_dimensions[get_column_letter(idx)].width = w


_INFO_HEADERS = ["Field", "Value"]
_PARAMS_HEADERS = ["Parameter", "Initial value", "Warning init", "Final value", "Warning final"]
_ANSWERS_HEADERS = ["Question", "Answer", "Motivations", "Comments"]
_EXAMPLES_HEADERS = ["Question", "Example text", "Transliteration", "Gloss", "Translation", "Reference"]


def build_submission_workbook(db: Session, sub: models.Submission) -> Workbook:
    """Workbook per una singola Submission (backup di una lingua).

    Sheet:
      - Info       : Language id/name, Backup date, Submitted by, Note
      - Parameters : value_orig, warning_orig, value_eval, warning_eval
      - Answers    : question_code, response, motivazioni, comments
      - Examples   : question_code, textarea, transliteration, gloss, translation, ref

    Tutti i dati derivano dallo snapshot Submission*: niente lookup sulle
    tabelle vive — il file riflette lo stato congelato del backup.
    """
    wb = Workbook()

    # === Info ===
    ws_info = wb.active
    ws_info.title = "Info"
    ws_info.append(_INFO_HEADERS)
    _bold_header_row(ws_info, len(_INFO_HEADERS))

    lang = sub.language
    submitter = (
        f"{sub.submitted_by.name or ''} {sub.submitted_by.surname or ''}".strip()
        or (sub.submitted_by.email if sub.submitted_by else "")
    ) if sub.submitted_by_id else "System"
    submitted_at_str = sub.submitted_at.strftime("%Y-%m-%d %H:%M UTC") if sub.submitted_at else ""

    ws_info.append(["Language ID", lang.id if lang else (sub.language_id or "")])
    ws_info.append(["Language name", lang.name_full if lang else ""])
    ws_info.append(["Backup date (UTC)", submitted_at_str])
    ws_info.append(["Submitted by", submitter])
    ws_info.append(["Note", sub.note or ""])
    _style_table(ws_info, "BackupInfo", len(_INFO_HEADERS), [22, 60])

    # === Parameters ===
    ws_par = wb.create_sheet("Parameters")
    ws_par.append(_PARAMS_HEADERS)
    _bold_header_row(ws_par, len(_PARAMS_HEADERS))
    params_sorted = sorted(sub.params, key=lambda p: p.parameter_id or "")
    for p in params_sorted:
        ws_par.append([
            p.parameter_id or "",
            p.value_orig or "",
            "Yes" if p.warning_orig else "",
            p.value_eval or "",
            "Yes" if p.warning_eval else "",
        ])
    _style_table(ws_par, "BackupParameters", len(_PARAMS_HEADERS), [16, 14, 14, 14, 14])

    # Pre-aggrega le motivations per question_code (label fallback su code)
    mots_by_q: dict[str, list[str]] = {}
    for m in sub.answer_motivations:
        text = m.motivation_label or m.motivation_code or ""
        if text:
            mots_by_q.setdefault(m.question_code, []).append(text)

    # === Answers ===
    ws_ans = wb.create_sheet("Answers")
    ws_ans.append(_ANSWERS_HEADERS)
    _bold_header_row(ws_ans, len(_ANSWERS_HEADERS))
    answers_sorted = sorted(sub.answers, key=lambda a: a.question_code or "")
    for a in answers_sorted:
        resp = ""
        if a.response_text == "yes":
            resp = "YES"
        elif a.response_text == "no":
            resp = "NO"
        elif a.response_text:
            resp = a.response_text
        ws_ans.append([
            a.question_code or "",
            resp,
            "; ".join(mots_by_q.get(a.question_code, [])),
            a.comments or "",
        ])
    _style_table(ws_ans, "BackupAnswers", len(_ANSWERS_HEADERS), [16, 10, 30, 36])

    # === Examples ===
    ws_ex = wb.create_sheet("Examples")
    ws_ex.append(_EXAMPLES_HEADERS)
    _bold_header_row(ws_ex, len(_EXAMPLES_HEADERS))
    examples_sorted = sorted(sub.examples, key=lambda e: (e.question_code or "", e.id or 0))
    for e in examples_sorted:
        ws_ex.append([
            e.question_code or "",
            e.textarea or "",
            e.transliteration or "",
            e.gloss or "",
            e.translation or "",
            e.reference or "",
        ])
    _style_table(ws_ex, "BackupExamples", len(_EXAMPLES_HEADERS), [14, 36, 22, 22, 26, 22])

    apply_excel_citation(wb)
    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
    