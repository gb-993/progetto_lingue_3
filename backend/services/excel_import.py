"""
Servizio di import Excel.

Strategia (concordata):
  - Schema sheets (Motivations/Parameters/Questions/QuestionAllowedMotivations):
    "Strict update". Ogni riga è un upsert vincolato: se l'id esiste in DB,
    rimpiazzo TUTTI i campi col valore del file. Se l'id NON esiste, è un errore
    (riga saltata, log nel report). Nessun delete delle entità non menzionate.

  - Database_model (compilation della singola lingua):
    Replace totale. Cancello tutte le risposte/esempi/motivazioni della lingua,
    poi inserisco solo le righe valide. Le righe errate vengono saltate, e la
    domanda corrispondente resta visibile come "non risposta" → l'admin sa
    esattamente cosa è andato storto guardando il report.

  - Errori a cascata espliciti, mai silenziosi: se Motivation X fallisce e una
    QuestionAllowedMotivation la referenzia, la QAM va nel report con motivo
    "Motivation X non disponibile (errore upstream)".

  - Savepoint per riga: un errore in una riga non rovina le righe successive.
"""
from __future__ import annotations
from dataclasses import dataclass, field, asdict
from typing import List, Optional, Dict, Set, Any, Tuple
from datetime import datetime
import io

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError, DataError

import models
from services.logic_parser import validate_expression, ParseException
from services.versioning import record_version
from services.language_alias import resolve_language


# ============================================================================
# Data classes per il report
# ============================================================================

@dataclass
class ImportError:
    sheet: str
    row: int                      # numero di riga 1-based (la 1 è l'header)
    column: Optional[str] = None
    value: Optional[str] = None
    reason: str = ""

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class SheetSummary:
    rows_total: int = 0
    updated: int = 0
    inserted: int = 0
    skipped: int = 0
    errors: int = 0

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class ImportReport:
    sheets_processed: List[str] = field(default_factory=list)
    by_sheet: Dict[str, SheetSummary] = field(default_factory=dict)
    errors: List[ImportError] = field(default_factory=list)
    target_language_id: Optional[str] = None
    target_language_name: Optional[str] = None

    def to_dict(self) -> dict:
        return {
            "sheets_processed": self.sheets_processed,
            "by_sheet": {k: v.to_dict() for k, v in self.by_sheet.items()},
            "errors": [e.to_dict() for e in self.errors],
            "target_language_id": self.target_language_id,
            "target_language_name": self.target_language_name,
            "total_errors": len(self.errors),
        }


# ============================================================================
# Helper di parsing celle
# ============================================================================

SCHEMA_SHEETS = ("Motivations", "Parameters", "Questions", "QuestionAllowedMotivations")
COMPILATION_SHEET = "Database_model"

# Ordine di processing (rispetta le dipendenze: Questions→Parameters,
# QAM→Questions+Motivations, Database_model→Questions).
SUPPORTED_SHEET_TYPES = (
    "Motivations",
    "Parameters",
    "Questions",
    "QuestionAllowedMotivations",
    "Languages",
    "Glossary",
    COMPILATION_SHEET,
)

# Signature di header per ciascun tipo di sheet. Usate solo come fallback
# quando il nome della tab non matcha (es. tab lasciate come "Sheet1"/"Foglio1"
# dopo copia-incolla). Le signature sono set di colonne che DEVONO essere tutte
# presenti nella riga 1; sono scelte per essere distintive e non collidere fra
# tipi diversi.
SHEET_SIGNATURES: Dict[str, Set[str]] = {
    # 4 colonne tutte specifiche, zero collisioni possibili
    COMPILATION_SHEET: {"Language", "Parameter_Label", "Question_ID", "Language_Answer"},
    # "Code" da solo non appare in nessun altro sheet (QAM usa "Motivation Code")
    "Motivations": {"Code"},
    # "Schema" è esclusivo di Parameters (distingue da Questions che ha "ID"+"Text")
    "Parameters": {"ID", "Schema"},
    # "Parameter ID" distingue da Parameters; "Text" da Motivations
    "Questions": {"ID", "Parameter ID", "Text"},
    "QuestionAllowedMotivations": {"Question ID", "Motivation Code"},
    # "ISO code" evita collisione con sheet generici che hanno "ID"+"Name"
    "Languages": {"ID", "Name", "ISO code"},
    # "Word" non appare in nessun altro sheet
    "Glossary": {"Word", "Description"},
}


def _str(v: Any) -> str:
    """Cella → stringa trim. None/vuoto → ''."""
    if v is None:
        return ""
    return str(v).strip()


def _bool_yn(v: Any) -> bool:
    """'Yes'/'No' → bool. Default False."""
    s = _str(v).lower()
    return s in ("yes", "y", "true", "1", "x")


def _none_if_empty(v: Any) -> Optional[str]:
    s = _str(v)
    return s if s else None


def _build_header_map(ws: Worksheet) -> Dict[str, int]:
    """Mappa nome colonna → indice (0-based) leggendo la riga 1."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {(_str(h)): i for i, h in enumerate(header_row) if _str(h)}


def _get(row: Tuple, header_map: Dict[str, int], col_name: str) -> Any:
    """Estrae il valore della colonna col_name dalla riga. None se mancante."""
    idx = header_map.get(col_name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


# ============================================================================
# Sheet detection: nome esatto + fallback per header signature
# ============================================================================

def _detect_sheet_type(ws: Worksheet) -> Optional[str]:
    """Identifica il tipo di sheet leggendo la riga 1 (header).
    Ritorna il primo tipo la cui SHEET_SIGNATURES è interamente contenuta
    nelle colonne presenti. None se nessun match. I confronti sono
    case-sensitive: gli utenti devono scrivere gli header esatti."""
    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return None
    headers = {_str(h) for h in header_row if _str(h)}
    if not headers:
        return None
    for sheet_type in SUPPORTED_SHEET_TYPES:
        signature = SHEET_SIGNATURES.get(sheet_type)
        if signature and signature.issubset(headers):
            return sheet_type
    return None


def _resolve_sheets(wb) -> Dict[str, Worksheet]:
    """Mappa tipo→worksheet da processare.

    Priorità:
      1. Match per **nome esatto** della tab (comportamento storico, sempre
         vince).
      2. Fallback per **header signature** sulle tab rimanenti — utile per
         file in cui la tab è rimasta "Sheet1"/"Foglio1" o ha il nome della
         lingua. Se più tab anonime matchano lo stesso tipo, vince la prima
         in ordine di workbook.
    """
    resolved: Dict[str, Worksheet] = {}
    for sheet_type in SUPPORTED_SHEET_TYPES:
        if sheet_type in wb.sheetnames:
            resolved[sheet_type] = wb[sheet_type]

    for ws in wb.worksheets:
        if ws.title in SUPPORTED_SHEET_TYPES:
            continue  # tab presa per nome esatto (anche se non era resolved)
        detected = _detect_sheet_type(ws)
        if detected and detected not in resolved:
            resolved[detected] = ws

    return resolved


# ============================================================================
# Main entry point
# ============================================================================

def import_excel(
    db: Session,
    file_bytes: bytes,
    current_user_id: int,
    *,
    create_missing: bool = False,
) -> ImportReport:
    """
    Punto di ingresso. Apre il file, riconosce i sheet presenti, processa
    in ordine di dipendenza, ritorna un ImportReport completo.

    `create_missing`: se True, gli importer schema (Motivations / Parameters /
    Questions) creano la entità invece di errorare quando l'ID non esiste.
    Pensato per il backup-restore (dove dopo wipe schema è vuoto). Default
    False mantiene il comportamento "strict update" per gli upload manuali da
    UI, che non devono creare schema accidentalmente.
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        report = ImportReport()
        report.errors.append(ImportError(
            sheet="(file)", row=0, reason=f"File non leggibile: {e}"
        ))
        return report

    report = ImportReport()

    # ID falliti che bloccano dipendenze a cascata
    failed_motivation_codes: Set[str] = set()
    failed_parameter_ids: Set[str] = set()
    failed_question_ids: Set[str] = set()

    # Risoluzione tab: nome esatto + fallback per header signature.
    sheets = _resolve_sheets(wb)

    def _run(sheet_type: str, fn) -> None:
        ws = sheets.get(sheet_type)
        if ws is None:
            return
        fn(ws)
        try:
            db.commit()
        except Exception as e:
            db.rollback()
            report.errors.append(ImportError(sheet=sheet_type, row=0, reason=f"Commit fallito: {e}"))

    _run("Motivations", lambda ws: _import_motivations(
        db, ws, report, failed_motivation_codes,
        user_id=current_user_id, create_missing=create_missing))

    _run("Parameters", lambda ws: _import_parameters(
        db, ws, report, current_user_id, failed_parameter_ids,
        create_missing=create_missing))

    _run("Questions", lambda ws: _import_questions(
        db, ws, report, current_user_id,
        failed_parameter_ids, failed_question_ids,
        create_missing=create_missing))

    _run("QuestionAllowedMotivations", lambda ws: _import_qam(
        db, ws, report, failed_motivation_codes, failed_question_ids))

    # "Languages" e "Glossary" sono presenti nei file del backup-zip
    # (languages_metadata.xlsx, glossary.xlsx). Li gestiamo qui così l'import
    # totale può semplicemente alimentare ogni xlsx del bundle a import_excel.
    _run("Languages", lambda ws: _import_languages_metadata(db, ws, report))
    _run("Glossary", lambda ws: _import_glossary(db, ws, report))

    _run(COMPILATION_SHEET, lambda ws: _import_compilation(
        db, ws, report, failed_question_ids))

    return report


# ============================================================================
# Helper: tenta l'esecuzione di una funzione su un savepoint per riga
# ============================================================================

def _safe_apply(db: Session, fn) -> Tuple[bool, Optional[str]]:
    """
    Esegue fn() in un savepoint. Se fallisce con errore di DB, rollback al
    savepoint e ritorna (False, msg). Altrimenti (True, None).
    """
    sp = db.begin_nested()
    try:
        fn()
        db.flush()
        sp.commit()
        return True, None
    except (IntegrityError, DataError) as e:
        sp.rollback()
        return False, _format_db_error(e)
    except Exception as e:
        sp.rollback()
        return False, str(e)


def _format_db_error(e: Exception) -> str:
    msg = str(getattr(e, "orig", e))
    if len(msg) > 200:
        msg = msg[:200] + "…"
    return msg


# ============================================================================
# 1. MOTIVATIONS — strict update
# ============================================================================

def _import_motivations(db: Session, ws: Worksheet, report: ImportReport,
                        failed_codes: Set[str], user_id: Optional[int] = None,
                        *, create_missing: bool = False) -> None:
    summary = SheetSummary()
    report.sheets_processed.append("Motivations")
    report.by_sheet["Motivations"] = summary

    hmap = _build_header_map(ws)
    if "Code" not in hmap:
        report.errors.append(ImportError(sheet="Motivations", row=1,
                                         reason="Colonna 'Code' mancante"))
        return

    # Pre-load tutte le motivations per code (chiave normalizzata in upper-case
    # per match case-insensitive sul file Excel; il code DB resta canonico).
    by_code = {m.code.upper(): m for m in db.query(models.Motivation).all()}

    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _str(v) == "" for v in row):
            continue
        summary.rows_total += 1

        code = _str(_get(row, hmap, "Code"))
        if not code:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet="Motivations", row=ridx, column="Code",
                reason="Codice vuoto, riga saltata"
            ))
            continue

        code_key = code.upper()
        existing = by_code.get(code_key)
        label = _str(_get(row, hmap, "Label"))

        if not existing:
            if not create_missing:
                summary.errors += 1
                failed_codes.add(code_key)
                report.errors.append(ImportError(
                    sheet="Motivations", row=ridx, column="Code", value=code,
                    reason=f"Motivation '{code}' does not exist in the DB. Create it via the UI before importing."
                ))
                continue
            # create_missing path
            def apply_create():
                m = models.Motivation(code=code, label=label or "")
                db.add(m)
                db.flush()
                by_code[code_key] = m

            ok, err = _safe_apply(db, apply_create)
            if ok:
                summary.inserted += 1
                record_version(db, by_code[code_key], operation="create",
                               source="backup_restore", user_id=user_id,
                               note="Backup restore")
            else:
                summary.errors += 1
                failed_codes.add(code_key)
                report.errors.append(ImportError(
                    sheet="Motivations", row=ridx, value=code, reason=err
                ))
            continue

        def apply():
            existing.label = label or existing.label

        ok, err = _safe_apply(db, apply)
        if ok:
            summary.updated += 1
            record_version(db, existing, operation="update", source="excel_import",
                           user_id=user_id, note="Import Excel")
        else:
            summary.errors += 1
            failed_codes.add(code_key)
            report.errors.append(ImportError(
                sheet="Motivations", row=ridx, value=code, reason=err
            ))


# ============================================================================
# 2. PARAMETERS — strict update + ParameterChangeLog
# ============================================================================

PARAM_FIELDS = (
    ("Name", "name", _str),
    ("Schema", "schema", _str),
    ("Type", "param_type", _str),
    ("Level", "level_of_comparison", _str),
    ("Short Description", "short_description", _str),
    ("Long Description", "long_description", _str),
    ("Implicational Condition", "implicational_condition", _none_if_empty),
    ("Explanation of Implicational Condition", "description_of_the_implicational_condition", _str),
)


def _import_parameters(db: Session, ws: Worksheet, report: ImportReport,
                       user_id: int, failed_ids: Set[str],
                       *, create_missing: bool = False) -> None:
    summary = SheetSummary()
    report.sheets_processed.append("Parameters")
    report.by_sheet["Parameters"] = summary

    hmap = _build_header_map(ws)
    if "ID" not in hmap:
        report.errors.append(ImportError(sheet="Parameters", row=1,
                                         reason="Missing 'ID' column"))
        return

    # Chiave normalizzata upper-case per match case-insensitive sul file Excel;
    # l'ID DB resta canonico.
    by_id = {p.id.upper(): p for p in db.query(models.ParameterDef).all()}

    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _str(v) == "" for v in row):
            continue
        summary.rows_total += 1

        pid = _str(_get(row, hmap, "ID"))
        if not pid:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet="Parameters", row=ridx, column="ID",
                reason="Empty ID, row skipped"
            ))
            continue

        pid_key = pid.upper()
        existing = by_id.get(pid_key)

        # Validazione condition (se presente). La facciamo prima del branch
        # create-vs-update così si applica anche al nuovo parametro.
        cond_raw = _none_if_empty(_get(row, hmap, "Implicational Condition"))
        if cond_raw:
            try:
                validate_expression(cond_raw)
            except ParseException as e:
                summary.errors += 1
                failed_ids.add(pid_key)
                report.errors.append(ImportError(
                    sheet="Parameters", row=ridx,
                    column="Implicational Condition", value=cond_raw,
                    reason=f"Wrong formula syntax: {e}"
                ))
                continue

        # CREATE branch (solo se create_missing=True e parametro non esiste)
        if existing is None:
            if not create_missing:
                summary.errors += 1
                failed_ids.add(pid_key)
                report.errors.append(ImportError(
                    sheet="Parameters", row=ridx, column="ID", value=pid,
                    reason=f"Parameter '{pid}' does not exist in the DB. Create it via the UI before importing."
                ))
                continue

            new_position = _get(row, hmap, "Position")
            try:
                new_position = int(new_position) if new_position not in (None, "") else 0
            except (TypeError, ValueError):
                new_position = 0
            new_is_active = _bool_yn(_get(row, hmap, "Is Active"))

            def apply_create():
                kwargs = {"id": pid, "position": new_position, "is_active": new_is_active}
                for col_name, attr_name, parser in PARAM_FIELDS:
                    kwargs[attr_name] = parser(_get(row, hmap, col_name))
                p = models.ParameterDef(**kwargs)
                db.add(p)
                db.flush()
                by_id[pid_key] = p

            ok, err = _safe_apply(db, apply_create)
            if ok:
                summary.inserted += 1
                record_version(db, by_id[pid_key], operation="create",
                               source="backup_restore", user_id=user_id,
                               note="Backup restore")
            else:
                summary.errors += 1
                failed_ids.add(pid_key)
                report.errors.append(ImportError(
                    sheet="Parameters", row=ridx, value=pid, reason=err
                ))
            continue

        # UPDATE branch (parametro esistente)
        old_snapshot = {f[1]: getattr(existing, f[1]) for f in PARAM_FIELDS}
        old_position = existing.position
        old_is_active = existing.is_active

        new_position = _get(row, hmap, "Position")
        try:
            new_position = int(new_position) if new_position not in (None, "") else existing.position
        except (TypeError, ValueError):
            new_position = existing.position

        new_is_active = _bool_yn(_get(row, hmap, "Is Active"))

        def apply():
            for col_name, attr_name, parser in PARAM_FIELDS:
                val = _get(row, hmap, col_name)
                setattr(existing, attr_name, parser(val))
            existing.position = new_position
            existing.is_active = new_is_active

        ok, err = _safe_apply(db, apply)
        if not ok:
            summary.errors += 1
            failed_ids.add(pid_key)
            report.errors.append(ImportError(
                sheet="Parameters", row=ridx, value=pid, reason=err
            ))
            continue

        # Diff per ChangeLog
        diff_parts = []
        for f in PARAM_FIELDS:
            attr = f[1]
            if old_snapshot[attr] != getattr(existing, attr):
                diff_parts.append(attr)
        if old_position != new_position:
            diff_parts.append("position")
        if old_is_active != new_is_active:
            diff_parts.append("is_active")

        if diff_parts:
            log = models.ParameterChangeLog(
                parameter_id=existing.id, user_id=user_id,
                change_note=f"[Excel import] Updated: {', '.join(diff_parts)}"
            )
            db.add(log)

        record_version(db, existing, operation="update", source="excel_import",
                       user_id=user_id,
                       note=f"Import Excel ({', '.join(diff_parts) or 'no changes'})")
        summary.updated += 1


# ============================================================================
# 3. QUESTIONS — strict update + ParameterChangeLog sul parent param
# ============================================================================

QUESTION_FIELDS = (
    ("Text", "text", _str),
    ("Template Type", "template_type", _str),
    ("Instruction", "instruction", _none_if_empty),
    ("Instruction YES", "instruction_yes", _none_if_empty),
    ("Instruction NO", "instruction_no", _none_if_empty),
    ("Example YES", "example_yes", _none_if_empty),
    ("Help Info", "help_info", _none_if_empty),
)


def _import_questions(db: Session, ws: Worksheet, report: ImportReport,
                      user_id: int, failed_param_ids: Set[str],
                      failed_question_ids: Set[str],
                      *, create_missing: bool = False) -> None:
    summary = SheetSummary()
    report.sheets_processed.append("Questions")
    report.by_sheet["Questions"] = summary

    hmap = _build_header_map(ws)
    if "ID" not in hmap:
        report.errors.append(ImportError(sheet="Questions", row=1,
                                         reason="Missing 'ID' column"))
        return

    # Chiavi normalizzate upper-case per match case-insensitive sul file Excel;
    # gli ID DB restano canonici e vengono usati per le FK.
    by_id = {q.id.upper(): q for q in db.query(models.Question).all()}
    param_id_by_upper = {p.id.upper(): p.id for p in db.query(models.ParameterDef.id).all()}

    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _str(v) == "" for v in row):
            continue
        summary.rows_total += 1

        qid = _str(_get(row, hmap, "ID"))
        if not qid:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet="Questions", row=ridx, column="ID",
                reason="Empty ID, row skipped"
            ))
            continue

        qid_key = qid.upper()
        existing = by_id.get(qid_key)
        new_param_id = _str(_get(row, hmap, "Parameter ID"))
        new_param_id_key = new_param_id.upper() if new_param_id else ""
        # Risolvi l'ID canonico del parametro (case-insensitive); None se non esiste.
        canonical_new_param_id = param_id_by_upper.get(new_param_id_key) if new_param_id else None

        # CREATE branch (solo se create_missing=True e domanda non esiste)
        if existing is None:
            if not create_missing:
                summary.errors += 1
                failed_question_ids.add(qid_key)
                report.errors.append(ImportError(
                    sheet="Questions", row=ridx, column="ID", value=qid,
                    reason=f"Question '{qid}' does not exist in the DB. Create it via the UI before importing."
                ))
                continue
            if not new_param_id:
                summary.errors += 1
                failed_question_ids.add(qid_key)
                report.errors.append(ImportError(
                    sheet="Questions", row=ridx, column="Parameter ID", value=qid,
                    reason="Empty Parameter ID for new question"
                ))
                continue
            if canonical_new_param_id is None:
                summary.errors += 1
                failed_question_ids.add(qid_key)
                report.errors.append(ImportError(
                    sheet="Questions", row=ridx, column="Parameter ID", value=new_param_id,
                    reason=f"Parameter '{new_param_id}' does not exist"
                ))
                continue

            new_stop = _bool_yn(_get(row, hmap, "Is Stop Question"))
            new_active = _bool_yn(_get(row, hmap, "Is Active"))

            def apply_create():
                kwargs = {
                    "id": qid, "parameter_id": canonical_new_param_id,
                    "is_stop_question": new_stop, "is_active": new_active,
                }
                for col_name, attr_name, parser in QUESTION_FIELDS:
                    kwargs[attr_name] = parser(_get(row, hmap, col_name))
                q = models.Question(**kwargs)
                db.add(q)
                db.flush()
                by_id[qid_key] = q

            ok, err = _safe_apply(db, apply_create)
            if ok:
                summary.inserted += 1
                record_version(db, by_id[qid_key], operation="create",
                               source="backup_restore", user_id=user_id,
                               note="Backup restore")
            else:
                summary.errors += 1
                failed_question_ids.add(qid_key)
                report.errors.append(ImportError(
                    sheet="Questions", row=ridx, value=qid, reason=err
                ))
            continue

        # Cambio parent? Confronto case-insensitive contro l'ID corrente del param.
        parent_changing = bool(new_param_id) and new_param_id_key != existing.parameter_id.upper()
        if parent_changing:
            if new_param_id_key in failed_param_ids:
                summary.errors += 1
                failed_question_ids.add(qid_key)
                report.errors.append(ImportError(
                    sheet="Questions", row=ridx, column="Parameter ID", value=new_param_id,
                    reason=f"Parameter '{new_param_id}' failed during import (upstream error)"
                ))
                continue
            if canonical_new_param_id is None:
                summary.errors += 1
                failed_question_ids.add(qid_key)
                report.errors.append(ImportError(
                    sheet="Questions", row=ridx, column="Parameter ID", value=new_param_id,
                    reason=f"Parameter '{new_param_id}' does not exist"
                ))
                continue

        old_param_id = existing.parameter_id
        old_snapshot = {f[1]: getattr(existing, f[1]) for f in QUESTION_FIELDS}
        old_stop = existing.is_stop_question
        old_active = existing.is_active

        new_stop = _bool_yn(_get(row, hmap, "Is Stop Question"))
        new_active = _bool_yn(_get(row, hmap, "Is Active"))

        def apply():
            if parent_changing and canonical_new_param_id is not None:
                existing.parameter_id = canonical_new_param_id
            for col_name, attr_name, parser in QUESTION_FIELDS:
                setattr(existing, attr_name, parser(_get(row, hmap, col_name)))
            existing.is_stop_question = new_stop
            existing.is_active = new_active

        ok, err = _safe_apply(db, apply)
        if not ok:
            summary.errors += 1
            failed_question_ids.add(qid_key)
            report.errors.append(ImportError(
                sheet="Questions", row=ridx, value=qid, reason=err
            ))
            continue

        # Diff log nel parent param
        diff_parts = []
        for f in QUESTION_FIELDS:
            if old_snapshot[f[1]] != getattr(existing, f[1]):
                diff_parts.append(f[1])
        if old_stop != new_stop:
            diff_parts.append("is_stop_question")
        if old_active != new_active:
            diff_parts.append("is_active")
        if old_param_id != existing.parameter_id:
            diff_parts.append("parameter_id")

        if diff_parts:
            log = models.ParameterChangeLog(
                parameter_id=existing.parameter_id, user_id=user_id,
                change_note=f"[Excel import] [Question {qid}] Updated: {', '.join(diff_parts)}"
            )
            db.add(log)

        record_version(db, existing, operation="update", source="excel_import",
                       user_id=user_id,
                       note=f"Import Excel ({', '.join(diff_parts) or 'no changes'})")
        summary.updated += 1


# ============================================================================
# 4. QUESTION_ALLOWED_MOTIVATIONS — replace dei link per le coppie nel file
# ============================================================================

def _import_qam(db: Session, ws: Worksheet, report: ImportReport,
                failed_motivation_codes: Set[str],
                failed_question_ids: Set[str]) -> None:
    summary = SheetSummary()
    report.sheets_processed.append("QuestionAllowedMotivations")
    report.by_sheet["QuestionAllowedMotivations"] = summary

    hmap = _build_header_map(ws)
    if "Question ID" not in hmap or "Motivation Code" not in hmap:
        report.errors.append(ImportError(
            sheet="QuestionAllowedMotivations", row=1,
            reason="Required columns: 'Question ID', 'Motivation Code'"
        ))
        return

    # Chiavi normalizzate upper-case per match case-insensitive sul file Excel;
    # gli ID DB restano canonici e vengono usati per le FK.
    by_qid = {q.id.upper(): q for q in db.query(models.Question).all()}
    by_code = {m.code.upper(): m for m in db.query(models.Motivation).all()}

    # Strategia: per ogni question_id presente nel file, cancello i link esistenti
    # e li ri-creo dal file. Questo evita di accumulare link orfani.
    questions_seen: Set[str] = set()
    pairs_to_create: List[Tuple[str, int]] = []

    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _str(v) == "" for v in row):
            continue
        summary.rows_total += 1

        qid = _str(_get(row, hmap, "Question ID"))
        code = _str(_get(row, hmap, "Motivation Code"))

        if not qid or not code:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet="QuestionAllowedMotivations", row=ridx,
                reason="Empty Question ID or Motivation Code"
            ))
            continue

        qid_key = qid.upper()
        code_key = code.upper()

        if qid_key in failed_question_ids:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet="QuestionAllowedMotivations", row=ridx,
                column="Question ID", value=qid,
                reason=f"Question '{qid}' failed during import (upstream error)"
            ))
            continue
        question_db = by_qid.get(qid_key)
        if question_db is None:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet="QuestionAllowedMotivations", row=ridx,
                column="Question ID", value=qid,
                reason=f"Question '{qid}' does not exist"
            ))
            continue

        if code_key in failed_motivation_codes:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet="QuestionAllowedMotivations", row=ridx,
                column="Motivation Code", value=code,
                reason=f"Motivation '{code}' failed during import (upstream error)"
            ))
            continue
        motivation_db = by_code.get(code_key)
        if motivation_db is None:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet="QuestionAllowedMotivations", row=ridx,
                column="Motivation Code", value=code,
                reason=f"Motivation '{code}' does not exist"
            ))
            continue

        questions_seen.add(question_db.id)
        pairs_to_create.append((question_db.id, motivation_db.id))

    # Replace dei link per le sole question viste nel file
    if questions_seen:
        db.query(models.QuestionAllowedMotivation).filter(
            models.QuestionAllowedMotivation.question_id.in_(questions_seen)
        ).delete(synchronize_session=False)
        db.flush()

        for qid, mid in pairs_to_create:
            db.add(models.QuestionAllowedMotivation(question_id=qid, motivation_id=mid))
            summary.inserted += 1


# ============================================================================
# 5. DATABASE_MODEL — replace della compilation della singola lingua
# ============================================================================

def _split_lines(v: Any) -> List[str]:
    s = _str(v)
    if not s:
        return []
    return [line.strip() for line in s.replace("\r\n", "\n").replace("\r", "\n").split("\n")]


def _import_compilation(db: Session, ws: Worksheet, report: ImportReport,
                        failed_question_ids: Set[str]) -> None:
    summary = SheetSummary()
    report.sheets_processed.append(COMPILATION_SHEET)
    report.by_sheet[COMPILATION_SHEET] = summary

    hmap = _build_header_map(ws)
    required = ["Language", "Parameter_Label", "Question_ID", "Language_Answer"]
    missing = [c for c in required if c not in hmap]
    if missing:
        report.errors.append(ImportError(
            sheet=COMPILATION_SHEET, row=1,
            reason=f"Missing columns: {', '.join(missing)}"
        ))
        return

    # 1. Identificazione lingua dal valore "Language" (deve essere unico)
    lang_values: Set[str] = set()
    rows = []
    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _str(v) == "" for v in row):
            continue
        rows.append((ridx, row))
        lv = _str(_get(row, hmap, "Language"))
        if lv:
            lang_values.add(lv)

    if not rows:
        return
    if len(lang_values) != 1:
        report.errors.append(ImportError(
            sheet=COMPILATION_SHEET, row=0,
            reason=f"The file must contain a single language. Found: {sorted(lang_values)}"
        ))
        return

    lang_name = next(iter(lang_values))
    lang = db.query(models.Language).filter(
        models.Language.name_full == lang_name
    ).first()
    if not lang:
        report.errors.append(ImportError(
            sheet=COMPILATION_SHEET, row=0, value=lang_name,
            reason=f"Language '{lang_name}' not found (search by name_full)"
        ))
        return

    report.target_language_id = lang.id
    report.target_language_name = lang.name_full

    # 2. REPLACE TOTALE: cancella tutte le risposte della lingua + esempi e motivazioni.
    # Cancellazione esplicita perché db.query().delete() bulk NON triggera il
    # cascade ORM, e ondelete=CASCADE a livello DB non è uniforme (SQLite di test
    # vs Postgres prod). Approccio sicuro su entrambi.
    old_answer_ids = [
        a_id for (a_id,) in db.query(models.Answer.id).filter(
            models.Answer.language_id == lang.id
        ).all()
    ]
    if old_answer_ids:
        db.query(models.Example).filter(
            models.Example.answer_id.in_(old_answer_ids)
        ).delete(synchronize_session=False)
        db.query(models.AnswerMotivation).filter(
            models.AnswerMotivation.answer_id.in_(old_answer_ids)
        ).delete(synchronize_session=False)
        db.query(models.Answer).filter(
            models.Answer.id.in_(old_answer_ids)
        ).delete(synchronize_session=False)
    # Reset delle admin notes per questa lingua. Settiamo a None senza cancellare
    # la riga di LanguageParameterStatus, così preserviamo `is_unsure`. Le note
    # verranno re-applicate dal file in fondo all'import.
    db.query(models.LanguageParameterStatus).filter(
        models.LanguageParameterStatus.language_id == lang.id
    ).update({"admin_note": None}, synchronize_session=False)
    db.flush()

    # 3. Pre-load di tutte le question + motivations per id e per code.
    # Chiavi normalizzate upper-case per match case-insensitive sul file Excel;
    # gli ID DB restano canonici (servono come FK su Answer.question_id ecc.).
    q_id_by_upper = {q.id.upper(): q.id for q in db.query(models.Question.id).all()}
    mot_by_code = {m.code.upper(): m for m in db.query(models.Motivation).all()}
    param_id_by_upper = {p.id.upper(): p.id for p in db.query(models.ParameterDef.id).all()}

    # Admin notes da applicare a fine import. Una riga per parametro.
    # Se la stessa Admin_Note compare su più question dello stesso parametro
    # (caso normale dell'export, che la duplica), l'ultima vince — sono uguali.
    admin_notes_by_pid: dict[str, str] = {}

    # 4. Per ogni riga del file, tenta l'inserimento
    for ridx, row in rows:
        summary.rows_total += 1

        qid = _str(_get(row, hmap, "Question_ID"))
        if not qid:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet=COMPILATION_SHEET, row=ridx, column="Question_ID",
                reason="Empty Question_ID"
            ))
            continue

        qid_key = qid.upper()
        if qid_key in failed_question_ids:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet=COMPILATION_SHEET, row=ridx, column="Question_ID", value=qid,
                reason=f"Question '{qid}' failed during import (upstream error)"
            ))
            continue
        canonical_qid = q_id_by_upper.get(qid_key)
        if canonical_qid is None:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet=COMPILATION_SHEET, row=ridx, column="Question_ID", value=qid,
                reason=f"Question '{qid}' does not exist"
            ))
            continue
        # Da qui in poi usa l'ID canonico DB per la FK Answer.question_id.
        qid = canonical_qid

        raw_ans = _str(_get(row, hmap, "Language_Answer")).upper()
        if raw_ans in ("YES", "Y"):
            response = "yes"
        elif raw_ans in ("NO", "N"):
            response = "no"
        elif raw_ans == "":
            # Risposta vuota → la domanda resta non risposta. Non è un errore.
            summary.skipped += 1
            continue
        else:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet=COMPILATION_SHEET, row=ridx, column="Language_Answer", value=raw_ans,
                reason=f"Invalid value (expected YES/NO/empty): '{raw_ans}'"
            ))
            continue

        comments = _str(_get(row, hmap, "Language_Comments"))

        # Esempi multilinea
        ex_texts = _split_lines(_get(row, hmap, "Language_Examples"))
        gloss_lines = _split_lines(_get(row, hmap, "Language_Example_Gloss"))
        transl_lines = _split_lines(_get(row, hmap, "Language_Example_Translation"))
        ref_lines = _split_lines(_get(row, hmap, "Language_References"))

        # Motivations (colonna opzionale, presente nei file dal 2026-05).
        # Codici separati da `;` o `,`. Codici sconosciuti: errore non bloccante,
        # la motivation viene saltata ma l'answer e le altre motivations valide
        # vengono comunque inserite.
        mot_codes_raw = _str(_get(row, hmap, "Motivations"))
        mot_codes_to_apply: list[int] = []
        if mot_codes_raw:
            for token in mot_codes_raw.replace(",", ";").split(";"):
                code = token.strip()
                if not code:
                    continue
                m = mot_by_code.get(code.upper())
                if m is None:
                    report.errors.append(ImportError(
                        sheet=COMPILATION_SHEET, row=ridx, column="Motivations", value=code,
                        reason=f"Motivation code '{code}' not found"
                    ))
                    continue
                mot_codes_to_apply.append(m.id)

        # Admin_Note: associata al parametro, non alla question. Accumula a
        # fine loop per applicarla una volta sola per parametro (usando l'ID
        # canonico DB, così la chiave del dict resta consistente anche se il
        # file scrive il Parameter_Label con casing diverso).
        note_cell = _str(_get(row, hmap, "Admin_Note"))
        param_label = _str(_get(row, hmap, "Parameter_Label"))
        canonical_pid = param_id_by_upper.get(param_label.upper()) if param_label else None
        if note_cell and canonical_pid is not None:
            admin_notes_by_pid[canonical_pid] = note_cell

        # Crea Answer + Examples + AnswerMotivations
        def apply(mot_ids=mot_codes_to_apply):
            answer = models.Answer(
                language_id=lang.id, question_id=qid,
                response_text=response, comments=comments or None,
                status="pending",
            )
            db.add(answer)
            db.flush()  # per ottenere answer.id

            for i, txt in enumerate(ex_texts):
                if not txt and not (
                    (gloss_lines[i] if i < len(gloss_lines) else "") or
                    (transl_lines[i] if i < len(transl_lines) else "") or
                    (ref_lines[i] if i < len(ref_lines) else "")
                ):
                    continue
                ex = models.Example(
                    answer_id=answer.id,
                    number=str(i + 1),
                    textarea=txt or None,
                    gloss=(gloss_lines[i] if i < len(gloss_lines) else None) or None,
                    translation=(transl_lines[i] if i < len(transl_lines) else None) or None,
                    reference=(ref_lines[i] if i < len(ref_lines) else None) or None,
                )
                db.add(ex)

            # Dedup difensivo: stesso codice ripetuto nella cella → 1 sola riga.
            for mid in set(mot_ids):
                db.add(models.AnswerMotivation(answer_id=answer.id, motivation_id=mid))

        ok, err = _safe_apply(db, apply)
        if ok:
            summary.inserted += 1
        else:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet=COMPILATION_SHEET, row=ridx, value=qid, reason=err
            ))

    # 5. Applica admin notes accumulate (uno-shot, una per parametro).
    for pid, note in admin_notes_by_pid.items():
        status = db.query(models.LanguageParameterStatus).filter(
            models.LanguageParameterStatus.language_id == lang.id,
            models.LanguageParameterStatus.parameter_id == pid,
        ).first()
        if status is None:
            db.add(models.LanguageParameterStatus(
                language_id=lang.id,
                parameter_id=pid,
                admin_note=note,
                is_unsure=False,
            ))
        else:
            status.admin_note = note


# ============================================================================
# 7. LANGUAGES METADATA — upsert per ID
# ============================================================================
#
# Foglio "Languages" prodotto da build_language_list_workbook (export "language
# metadata" e file `languages_metadata.xlsx` del backup-zip).
#
# Strategia: upsert per ID. Se la lingua esiste in DB → aggiorno i campi
# scrivibili. Se non esiste → la creo. Mai cancello lingue non menzionate.
# Campi NON ripristinati: assigned_user (richiede una lookup utenti per email,
# meglio gestita lato UI), submitted_at, reviewed_at, updated_at (auto da
# SQLAlchemy onupdate). I dati di compilazione (Answer/Example/Motivation)
# vengono dai per-lingua xlsx, non da qui.
# ============================================================================

def _bool_yn_or_none(v: Any) -> Optional[bool]:
    if v is None:
        return None
    s = _str(v).lower()
    if s in ("yes", "y", "true", "1"):
        return True
    if s in ("no", "n", "false", "0"):
        return False
    return None


def _float_or_none(v: Any) -> Optional[float]:
    if v is None or _str(v) == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


_LANGUAGE_VALID_STATUSES = {"pending", "waiting_for_approval", "approved", "rejected"}


def _import_languages_metadata(db: Session, ws: Worksheet, report: ImportReport) -> None:
    summary = SheetSummary()
    report.sheets_processed.append("Languages")
    report.by_sheet["Languages"] = summary

    hmap = _build_header_map(ws)
    if "ID" not in hmap or "Name" not in hmap:
        report.errors.append(ImportError(
            sheet="Languages", row=1,
            reason="Colonne 'ID' e 'Name' obbligatorie."
        ))
        return

    # Per i nuovi inserimenti la `position` viene calcolata progressivamente
    # come max(position attuale) + 1 al momento dell'inserimento (vedi apply()).

    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _str(v) == "" for v in row):
            continue
        summary.rows_total += 1

        lid = _str(_get(row, hmap, "ID"))
        name = _str(_get(row, hmap, "Name"))
        if not lid:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet="Languages", row=ridx, column="ID", reason="Empty ID"
            ))
            continue
        if not name:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet="Languages", row=ridx, column="Name", value=lid, reason="Empty Name"
            ))
            continue

        status_raw = _str(_get(row, hmap, "Status")).lower()
        status = status_raw if status_raw in _LANGUAGE_VALID_STATUSES else "pending"

        top_str = _str(_get(row, hmap, "Top-level family")) or ""
        fam_str = _str(_get(row, hmap, "Family")) or ""
        grp_str = _str(_get(row, hmap, "Group")) or ""

        # Reverse lookup stringa→FK, stesso pattern di resolve_taxonomy in
        # languages.py. Niente forward-propagation: la riga Excel è
        # source-of-truth per le stringhe, non riscriviamo "Family" usando
        # il nome del parent del Group. Se il nome non matcha nessuna entità
        # in /taxonomy l'FK resta NULL e la stringa appare in "unnormalized".
        top_obj = db.query(models.TopFamily).filter(models.TopFamily.name == top_str).first() if top_str else None
        fam_obj = db.query(models.Family).filter(models.Family.name == fam_str).first() if fam_str else None
        grp_obj = db.query(models.Group).filter(models.Group.name == grp_str).first() if grp_str else None

        fields = {
            "name_full": name,
            "top_level_family": top_str,
            "family": fam_str,
            "grp": grp_str,
            "top_family_id": top_obj.id if top_obj else None,
            "family_id": fam_obj.id if fam_obj else None,
            "group_id": grp_obj.id if grp_obj else None,
            "isocode": _str(_get(row, hmap, "ISO code")) or "",
            "glottocode": _str(_get(row, hmap, "Glottocode")) or "",
            "location": _str(_get(row, hmap, "Location")) or "",
            "latitude": _float_or_none(_get(row, hmap, "Latitude")),
            "longitude": _float_or_none(_get(row, hmap, "Longitude")),
            "supervisor": _str(_get(row, hmap, "Supervisor")) or "",
            "informant": _str(_get(row, hmap, "Informant")) or "",
            "historical_language": _bool_yn_or_none(_get(row, hmap, "Historical")) or False,
            "source": _str(_get(row, hmap, "Source")) or "",
            "status": status,
        }

        # Lookup con fallback su alias storici: se l'id del file non
        # corrisponde a una lingua corrente, prova a riconoscerlo come
        # vecchio id di una lingua rinominata via UI admin. In quel caso
        # aggiorna i campi della lingua esistente senza toccarne l'id.
        resolved = resolve_language(db, lid, file_glottocode=fields["glottocode"])
        if resolved.glottocode_mismatch:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet="Languages", row=ridx, value=lid,
                reason=resolved.glottocode_mismatch,
            ))
            continue
        existing = resolved.language

        def apply():
            if existing is None:
                last = db.query(models.Language).order_by(
                    models.Language.position.desc()
                ).first()
                pos = (last.position + 1) if last else 1
                lang = models.Language(id=lid, position=pos, **fields)
                db.add(lang)
            else:
                # NB: non tocchiamo `existing.id`. Se il match e' via alias
                # l'id corrente (post-rename) e' quello giusto, non quello
                # del file. Aggiorniamo solo i campi metadata.
                for k, v in fields.items():
                    setattr(existing, k, v)

        ok, err = _safe_apply(db, apply)
        if ok:
            if existing is None:
                summary.inserted += 1
            else:
                summary.updated += 1
        else:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet="Languages", row=ridx, value=lid, reason=err
            ))


# ============================================================================
# 8. GLOSSARY — upsert per word
# ============================================================================

def _import_glossary(db: Session, ws: Worksheet, report: ImportReport) -> None:
    summary = SheetSummary()
    report.sheets_processed.append("Glossary")
    report.by_sheet["Glossary"] = summary

    hmap = _build_header_map(ws)
    if "Word" not in hmap or "Description" not in hmap:
        report.errors.append(ImportError(
            sheet="Glossary", row=1,
            reason="Colonne 'Word' e 'Description' obbligatorie."
        ))
        return

    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _str(v) == "" for v in row):
            continue
        summary.rows_total += 1

        word = _str(_get(row, hmap, "Word"))
        desc = _str(_get(row, hmap, "Description"))
        if not word:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet="Glossary", row=ridx, column="Word", reason="Empty Word"
            ))
            continue
        if not desc:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet="Glossary", row=ridx, column="Description", value=word,
                reason="Empty Description"
            ))
            continue

        existing = db.query(models.Glossary).filter(models.Glossary.word == word).first()

        def apply():
            if existing is None:
                db.add(models.Glossary(word=word, description=desc))
            else:
                existing.description = desc

        ok, err = _safe_apply(db, apply)
        if ok:
            if existing is None:
                summary.inserted += 1
            else:
                summary.updated += 1
        else:
            summary.errors += 1
            report.errors.append(ImportError(
                sheet="Glossary", row=ridx, value=word, reason=err
            ))
