"""
Servizio archivio domande obsolete.

Quando una Question viene modificata in modo non compatibile con i dati
linguistici raccolti, l'admin puo' richiedere il "wipe": le Answer/Example/
AnswerMotivation della question vengono spostate in tabelle archive
(insieme a uno snapshot della question stessa) e cancellate dalle tabelle
attive. La Question resta viva con il nuovo testo, pronta a raccogliere
nuovi dati.

API principali:
  - archive_and_wipe(db, question, user_id, archive_note) -> ArchivedQuestion
  - count_linked_data(db, question_id) -> dict {answers, examples, languages}
  - build_archived_question_workbook(db, archived_question) -> Workbook
"""
from __future__ import annotations
from typing import Dict
from datetime import datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import Session
from sqlalchemy import func

import models


# ============================================================================
# STATS: contatore per la conferma pre-wipe (mostrato in UI)
# ============================================================================

def count_linked_data(db: Session, question_id: str) -> Dict[str, int]:
    """Quante Answer/Example/lingue sarebbero archiviate dal wipe."""
    answers_count = (
        db.query(func.count(models.Answer.id))
        .filter(models.Answer.question_id == question_id)
        .scalar()
        or 0
    )
    languages_count = (
        db.query(func.count(func.distinct(models.Answer.language_id)))
        .filter(models.Answer.question_id == question_id)
        .scalar()
        or 0
    )
    examples_count = (
        db.query(func.count(models.Example.id))
        .join(models.Answer, models.Example.answer_id == models.Answer.id)
        .filter(models.Answer.question_id == question_id)
        .scalar()
        or 0
    )
    return {
        "answers": int(answers_count),
        "examples": int(examples_count),
        "languages": int(languages_count),
    }


# ============================================================================
# WIPE: snapshot + sposta dati nell'archivio + cancella dai tavoli attivi
# ============================================================================

def archive_and_wipe(
    db: Session,
    question: models.Question,
    user_id: int | None,
    archive_note: str | None = None,
) -> models.ArchivedQuestion:
    """
    Crea una ArchivedQuestion con lo snapshot della question PRIMA della
    modifica (chi chiama deve invocarla *prima* di applicare le modifiche
    al testo), copia tutte le Answer/Example/AnswerMotivation collegate
    nelle tabelle archive, poi cancella quelle attive.

    NON committa: chi chiama gestisce la transazione.
    """
    # Lookup nome del parametro per lo snapshot (denormalizzato).
    param = (
        db.query(models.ParameterDef)
        .filter(models.ParameterDef.id == question.parameter_id)
        .first()
    )
    param_name = param.name if param else ""

    # Crea il record archive con lo snapshot della question.
    archived = models.ArchivedQuestion(
        original_question_id=question.id,
        parameter_id=question.parameter_id,
        parameter_name=param_name,
        text=question.text or "",
        template_type=question.template_type or "",
        instruction=question.instruction,
        instruction_yes=question.instruction_yes,
        instruction_no=question.instruction_no,
        example_yes=question.example_yes,
        help_info=question.help_info,
        is_stop_question=bool(question.is_stop_question),
        is_active=bool(question.is_active),
        archived_at=datetime.utcnow(),
        archived_by_id=user_id,
        archive_note=(archive_note or "").strip(),
    )
    db.add(archived)
    db.flush()

    # Allowed motivations snapshot (code + label congelati).
    allowed = (
        db.query(models.QuestionAllowedMotivation, models.Motivation)
        .join(models.Motivation, models.QuestionAllowedMotivation.motivation_id == models.Motivation.id)
        .filter(models.QuestionAllowedMotivation.question_id == question.id)
        .all()
    )
    for _qam, mot in allowed:
        db.add(models.ArchivedQuestionMotivation(
            archived_question_id=archived.id,
            motivation_code=mot.code or "",
            motivation_label=mot.label or "",
        ))

    # Lookup denormalizzato del nome lingua per ogni Answer.
    lang_name_by_id: Dict[str, str] = {
        l.id: l.name_full
        for l in db.query(models.Language.id, models.Language.name_full).all()
    }

    # Copia tutte le Answer + Example + AnswerMotivation collegate.
    answers = (
        db.query(models.Answer)
        .filter(models.Answer.question_id == question.id)
        .all()
    )
    answers_count = 0
    examples_count = 0
    for a in answers:
        arch_a = models.ArchivedAnswer(
            archived_question_id=archived.id,
            language_id=a.language_id,
            language_name_full=lang_name_by_id.get(a.language_id, "") or "",
            status=a.status,
            response_text=a.response_text,
            comments=a.comments,
            original_updated_at=a.updated_at,
        )
        db.add(arch_a)
        db.flush()
        answers_count += 1

        for ex in a.examples:
            db.add(models.ArchivedExample(
                archived_answer_id=arch_a.id,
                number=ex.number or "",
                textarea=ex.textarea,
                transliteration=ex.transliteration,
                gloss=ex.gloss,
                translation=ex.translation,
                reference=ex.reference,
            ))
            examples_count += 1

        for am in a.answer_motivations:
            mot = am.motivation
            db.add(models.ArchivedAnswerMotivation(
                archived_answer_id=arch_a.id,
                motivation_code=(mot.code if mot else "") or "",
                motivation_label=(mot.label if mot else "") or "",
            ))

    archived.answers_count = answers_count
    archived.examples_count = examples_count

    # Cancella i dati attivi. Le cascate "all, delete-orphan" su Answer.examples
    # e Answer.answer_motivations gestiscono Example/AnswerMotivation.
    for a in answers:
        db.delete(a)

    db.flush()
    return archived


# ============================================================================
# EXPORT XLSX (singolo "Database_model" come da richiesta linguisti)
# ============================================================================

ARCHIVED_DB_HEADERS = [
    "Language",
    "Language ID",
    "Parameter ID",
    "Parameter Name",
    "Question ID",
    "Question Text (archived)",
    "Question Examples YES (archived)",
    "Question Instructions (archived)",
    "Language Answer",
    "Language Comments",
    "Language Motivations",
    "Language Examples",
    "Language Example Transliteration",
    "Language Example Gloss",
    "Language Example Translation",
    "Language References",
]

_BOLD_WHITE = Font(bold=True, color="FFFFFF")


def _bold_header_row(ws, n_cols: int):
    for i in range(1, n_cols + 1):
        ws.cell(row=1, column=i).font = _BOLD_WHITE


def _style_table(ws, name: str, n_cols: int, widths):
    if ws.max_row < 2:
        for idx, w in enumerate(widths, start=1):
            if idx > n_cols:
                break
            ws.column_dimensions[get_column_letter(idx)].width = w
        return
    ref = f"A1:{get_column_letter(n_cols)}{ws.max_row}"
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)
    ws.freeze_panes = "A2"
    for idx, w in enumerate(widths, start=1):
        if idx > n_cols:
            break
        ws.column_dimensions[get_column_letter(idx)].width = w


def build_archived_question_workbook(
    db: Session, archived: models.ArchivedQuestion
) -> Workbook:
    """
    Workbook con un singolo sheet "Database_model": una riga per lingua
    archiviata, con esempi/glosse/traduzioni concatenati per cella (\\n).
    Tutte le info derivano dallo snapshot archive: niente lookup vivi.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Database_model"
    ws.append(ARCHIVED_DB_HEADERS)
    _bold_header_row(ws, len(ARCHIVED_DB_HEADERS))

    answers = sorted(archived.answers, key=lambda a: a.language_id)
    for a in answers:
        # Risposta: stringa stile vecchio progetto (YES/NO)
        if a.response_text == "yes":
            lang_answer = "YES"
        elif a.response_text == "no":
            lang_answer = "NO"
        else:
            lang_answer = ""

        ex_list = sorted(a.examples, key=lambda e: (e.number or "", e.id or 0))
        cell_examples = "\n".join((ex.textarea or "") for ex in ex_list) if ex_list else ""
        cell_translit = "\n".join((ex.transliteration or "") for ex in ex_list) if ex_list else ""
        cell_gloss = "\n".join((ex.gloss or "") for ex in ex_list) if ex_list else ""
        cell_transl = "\n".join((ex.translation or "") for ex in ex_list) if ex_list else ""
        cell_refs = "\n".join((ex.reference or "") for ex in ex_list) if ex_list else ""

        # Motivations selezionate: code + label per leggibilita'.
        mots_str = ", ".join(
            f"{m.motivation_code} ({m.motivation_label})" if m.motivation_label else m.motivation_code
            for m in a.answer_motivations
        )

        ws.append([
            a.language_name_full or "",
            a.language_id or "",
            archived.parameter_id or "",
            archived.parameter_name or "",
            archived.original_question_id or "",
            archived.text or "",
            archived.example_yes or "",
            archived.instruction or "",
            lang_answer,
            a.comments or "",
            mots_str,
            cell_examples,
            cell_translit,
            cell_gloss,
            cell_transl,
            cell_refs,
        ])

    _style_table(
        ws, "ArchivedDatabaseModel", len(ARCHIVED_DB_HEADERS),
        [22, 12, 12, 22, 14, 36, 26, 26, 12, 22, 26, 30, 24, 22, 26, 24],
    )
    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
