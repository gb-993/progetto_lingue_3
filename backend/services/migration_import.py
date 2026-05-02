"""
Migration Bundle Importer.

Importa in blocco lo stato del vecchio sito Django partendo da un ZIP che
contiene fogli Excel + xlsx Database_model per lingua. Pensato per un'unica
operazione di seed alla messa online del nuovo sito.

Contratto del bundle (vedi `docs` o l'export Django gemello):

    PCM_migration_<ts>.zip
    ├── 00_languages.xlsx
    │     ID, Name, Position, Top-level family, Family, Group,
    │     ISO code, Glottocode, Location, Latitude, Longitude,
    │     Supervisor, Informant, Historical, Source
    ├── 01_motivations.xlsx        (ID, Code, Label)
    ├── 02_parameters.xlsx         (ID, Position, Name, Schema, Type, Level,
    │                                Short Description, Long Description,
    │                                Implicational Condition,
    │                                Explanation of Implicational Condition,
    │                                Is Active)
    ├── 03_questions.xlsx          (ID, Parameter ID, Text, Template Type,
    │                                Instruction, Instruction YES, Instruction NO,
    │                                Example YES, Help Info,
    │                                Is Stop Question, Is Active)
    ├── 04_question_allowed_motivations.xlsx (Question ID, Motivation Code)
    ├── 06_glossary.xlsx           (Word, Description)
    ├── 08_unsure_flags.xlsx       (Language ID, Parameter ID)   # opzionale
    └── data/
        ├── <Lingua1>.xlsx         (foglio Database_model)
        └── ...

Strategia operativa:
    - se `wipe=True`, TRUNCATE delle tabelle dati in ordine FK-safe prima dell'import;
    - upsert per id/code/word su tutte le entità di schema;
    - tassonomia (top_families/families/groups) ricavata dalle stringhe presenti
      su 00_languages.xlsx;
    - per ogni file in data/: cancella eventuali risposte esistenti della lingua
      e ricrea Answer come `approved` + Example;
    - alla fine: per ogni lingua, calcola LanguageParameter (consolidate) + esegue il DAG;
    - admin di default ricreato sempre (env: ADMIN_EMAIL, ADMIN_PASSWORD);
    - per ogni entità versionata (Motivation, Parameter, Question, Language,
      Answer) viene registrata una EntityVersion con operation='create' e
      source='migration_import'. Questo dà a ciascuna entità una "prima
      versione" da cui partire: ogni modifica futura mostrerà diff puliti
      contro questo snapshot iniziale, come se i dati fossero stati immessi
      manualmente al momento del seed.

Endpoint chiamante: routers/migration.py
"""
from __future__ import annotations
from dataclasses import dataclass, field, asdict
from typing import List, Optional, Dict, Set, Any, Tuple, Iterable
from datetime import datetime
from time_utils import utc_now
import io
import os
import zipfile

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError, DataError
import bcrypt

import models
from services.dag_eval import run_dag_for_language
from services.param_consolidate import recompute_and_persist_language_parameter
from services.versioning import record_version
from services.migration_progress import ProgressReporter, NULL_PROGRESS


# ============================================================================
# Report
# ============================================================================

@dataclass
class MigrationError:
    section: str
    row: int = 0
    column: Optional[str] = None
    value: Optional[str] = None
    reason: str = ""

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class SectionSummary:
    rows_total: int = 0
    inserted: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class MigrationReport:
    sections: List[str] = field(default_factory=list)
    by_section: Dict[str, SectionSummary] = field(default_factory=dict)
    errors: List[MigrationError] = field(default_factory=list)
    languages_imported: List[str] = field(default_factory=list)
    languages_dag_failed: List[Tuple[str, str]] = field(default_factory=list)
    admin_email: Optional[str] = None
    wipe_performed: bool = False

    def to_dict(self) -> dict:
        return {
            "sections": self.sections,
            "by_section": {k: v.to_dict() for k, v in self.by_section.items()},
            "errors": [e.to_dict() for e in self.errors],
            "languages_imported": self.languages_imported,
            "languages_dag_failed": [
                {"language_id": lid, "error": err}
                for (lid, err) in self.languages_dag_failed
            ],
            "admin_email": self.admin_email,
            "wipe_performed": self.wipe_performed,
            "total_errors": len(self.errors),
        }

    def section(self, name: str) -> SectionSummary:
        if name not in self.by_section:
            self.by_section[name] = SectionSummary()
            self.sections.append(name)
        return self.by_section[name]


# ============================================================================
# Helpers
# ============================================================================

def _str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _none_if_empty(v: Any) -> Optional[str]:
    s = _str(v)
    return s if s else None


def _bool_yn(v: Any) -> bool:
    s = _str(v).lower()
    return s in ("yes", "y", "true", "1", "x")


def _int_or(v: Any, default: int = 0) -> int:
    s = _str(v)
    if not s:
        return default
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return default


def _float_or_none(v: Any) -> Optional[float]:
    s = _str(v)
    if not s:
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _build_header_map(ws: Worksheet) -> Dict[str, int]:
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {_str(h): i for i, h in enumerate(header_row) if _str(h)}


def _get(row: Tuple, hmap: Dict[str, int], col: str) -> Any:
    idx = hmap.get(col)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _open_xlsx_from_zip(zf: zipfile.ZipFile, name: str) -> Optional[Worksheet]:
    """Ritorna il worksheet attivo del file `name` dentro lo zip, o None se assente."""
    if name not in zf.namelist():
        return None
    with zf.open(name) as fp:
        wb = load_workbook(io.BytesIO(fp.read()), data_only=True)
    return wb.active


# ============================================================================
# WIPE: TRUNCATE FK-safe
# ============================================================================

# Ordine di cancellazione: prima foglie, poi tabelle parent.
WIPE_ORDER = [
    "entity_versions",
    "submission_params",
    "submission_answer_motivations",
    "submission_examples",
    "submission_answers",
    "submissions",
    "answer_motivations",
    "examples",
    "answers",
    "language_parameter_evals",
    "language_parameters",
    "language_parameter_statuses",
    "parameter_change_logs",
    "site_contents",
    "glossary",
    "question_allowed_motivations",
    "questions",
    "motivations",
    "parameter_defs",
    "param_schemas",
    "param_types",
    "param_levels_of_comparison",
    "languages",
    "groups",
    "families",
    "top_families",
]

# Tabella users non viene mai svuotata interamente: lasciamo i record esistenti
# (es. eventuali account admin custom). Per il flusso "DB pulito" l'utente lo
# avrà comunque azzerato a monte.


def _wipe_all(db: Session) -> None:
    """TRUNCATE in ordine FK-safe. Postgres: TRUNCATE ... CASCADE per sicurezza."""
    from sqlalchemy import text
    for table in WIPE_ORDER:
        db.execute(text(f'TRUNCATE TABLE "{table}" RESTART IDENTITY CASCADE'))
    db.flush()


# ============================================================================
# DEFAULT ADMIN
# ============================================================================

def _ensure_default_admin(db: Session) -> str:
    """Crea/aggiorna l'admin di default. Ritorna l'email."""
    email = os.getenv("ADMIN_EMAIL", "admin@pcm.local").strip().lower()
    password = os.getenv("ADMIN_PASSWORD", "admin")

    user = db.query(models.User).filter(models.User.email == email).first()
    hashed = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

    if user:
        user.hashed_password = hashed
        user.role = "admin"
        user.is_active = True
        user.terms_accepted = True
        user.terms_accepted_at = user.terms_accepted_at or utc_now()
    else:
        user = models.User(
            email=email,
            hashed_password=hashed,
            name="Admin",
            surname="PCM",
            role="admin",
            terms_accepted=True,
            terms_accepted_at=utc_now(),
            is_active=True,
            date_joined=utc_now(),
        )
        db.add(user)
    db.flush()
    return email


# ============================================================================
# 1. MOTIVATIONS — upsert per code
# ============================================================================

def _import_motivations(db: Session, ws: Worksheet, report: MigrationReport) -> None:
    summary = report.section("Motivations")
    hmap = _build_header_map(ws)
    if "Code" not in hmap:
        report.errors.append(MigrationError(
            section="Motivations", row=1, reason="Missing 'Code' column"
        ))
        return

    by_code = {m.code: m for m in db.query(models.Motivation).all()}

    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _str(v) == "" for v in row):
            continue
        summary.rows_total += 1
        code = _str(_get(row, hmap, "Code"))
        if not code:
            summary.errors += 1
            report.errors.append(MigrationError(
                section="Motivations", row=ridx, column="Code", reason="Empty Code"
            ))
            continue
        label = _str(_get(row, hmap, "Label"))
        existing = by_code.get(code)
        if existing:
            existing.label = label or existing.label
            summary.updated += 1
        else:
            obj = models.Motivation(code=code, label=label or code)
            db.add(obj)
            db.flush()
            by_code[code] = obj
            record_version(db, obj, operation="create", source="migration_import",
                           user_id=None, note="Initial seed (migration import)")
            summary.inserted += 1


# ============================================================================
# 2. PARAMETERS — upsert per id
# ============================================================================

PARAM_FIELDS = (
    ("Name", "name", _str),
    ("Schema", "schema", _str),
    ("Type", "param_type", _str),
    ("Level", "level_of_comparison", _str),
    ("Short Description", "short_description", _str),
    ("Long Description", "long_description", _str),
    ("Implicational Condition", "implicational_condition", _none_if_empty),
    ("Explanation of Implicational Condition", "description_of_the_implicational_condition", _str),
)


def _ensure_param_lookup(
    db: Session,
    model_cls,
    raw_label: str,
    cache: Dict[str, Any],
) -> str:
    """
    Upsert su una lookup table di parametri (ParamSchema/ParamType/ParamLevelOfComparison)
    con matching case-insensitive. Ritorna la label canonica (quella effettivamente
    presente in DB) così che il valore stringa sul ParameterDef venga normalizzato e
    coincida esattamente con un'opzione delle tendine del form.
    """
    if raw_label is None:
        return ""
    label = str(raw_label).strip()
    if not label:
        return ""
    key = label.lower()
    if key in cache:
        return cache[key].label
    existing = db.query(model_cls).filter(func.lower(model_cls.label) == key).first()
    if existing:
        cache[key] = existing
        return existing.label
    obj = model_cls(label=label)
    db.add(obj)
    db.flush()
    cache[key] = obj
    return obj.label


def _import_parameters(db: Session, ws: Worksheet, report: MigrationReport) -> None:
    summary = report.section("Parameters")
    hmap = _build_header_map(ws)
    if "ID" not in hmap:
        report.errors.append(MigrationError(
            section="Parameters", row=1, reason="Missing 'ID' column"
        ))
        return

    by_id = {p.id: p for p in db.query(models.ParameterDef).all()}
    next_position = 1 + (
        db.query(models.ParameterDef).count() and
        max((p.position or 0) for p in db.query(models.ParameterDef).all()) or 0
    )

    # Cache locali per le lookup (riempite on-the-fly, case-insensitive)
    schema_cache: Dict[str, models.ParamSchema] = {}
    type_cache: Dict[str, models.ParamType] = {}
    level_cache: Dict[str, models.ParamLevelOfComparison] = {}

    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _str(v) == "" for v in row):
            continue
        summary.rows_total += 1
        pid = _str(_get(row, hmap, "ID"))
        if not pid:
            summary.errors += 1
            report.errors.append(MigrationError(
                section="Parameters", row=ridx, column="ID", reason="Empty ID"
            ))
            continue

        position = _int_or(_get(row, hmap, "Position"), default=next_position)
        is_active_raw = _get(row, hmap, "Is Active")
        is_active = True if is_active_raw is None or _str(is_active_raw) == "" else _bool_yn(is_active_raw)

        existing = by_id.get(pid)
        try:
            if existing:
                for col, attr, parser in PARAM_FIELDS:
                    setattr(existing, attr, parser(_get(row, hmap, col)))
                existing.position = position
                existing.is_active = is_active
                target = existing
                summary.updated += 1
            else:
                kwargs = {col_attr[1]: col_attr[2](_get(row, hmap, col_attr[0]))
                          for col_attr in PARAM_FIELDS}
                obj = models.ParameterDef(
                    id=pid, position=position, is_active=is_active, **kwargs,
                )
                db.add(obj)
                db.flush()
                by_id[pid] = obj
                target = obj
                summary.inserted += 1
                next_position = max(next_position, position) + 1

            # Popola le lookup table e normalizza il valore stringa sul ParameterDef
            # (case-insensitive: la prima occorrenza definisce la forma canonica).
            target.schema = _ensure_param_lookup(db, models.ParamSchema, target.schema, schema_cache)
            target.param_type = _ensure_param_lookup(db, models.ParamType, target.param_type, type_cache)
            target.level_of_comparison = _ensure_param_lookup(
                db, models.ParamLevelOfComparison, target.level_of_comparison, level_cache
            )

            # Versione iniziale solo per gli inserimenti (existing == None).
            if existing is None:
                db.flush()
                record_version(db, target, operation="create", source="migration_import",
                               user_id=None, note="Initial seed (migration import)")
        except (IntegrityError, DataError) as e:
            db.rollback()
            summary.errors += 1
            report.errors.append(MigrationError(
                section="Parameters", row=ridx, value=pid, reason=str(e)[:200]
            ))


# ============================================================================
# 3. QUESTIONS — upsert per id
# ============================================================================

QUESTION_FIELDS = (
    ("Text", "text", _str),
    ("Template Type", "template_type", _str),
    ("Instruction", "instruction", _none_if_empty),
    ("Instruction YES", "instruction_yes", _none_if_empty),
    ("Instruction NO", "instruction_no", _none_if_empty),
    ("Example YES", "example_yes", _none_if_empty),
    ("Help Info", "help_info", _none_if_empty),
)


def _import_questions(db: Session, ws: Worksheet, report: MigrationReport) -> None:
    summary = report.section("Questions")
    hmap = _build_header_map(ws)
    if "ID" not in hmap or "Parameter ID" not in hmap:
        report.errors.append(MigrationError(
            section="Questions", row=1,
            reason="Required columns: 'ID', 'Parameter ID'"
        ))
        return

    by_id = {q.id: q for q in db.query(models.Question).all()}
    valid_param_ids = {p for (p,) in db.query(models.ParameterDef.id).all()}

    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _str(v) == "" for v in row):
            continue
        summary.rows_total += 1
        qid = _str(_get(row, hmap, "ID"))
        param_id = _str(_get(row, hmap, "Parameter ID"))
        if not qid or not param_id:
            summary.errors += 1
            report.errors.append(MigrationError(
                section="Questions", row=ridx,
                reason="Missing ID or Parameter ID"
            ))
            continue
        if param_id not in valid_param_ids:
            summary.errors += 1
            report.errors.append(MigrationError(
                section="Questions", row=ridx, column="Parameter ID",
                value=param_id, reason=f"Parameter '{param_id}' not found"
            ))
            continue

        is_stop = _bool_yn(_get(row, hmap, "Is Stop Question"))
        is_active_raw = _get(row, hmap, "Is Active")
        is_active = True if is_active_raw is None or _str(is_active_raw) == "" else _bool_yn(is_active_raw)

        existing = by_id.get(qid)
        try:
            if existing:
                existing.parameter_id = param_id
                for col, attr, parser in QUESTION_FIELDS:
                    setattr(existing, attr, parser(_get(row, hmap, col)))
                existing.is_stop_question = is_stop
                existing.is_active = is_active
                summary.updated += 1
            else:
                kwargs = {ca[1]: ca[2](_get(row, hmap, ca[0])) for ca in QUESTION_FIELDS}
                obj = models.Question(
                    id=qid, parameter_id=param_id,
                    is_stop_question=is_stop, is_active=is_active, **kwargs,
                )
                db.add(obj)
                db.flush()
                by_id[qid] = obj
                # NB: la versione 'create' per Question viene registrata dopo
                # _import_qam, così lo snapshot include allowed_motivation_codes.
                summary.inserted += 1
        except (IntegrityError, DataError) as e:
            db.rollback()
            summary.errors += 1
            report.errors.append(MigrationError(
                section="Questions", row=ridx, value=qid, reason=str(e)[:200]
            ))


# ============================================================================
# 4. QUESTION ALLOWED MOTIVATIONS — replace per question
# ============================================================================

def _import_qam(db: Session, ws: Worksheet, report: MigrationReport) -> None:
    summary = report.section("QuestionAllowedMotivations")
    hmap = _build_header_map(ws)
    if "Question ID" not in hmap or "Motivation Code" not in hmap:
        report.errors.append(MigrationError(
            section="QuestionAllowedMotivations", row=1,
            reason="Required columns: 'Question ID', 'Motivation Code'"
        ))
        return

    by_qid = {q for (q,) in db.query(models.Question.id).all()}
    by_code = {m.code: m for m in db.query(models.Motivation).all()}

    questions_seen: Set[str] = set()
    pairs: List[Tuple[str, int]] = []

    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _str(v) == "" for v in row):
            continue
        summary.rows_total += 1
        qid = _str(_get(row, hmap, "Question ID"))
        code = _str(_get(row, hmap, "Motivation Code"))
        if not qid or not code:
            summary.errors += 1
            report.errors.append(MigrationError(
                section="QuestionAllowedMotivations", row=ridx,
                reason="Missing Question ID or Motivation Code"
            ))
            continue
        if qid not in by_qid:
            summary.errors += 1
            report.errors.append(MigrationError(
                section="QuestionAllowedMotivations", row=ridx, column="Question ID",
                value=qid, reason=f"Question '{qid}' not found"
            ))
            continue
        if code not in by_code:
            summary.errors += 1
            report.errors.append(MigrationError(
                section="QuestionAllowedMotivations", row=ridx, column="Motivation Code",
                value=code, reason=f"Motivation '{code}' not found"
            ))
            continue
        questions_seen.add(qid)
        pairs.append((qid, by_code[code].id))

    if questions_seen:
        db.query(models.QuestionAllowedMotivation).filter(
            models.QuestionAllowedMotivation.question_id.in_(questions_seen)
        ).delete(synchronize_session=False)
        db.flush()
        for qid, mid in pairs:
            db.add(models.QuestionAllowedMotivation(question_id=qid, motivation_id=mid))
            summary.inserted += 1


def _record_initial_create_for_questions(db: Session) -> None:
    """Registra la versione 'create' iniziale per le Question prive di
    EntityVersion. Da chiamare DOPO _import_qam così che lo snapshot includa
    correttamente allowed_motivation_codes."""
    questions = db.query(models.Question).all()
    if not questions:
        return
    versioned_qids = {
        qid for (qid,) in db.query(models.EntityVersion.entity_id).filter(
            models.EntityVersion.entity_type == "question"
        ).distinct().all()
    }
    for q in questions:
        if q.id in versioned_qids:
            continue
        record_version(db, q, operation="create", source="migration_import",
                       user_id=None, note="Initial seed (migration import)")


# ============================================================================
# 5. TAXONOMY (top_families/families/groups) — derivata da 00_languages
# ============================================================================

def _ensure_top_family(db: Session, name: str, cache: Dict[str, models.TopFamily]) -> Optional[models.TopFamily]:
    name = name.strip()
    if not name:
        return None
    if name in cache:
        return cache[name]
    obj = db.query(models.TopFamily).filter(models.TopFamily.name == name).first()
    if not obj:
        obj = models.TopFamily(name=name, position=len(cache))
        db.add(obj)
        db.flush()
    cache[name] = obj
    return obj


def _ensure_family(db: Session, name: str, top: Optional[models.TopFamily],
                   cache: Dict[str, models.Family]) -> Optional[models.Family]:
    name = name.strip()
    if not name:
        return None
    if name in cache:
        existing = cache[name]
        if top and existing.top_family_id is None:
            existing.top_family_id = top.id
        return existing
    obj = db.query(models.Family).filter(models.Family.name == name).first()
    if not obj:
        obj = models.Family(name=name, top_family_id=top.id if top else None,
                            position=len(cache))
        db.add(obj)
        db.flush()
    elif top and obj.top_family_id is None:
        obj.top_family_id = top.id
    cache[name] = obj
    return obj


def _ensure_group(db: Session, name: str, family: Optional[models.Family],
                  cache: Dict[str, models.Group]) -> Optional[models.Group]:
    name = name.strip()
    if not name:
        return None
    if name in cache:
        existing = cache[name]
        if family and existing.family_id is None:
            existing.family_id = family.id
        return existing
    obj = db.query(models.Group).filter(models.Group.name == name).first()
    if not obj:
        obj = models.Group(name=name, family_id=family.id if family else None,
                           position=len(cache))
        db.add(obj)
        db.flush()
    elif family and obj.family_id is None:
        obj.family_id = family.id
    cache[name] = obj
    return obj


# ============================================================================
# 6. LANGUAGES — upsert per id + popolamento taxonomy
# ============================================================================

LANGUAGE_FIELDS = (
    ("Name", "name_full", _str),
    ("ISO code", "isocode", _str),
    ("Glottocode", "glottocode", _str),
    ("Location", "location", _str),
    ("Supervisor", "supervisor", _str),
    ("Informant", "informant", _str),
    ("Source", "source", _str),
)


def _import_languages(db: Session, ws: Worksheet, report: MigrationReport) -> None:
    summary = report.section("Languages")
    hmap = _build_header_map(ws)
    if "ID" not in hmap or "Name" not in hmap:
        report.errors.append(MigrationError(
            section="Languages", row=1, reason="Required columns: 'ID', 'Name'"
        ))
        return

    existing_langs = {l.id: l for l in db.query(models.Language).all()}

    top_cache: Dict[str, models.TopFamily] = {}
    fam_cache: Dict[str, models.Family] = {}
    grp_cache: Dict[str, models.Group] = {}

    next_position = 1 + max(
        (l.position or 0 for l in existing_langs.values()), default=0
    )

    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _str(v) == "" for v in row):
            continue
        summary.rows_total += 1
        lid = _str(_get(row, hmap, "ID"))
        name_full = _str(_get(row, hmap, "Name"))
        if not lid or not name_full:
            summary.errors += 1
            report.errors.append(MigrationError(
                section="Languages", row=ridx, reason="Missing ID or Name"
            ))
            continue

        top_name = _str(_get(row, hmap, "Top-level family"))
        fam_name = _str(_get(row, hmap, "Family"))
        grp_name = _str(_get(row, hmap, "Group"))
        top = _ensure_top_family(db, top_name, top_cache)
        fam = _ensure_family(db, fam_name, top, fam_cache)
        grp = _ensure_group(db, grp_name, fam, grp_cache)

        position = _int_or(_get(row, hmap, "Position"), default=next_position)
        latitude = _float_or_none(_get(row, hmap, "Latitude"))
        longitude = _float_or_none(_get(row, hmap, "Longitude"))
        historical = _bool_yn(_get(row, hmap, "Historical"))

        existing = existing_langs.get(lid)
        try:
            if existing:
                for col, attr, parser in LANGUAGE_FIELDS:
                    val = _get(row, hmap, col)
                    if val is not None:
                        setattr(existing, attr, parser(val))
                existing.position = position
                existing.top_level_family = top_name
                existing.family = fam_name
                existing.grp = grp_name
                existing.top_family_id = top.id if top else None
                existing.family_id = fam.id if fam else None
                existing.group_id = grp.id if grp else None
                existing.latitude = latitude
                existing.longitude = longitude
                existing.historical_language = historical
                # workflow: tutte le lingue importate partono in pending
                existing.status = "pending"
                summary.updated += 1
            else:
                kwargs = {ca[1]: ca[2](_get(row, hmap, ca[0])) for ca in LANGUAGE_FIELDS}
                obj = models.Language(
                    id=lid, name_full=name_full, position=position,
                    top_level_family=top_name, family=fam_name, grp=grp_name,
                    top_family_id=top.id if top else None,
                    family_id=fam.id if fam else None,
                    group_id=grp.id if grp else None,
                    latitude=latitude, longitude=longitude,
                    historical_language=historical,
                    status="pending",
                    **{k: v for k, v in kwargs.items() if k != "name_full"},
                )
                db.add(obj)
                db.flush()
                existing_langs[lid] = obj
                record_version(db, obj, operation="create", source="migration_import",
                               user_id=None, note="Initial seed (migration import)")
                summary.inserted += 1
                next_position = max(next_position, position) + 1
        except (IntegrityError, DataError) as e:
            db.rollback()
            summary.errors += 1
            report.errors.append(MigrationError(
                section="Languages", row=ridx, value=lid, reason=str(e)[:200]
            ))


# ============================================================================
# 7. GLOSSARY — upsert per word
# ============================================================================

def _import_glossary(db: Session, ws: Worksheet, report: MigrationReport) -> None:
    summary = report.section("Glossary")
    hmap = _build_header_map(ws)
    if "Word" not in hmap:
        report.errors.append(MigrationError(
            section="Glossary", row=1, reason="Missing 'Word' column"
        ))
        return

    by_word = {g.word: g for g in db.query(models.Glossary).all()}

    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _str(v) == "" for v in row):
            continue
        summary.rows_total += 1
        word = _str(_get(row, hmap, "Word"))
        desc = _str(_get(row, hmap, "Description"))
        if not word:
            summary.errors += 1
            report.errors.append(MigrationError(
                section="Glossary", row=ridx, column="Word", reason="Empty Word"
            ))
            continue
        existing = by_word.get(word)
        if existing:
            existing.description = desc or existing.description
            summary.updated += 1
        else:
            db.add(models.Glossary(word=word, description=desc or ""))
            summary.inserted += 1


# ============================================================================
# 8. UNSURE FLAGS — replicano i ParameterReviewFlag del vecchio in is_unsure
# ============================================================================

def _import_unsure_flags(db: Session, ws: Worksheet, report: MigrationReport) -> None:
    summary = report.section("UnsureFlags")
    hmap = _build_header_map(ws)
    if "Language ID" not in hmap or "Parameter ID" not in hmap:
        report.errors.append(MigrationError(
            section="UnsureFlags", row=1,
            reason="Required columns: 'Language ID', 'Parameter ID'"
        ))
        return

    valid_lang = {l for (l,) in db.query(models.Language.id).all()}
    valid_param = {p for (p,) in db.query(models.ParameterDef.id).all()}

    seen: Set[Tuple[str, str]] = set()
    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _str(v) == "" for v in row):
            continue
        summary.rows_total += 1
        lid = _str(_get(row, hmap, "Language ID"))
        pid = _str(_get(row, hmap, "Parameter ID"))
        if not lid or not pid:
            summary.errors += 1
            continue
        if lid not in valid_lang or pid not in valid_param:
            summary.skipped += 1
            continue
        if (lid, pid) in seen:
            continue
        seen.add((lid, pid))
        existing = (
            db.query(models.LanguageParameterStatus)
            .filter(models.LanguageParameterStatus.language_id == lid,
                    models.LanguageParameterStatus.parameter_id == pid)
            .first()
        )
        if existing:
            existing.is_unsure = True
            summary.updated += 1
        else:
            db.add(models.LanguageParameterStatus(
                language_id=lid, parameter_id=pid, is_unsure=True
            ))
            summary.inserted += 1


# ============================================================================
# 9. COMPILATION — un xlsx Database_model per lingua, status=approved
# ============================================================================

DB_MODEL_REQUIRED = ["Language", "Parameter_Label", "Question_ID", "Language_Answer"]


def _split_lines(v: Any) -> List[str]:
    s = _str(v)
    if not s:
        return []
    return [line.strip() for line in s.replace("\r\n", "\n").replace("\r", "\n").split("\n")]


def _import_compilation_xlsx(db: Session, ws: Worksheet, source_name: str,
                             report: MigrationReport) -> Optional[str]:
    """Importa il foglio Database_model per una lingua. Ritorna l'id della lingua processata."""
    section = f"data/{source_name}"
    summary = report.section(section)
    hmap = _build_header_map(ws)
    missing = [c for c in DB_MODEL_REQUIRED if c not in hmap]
    if missing:
        report.errors.append(MigrationError(
            section=section, row=1,
            reason=f"Missing columns: {', '.join(missing)}"
        ))
        return None

    rows: List[Tuple[int, Tuple]] = []
    lang_values: Set[str] = set()
    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or _str(v) == "" for v in row):
            continue
        rows.append((ridx, row))
        lv = _str(_get(row, hmap, "Language"))
        if lv:
            lang_values.add(lv)

    if not rows:
        return None
    if len(lang_values) != 1:
        report.errors.append(MigrationError(
            section=section, row=0,
            reason=f"Expected exactly one language; found: {sorted(lang_values)}"
        ))
        return None

    lang_name = next(iter(lang_values))
    lang = db.query(models.Language).filter(
        models.Language.name_full == lang_name
    ).first()
    if not lang:
        report.errors.append(MigrationError(
            section=section, row=0, value=lang_name,
            reason=f"Language '{lang_name}' not found in 00_languages"
        ))
        return None

    # Replace totale: cancella risposte/esempi/motivazioni esistenti per la lingua
    old_answer_ids = [
        a_id for (a_id,) in db.query(models.Answer.id)
        .filter(models.Answer.language_id == lang.id).all()
    ]
    if old_answer_ids:
        db.query(models.Example).filter(
            models.Example.answer_id.in_(old_answer_ids)
        ).delete(synchronize_session=False)
        db.query(models.AnswerMotivation).filter(
            models.AnswerMotivation.answer_id.in_(old_answer_ids)
        ).delete(synchronize_session=False)
        db.query(models.Answer).filter(
            models.Answer.id.in_(old_answer_ids)
        ).delete(synchronize_session=False)
        db.flush()

    valid_qids = {q for (q,) in db.query(models.Question.id).all()}
    motivation_id_by_code = {m.code: m.id for m in db.query(models.Motivation).all()}

    for ridx, row in rows:
        summary.rows_total += 1
        qid = _str(_get(row, hmap, "Question_ID"))
        if not qid:
            summary.errors += 1
            report.errors.append(MigrationError(
                section=section, row=ridx, column="Question_ID",
                reason="Empty Question_ID"
            ))
            continue
        if qid not in valid_qids:
            summary.errors += 1
            report.errors.append(MigrationError(
                section=section, row=ridx, column="Question_ID",
                value=qid, reason=f"Question '{qid}' not found"
            ))
            continue

        raw_ans = _str(_get(row, hmap, "Language_Answer")).upper()
        if raw_ans in ("YES", "Y"):
            response = "yes"
        elif raw_ans in ("NO", "N"):
            response = "no"
        elif raw_ans == "":
            summary.skipped += 1
            continue
        else:
            summary.errors += 1
            report.errors.append(MigrationError(
                section=section, row=ridx, column="Language_Answer",
                value=raw_ans, reason="Invalid value (expected YES/NO/empty)"
            ))
            continue

        comments = _str(_get(row, hmap, "Language_Comments"))
        ex_texts = _split_lines(_get(row, hmap, "Language_Examples"))
        translit_lines = _split_lines(_get(row, hmap, "Language_Example_Transliteration"))
        gloss_lines = _split_lines(_get(row, hmap, "Language_Example_Gloss"))
        transl_lines = _split_lines(_get(row, hmap, "Language_Example_Translation"))
        ref_lines = _split_lines(_get(row, hmap, "Language_References"))

        # Codici motivazioni associate alla risposta: una sola cella, codici
        # separati da "," o ";". Codici sconosciuti vengono segnalati ma non
        # bloccano la creazione della risposta.
        mot_codes_raw = _str(_get(row, hmap, "Language_Motivations"))
        mot_codes: List[str] = []
        if mot_codes_raw:
            for token in mot_codes_raw.replace(";", ",").split(","):
                code = token.strip()
                if code:
                    mot_codes.append(code)

        try:
            answer = models.Answer(
                language_id=lang.id, question_id=qid,
                response_text=response, comments=comments or "",
                status="approved",
            )
            db.add(answer)
            db.flush()

            n_ex = max(len(ex_texts), len(translit_lines), len(gloss_lines),
                       len(transl_lines), len(ref_lines))
            for i in range(n_ex):
                txt = ex_texts[i] if i < len(ex_texts) else ""
                tl = translit_lines[i] if i < len(translit_lines) else ""
                gl = gloss_lines[i] if i < len(gloss_lines) else ""
                tr = transl_lines[i] if i < len(transl_lines) else ""
                rf = ref_lines[i] if i < len(ref_lines) else ""
                if not (txt or tl or gl or tr or rf):
                    continue
                db.add(models.Example(
                    answer_id=answer.id, number=str(i + 1),
                    textarea=txt,
                    transliteration=tl,
                    gloss=gl,
                    translation=tr,
                    reference=rf,
                ))

            seen_mot_ids: Set[int] = set()
            for code in mot_codes:
                mid = motivation_id_by_code.get(code)
                if mid is None:
                    report.errors.append(MigrationError(
                        section=section, row=ridx, column="Language_Motivations",
                        value=code, reason=f"Motivation '{code}' not found"
                    ))
                    continue
                if mid in seen_mot_ids:
                    continue
                seen_mot_ids.add(mid)
                db.add(models.AnswerMotivation(
                    answer_id=answer.id, motivation_id=mid,
                ))

            # Snapshot iniziale dell'Answer comprensivo di examples e
            # motivation_codes (richiede flush per popolare le relationship).
            db.flush()
            record_version(db, answer, operation="create", source="migration_import",
                           user_id=None, note="Initial seed (migration import)")
            summary.inserted += 1
        except (IntegrityError, DataError) as e:
            db.rollback()
            summary.errors += 1
            report.errors.append(MigrationError(
                section=section, row=ridx, value=qid, reason=str(e)[:200]
            ))
            continue

    db.flush()
    return lang.id


# ============================================================================
# 10. POST-IMPORT: consolidate + DAG per ogni lingua
# ============================================================================

def _run_post_import_evaluation(db: Session, language_ids: Iterable[str],
                                report: MigrationReport,
                                progress: ProgressReporter = NULL_PROGRESS) -> None:
    """Per ogni lingua: ricalcola LanguageParameter (consolidate) + esegue il DAG."""
    summary = report.section("DAG")
    param_ids = [p for (p,) in db.query(models.ParameterDef.id).all()]

    language_ids = list(language_ids)
    total = len(language_ids)
    for i, lid in enumerate(language_ids, start=1):
        progress.tick(current=i, label=f"DAG & consolidate ({i}/{total}): {lid}")
        summary.rows_total += 1
        # 1) consolidate per ogni parametro (popola value_orig)
        try:
            for pid in param_ids:
                recompute_and_persist_language_parameter(lid, pid, db)
            db.flush()
        except Exception as e:
            summary.errors += 1
            report.languages_dag_failed.append((lid, f"consolidate: {e}"))
            continue

        # 2) DAG (popola value_eval)
        try:
            run_dag_for_language(lid, db)
            db.flush()
            summary.inserted += 1
        except Exception as e:
            summary.errors += 1
            report.languages_dag_failed.append((lid, f"dag: {e}"))


# ============================================================================
# Main entry point
# ============================================================================

def import_migration_bundle(db: Session, file_bytes: bytes, wipe: bool = True,
                            progress: ProgressReporter = NULL_PROGRESS) -> MigrationReport:
    """Punto di ingresso. Apre il ZIP e orchestra tutte le fasi.

    `progress` è un ProgressReporter opzionale: se passato, le fasi e gli
    avanzamenti per-lingua vengono pubblicati nello stato job (vedi
    `services/migration_progress.py`).
    """
    report = MigrationReport()

    # 1. Apertura ZIP
    progress.phase("opening_zip", label="Opening migration bundle...")
    try:
        zf = zipfile.ZipFile(io.BytesIO(file_bytes), "r")
    except Exception as e:
        report.errors.append(MigrationError(section="(zip)", reason=f"Cannot open zip: {e}"))
        return report

    names = set(zf.namelist())

    # 2. Wipe (opzionale)
    if wipe:
        progress.phase("wipe", label="Wiping existing data...")
        try:
            _wipe_all(db)
            db.commit()
            report.wipe_performed = True
        except Exception as e:
            db.rollback()
            report.errors.append(MigrationError(section="(wipe)", reason=str(e)))
            return report

    # 3. Schema (motivations -> parameters -> questions -> qam -> glossary)
    section_files = [
        ("01_motivations.xlsx", _import_motivations, "motivations", "Importing motivations"),
        ("02_parameters.xlsx", _import_parameters, "parameters", "Importing parameters"),
        ("03_questions.xlsx", _import_questions, "questions", "Importing questions"),
        ("04_question_allowed_motivations.xlsx", _import_qam, "qam", "Importing question/motivation links"),
        ("06_glossary.xlsx", _import_glossary, "glossary", "Importing glossary"),
    ]
    for fname, importer, phase_id, phase_label in section_files:
        if fname not in names:
            continue
        ws = _open_xlsx_from_zip(zf, fname)
        if ws is None:
            continue
        progress.phase(phase_id, label=phase_label)
        try:
            importer(db, ws, report)
            db.commit()
        except Exception as e:
            db.rollback()
            report.errors.append(MigrationError(section=fname, reason=str(e)[:200]))

    # Registra la create iniziale per le Question dopo che le qam sono state
    # importate, così lo snapshot riflette le motivations associate.
    progress.phase("question_versions", label="Recording initial Question versions...")
    try:
        _record_initial_create_for_questions(db)
        db.commit()
    except Exception as e:
        db.rollback()
        report.errors.append(MigrationError(
            section="(versioning)",
            reason=f"Could not record initial Question versions: {str(e)[:200]}",
        ))

    # 4. Languages (popola anche taxonomy)
    if "00_languages.xlsx" in names:
        ws = _open_xlsx_from_zip(zf, "00_languages.xlsx")
        if ws is not None:
            progress.phase("languages", label="Importing languages and taxonomy")
            try:
                _import_languages(db, ws, report)
                db.commit()
            except Exception as e:
                db.rollback()
                report.errors.append(MigrationError(section="00_languages.xlsx", reason=str(e)[:200]))

    # 5. Compilation per ogni file in data/
    data_files = sorted([n for n in names if n.startswith("data/") and n.endswith(".xlsx")])
    processed_lang_ids: List[str] = []
    total_files = len(data_files)
    progress.phase("compilation", label=f"Importing compilation data (0/{total_files})", total=total_files)
    for i, fname in enumerate(data_files, start=1):
        ws = _open_xlsx_from_zip(zf, fname)
        if ws is None:
            continue
        base = os.path.basename(fname)
        progress.tick(current=i, label=f"Compilation ({i}/{total_files}): {base}")
        try:
            lid = _import_compilation_xlsx(db, ws, base, report)
            if lid:
                processed_lang_ids.append(lid)
                report.languages_imported.append(lid)
            db.commit()
        except Exception as e:
            db.rollback()
            report.errors.append(MigrationError(section=fname, reason=str(e)[:200]))

    # 6. Unsure flags (richiede lingue + parametri già importati)
    if "08_unsure_flags.xlsx" in names:
        ws = _open_xlsx_from_zip(zf, "08_unsure_flags.xlsx")
        if ws is not None:
            progress.phase("unsure_flags", label="Importing unsure flags")
            try:
                _import_unsure_flags(db, ws, report)
                db.commit()
            except Exception as e:
                db.rollback()
                report.errors.append(MigrationError(section="08_unsure_flags.xlsx", reason=str(e)[:200]))

    # 7. Consolidate + DAG
    if processed_lang_ids:
        progress.phase("dag", label=f"Computing parameters and DAG (0/{len(processed_lang_ids)})",
                       total=len(processed_lang_ids))
        try:
            _run_post_import_evaluation(db, processed_lang_ids, report, progress=progress)
            db.commit()
        except Exception as e:
            db.rollback()
            report.errors.append(MigrationError(section="DAG", reason=str(e)[:200]))

    # 8. Default admin (sempre, anche senza wipe — è idempotente)
    progress.phase("admin", label="Ensuring default admin user")
    try:
        report.admin_email = _ensure_default_admin(db)
        db.commit()
    except Exception as e:
        db.rollback()
        report.errors.append(MigrationError(section="(admin)", reason=str(e)[:200]))

    return report
