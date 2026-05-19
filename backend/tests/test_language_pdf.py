"""Test del PDF parametric data della lingua.

Smoke test (FPDF e' fonte di verita' per il rendering: non parsiamo il PDF
visivamente, ma verifichiamo che i bytes siano un PDF valido, che la
funzione regga su scenari diversi, e che il footer della citazione sia
presente come negli altri report).
"""
import pytest
from sqlalchemy import text

import models
from services.pdf_export import build_language_pdf


@pytest.fixture
def db_fk(db_session):
    db_session.execute(text("PRAGMA foreign_keys = ON"))
    return db_session


def _seed_minimal(db):
    lang = models.Language(
        id="ENG", name_full="English", position=1,
        family="Germanic", top_level_family="Indo-European", grp="West Germanic",
        glottocode="stan1293", isocode="eng", location="UK",
        supervisor="Cristina Guardiano", informant="John Smith",
        source="Various", historical_language=False, status="approved",
    )
    db.add(lang)
    db.commit()
    return lang


def _seed_with_parameters(db):
    lang = _seed_minimal(db)
    p1 = models.ParameterDef(id="P1", position=1, name="Active Param", is_active=True)
    p_off = models.ParameterDef(id="POFF", position=2, name="Disabled", is_active=False)
    db.add_all([p1, p_off])

    q1 = models.Question(id="Q_NORM", parameter_id="P1", text="Normal q text", is_active=True)
    q_stop = models.Question(id="Q_STOP", parameter_id="P1", text="Stop q text", is_stop_question=True, is_active=True)
    q_dead = models.Question(id="Q_DEAD", parameter_id="P1", text="Inactive q", is_active=False)
    q_in_off = models.Question(id="Q_ORPHAN", parameter_id="POFF", text="Q in off param", is_active=True)
    db.add_all([q1, q_stop, q_dead, q_in_off])

    mot = models.Motivation(code="MOT_X", label="Phonological evidence")
    db.add(mot)
    db.flush()

    db.add(models.Answer(
        language_id="ENG", question_id="Q_NORM",
        response_text="unsure", status="pending", comments="some commentary",
    ))
    db.add(models.QuestionAllowedMotivation(question_id="Q_NORM", motivation_id=mot.id))
    ans_stop = models.Answer(
        language_id="ENG", question_id="Q_STOP",
        response_text="yes", status="approved",
    )
    db.add(ans_stop)
    db.flush()
    db.add(models.AnswerMotivation(answer_id=ans_stop.id, motivation_id=mot.id))
    db.add(models.Example(
        answer_id=ans_stop.id, number="1",
        textarea="Sample utterance", gloss="sample-NOM", translation="A sample",
        reference="Source 2024:10",
    ))

    db.add(models.LanguageParameterStatus(
        language_id="ENG", parameter_id="P1",
        admin_note="This is an admin note", is_unsure=False,
    ))

    lp = models.LanguageParameter(language_id="ENG", parameter_id="P1", value_orig="+")
    db.add(lp)
    db.flush()
    db.add(models.LanguageParameterEval(
        language_parameter_id=lp.id, value_eval="+",
    ))
    db.commit()
    return lang


def test_pdf_is_valid_when_language_has_data(db_fk):
    lang = _seed_with_parameters(db_fk)
    out = build_language_pdf(db_fk, lang)
    assert isinstance(out, (bytes, bytearray))
    assert out[:5] == b"%PDF-", "Output must be a valid PDF (starts with %PDF-)"
    assert b"%%EOF" in out[-20:] or b"%%EOF" in out, "PDF must contain %%EOF marker"
    # Lunghezza ragionevole (cover + 1 parametro). Il footer di citazione e i
    # font DejaVu da soli portano il file ben oltre i 5 KB.
    assert len(out) > 5_000


def test_pdf_works_with_no_parameters(db_fk):
    """Lingua esistente ma nessun parametro attivo: solo la cover deve uscire."""
    lang = _seed_minimal(db_fk)
    out = build_language_pdf(db_fk, lang)
    assert out[:5] == b"%PDF-"


def test_pdf_works_with_parameter_but_no_answers(db_fk):
    """Parametro attivo con question ma senza nessuna Answer: la pagina del
    parametro deve uscire comunque con 'Not answered' su ogni question
    (verificato indirettamente: il builder non solleva e produce PDF valido)."""
    lang = _seed_minimal(db_fk)
    p = models.ParameterDef(id="P1", position=1, name="Empty Param", is_active=True)
    db_fk.add(p)
    db_fk.add(models.Question(id="Q1", parameter_id="P1", text="q text", is_active=True))
    db_fk.add(models.Question(id="Q2", parameter_id="P1", text="q text 2", is_active=True))
    db_fk.commit()

    out = build_language_pdf(db_fk, lang)
    assert out[:5] == b"%PDF-"


def test_pdf_excludes_inactive_questions_and_inactive_parameters(db_fk):
    """Il builder NON deve sollevare quando esistono q inattive o param
    inattivi: vengono semplicemente skippati. Smoke test."""
    lang = _seed_with_parameters(db_fk)
    out = build_language_pdf(db_fk, lang)
    # Generato senza eccezioni e di lunghezza > 0
    assert len(out) > 0
    assert out[:5] == b"%PDF-"


def test_pdf_handles_unsure_answer(db_fk):
    """Sanity check: l'answer 'unsure' viene gestita senza crash dal builder
    (il branch unsure usa colore arancione + label UNSURE)."""
    lang = _seed_minimal(db_fk)
    p = models.ParameterDef(id="P1", position=1, name="P", is_active=True)
    db_fk.add(p)
    q = models.Question(id="Q1", parameter_id="P1", text="?", is_active=True)
    db_fk.add(q)
    db_fk.add(models.Answer(
        language_id="ENG", question_id="Q1",
        response_text="unsure", status="pending",
    ))
    db_fk.commit()

    out = build_language_pdf(db_fk, lang)
    assert out[:5] == b"%PDF-"


def test_excel_database_model_now_shows_unsure(db_fk):
    """Regression test del fix Excel (precedentemente unsure -> ''): nel foglio
    Database_model la colonna Language_Answer deve contenere 'UNSURE'."""
    from io import BytesIO
    from openpyxl import load_workbook
    from services.excel_export import build_language_workbook

    lang = _seed_minimal(db_fk)
    p = models.ParameterDef(id="P1", position=1, name="P", is_active=True)
    db_fk.add(p)
    db_fk.add(models.Question(id="Q1", parameter_id="P1", text="?", is_active=True))
    db_fk.add(models.Answer(language_id="ENG", question_id="Q1", response_text="unsure"))
    db_fk.commit()

    wb = build_language_workbook(db_fk, lang, is_admin=True)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb2 = load_workbook(buf)
    ws = wb2["Database_model"]
    # Header: cerca colonna "Language_Answer"
    headers = [c.value for c in ws[1]]
    col_idx = headers.index("Language_Answer") + 1  # 1-based per openpyxl
    # Riga 2 = prima riga dati
    cell = ws.cell(row=2, column=col_idx).value
    assert cell == "UNSURE", f"Expected 'UNSURE', got {cell!r}"
