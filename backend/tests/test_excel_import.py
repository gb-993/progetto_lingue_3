"""
Test per l'import Excel.

Copre:
  - Round-trip end-to-end: export → import → assert state corretto.
  - Strict update per gli schema: ID inesistente → errore esplicito, no creazione.
  - Replace totale per Database_model: vecchie risposte cancellate, nuove inserite.
  - Skip righe con errore: domande svuotate, errore nel report.
  - Cascade errors: QAM bloccata se motivation è fallita.
  - Errori a cascata espliciti, mai silenziosi.
"""
import io
import pytest
from openpyxl import Workbook, load_workbook

import models
from services.excel_export import (
    build_language_workbook,
    build_schema_workbook,
)
from services.excel_import import import_excel


# ============================================================================
# Helpers di seed
# ============================================================================

def _seed_full(db_session):
    """Stato iniziale: 1 lingua, 1 user, 1 parametro, 2 domande, 1 motivazione,
    1 risposta yes con 2 esempi."""
    user = models.User(
        id=1, email="alice@test.it", hashed_password="x",
        name="Alice", surname="Smith", role="admin",
    )
    db_session.add(user)

    lang = models.Language(
        id="ITA", name_full="Italiano", position=1,
        family="Romance", top_level_family="Indo-European",
        latitude=42.5, longitude=12.0, historical_language=False,
    )
    db_session.add(lang)

    param = models.ParameterDef(
        id="FGM", position=1, name="Old name",
        short_description="Old short", long_description="",
        is_active=True,
    )
    db_session.add(param)

    q1 = models.Question(id="FGM_01", parameter_id="FGM", text="Old text 1",
                         is_stop_question=False, is_active=True)
    q2 = models.Question(id="FGM_02", parameter_id="FGM", text="Old text 2",
                         is_stop_question=False, is_active=True)
    db_session.add_all([q1, q2])

    mot = models.Motivation(id=1, code="MOT_X", label="Old label")
    db_session.add(mot)
    db_session.flush()

    qam = models.QuestionAllowedMotivation(question_id="FGM_02", motivation_id=mot.id)
    db_session.add(qam)

    ans1 = models.Answer(language_id="ITA", question_id="FGM_01",
                         response_text="yes", comments="old comment", status="approved")
    db_session.add(ans1)
    db_session.flush()

    db_session.add(models.Example(
        answer_id=ans1.id, number="1", textarea="Old example 1",
        gloss="g1", translation="t1", reference="r1",
    ))
    db_session.add(models.Example(
        answer_id=ans1.id, number="2", textarea="Old example 2",
        gloss="g2", translation="t2", reference="r2",
    ))
    db_session.commit()
    return lang, user


def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()


# ============================================================================
# Round-trip
# ============================================================================

def test_roundtrip_no_changes_keeps_state(db_session):
    """Export → import senza modifiche: lo stato del DB deve restare invariato."""
    lang, user = _seed_full(db_session)

    wb = build_language_workbook(db_session, lang, is_admin=True)
    file_bytes = _wb_to_bytes(wb)

    report = import_excel(db_session, file_bytes, user.id)
    db_session.commit()

    # Nessun errore atteso
    assert report.errors == [], f"Errori inattesi: {report.errors}"

    # Risposte: deve essercene ancora 1 yes su FGM_01
    answers = db_session.query(models.Answer).filter(
        models.Answer.language_id == "ITA"
    ).all()
    yes_answers = [a for a in answers if a.response_text == "yes"]
    assert len(yes_answers) == 1
    assert yes_answers[0].question_id == "FGM_01"

    # Esempi: 2 sulla yes answer
    yes_ans = yes_answers[0]
    assert len(yes_ans.examples) == 2

    # Motivazione: invariata
    mot = db_session.query(models.Motivation).filter_by(code="MOT_X").one()
    assert mot.label == "Old label"


def test_roundtrip_modify_param_via_schema_excel(db_session):
    """Da 2026-05 lo schema vive solo nel workbook schema dedicato (NON più
    replicato in ogni per-lingua xlsx). Modifico il name del parametro nel
    file schema, re-importo: il DB riflette il cambio + ParameterChangeLog."""
    lang, user = _seed_full(db_session)

    wb = build_schema_workbook(db_session)

    # Modifica della cella "Name" del parametro FGM nel sheet Parameters
    ws_par = wb["Parameters"]
    headers = [c.value for c in ws_par[1]]
    name_idx = headers.index("Name") + 1  # openpyxl 1-based
    id_idx = headers.index("ID") + 1
    for row_idx in range(2, ws_par.max_row + 1):
        if ws_par.cell(row=row_idx, column=id_idx).value == "FGM":
            ws_par.cell(row=row_idx, column=name_idx).value = "New name"
            break

    file_bytes = _wb_to_bytes(wb)
    report = import_excel(db_session, file_bytes, user.id)
    db_session.commit()

    assert report.errors == []
    p = db_session.query(models.ParameterDef).filter_by(id="FGM").one()
    assert p.name == "New name"

    # ChangeLog creato
    log = db_session.query(models.ParameterChangeLog).filter_by(parameter_id="FGM").first()
    assert log is not None
    assert "Excel import" in log.change_note
    assert "name" in log.change_note


# ============================================================================
# Strict update: ID inesistente → errore (no creazione)
# ============================================================================

def test_unknown_param_id_not_created(db_session):
    """Riga di Parameters con ID inesistente → errore, no nuovo parametro creato."""
    lang, user = _seed_full(db_session)

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Parameters")
    from services.excel_export import PARAMETERS_HEADERS
    ws.append(PARAMETERS_HEADERS)
    ws.append([
        "FGB", 99, "I am new", "", "", "",
        "short", "long", "", "", "Yes",  # FGB non esiste in DB
    ])

    file_bytes = _wb_to_bytes(wb)
    report = import_excel(db_session, file_bytes, user.id)
    db_session.commit()

    # Nessun nuovo parametro creato
    fgb = db_session.query(models.ParameterDef).filter_by(id="FGB").first()
    assert fgb is None, "FGB NON deve essere creato"

    # Un errore registrato
    assert len(report.errors) == 1
    assert report.errors[0].sheet == "Parameters"
    assert "non esiste" in report.errors[0].reason
    assert report.errors[0].value == "FGB"


def test_unknown_question_id_not_created(db_session):
    lang, user = _seed_full(db_session)
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Questions")
    from services.excel_export import QUESTIONS_HEADERS
    ws.append(QUESTIONS_HEADERS)
    ws.append(["FGM_99", "FGM", "new q", "", "", "", "", "", "", "No", "Yes"])

    report = import_excel(db_session, _wb_to_bytes(wb), user.id)
    db_session.commit()

    assert db_session.query(models.Question).filter_by(id="FGM_99").first() is None
    assert any("FGM_99" in e.value and "non esiste" in e.reason for e in report.errors)


def test_unknown_motivation_code_not_created(db_session):
    lang, user = _seed_full(db_session)
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Motivations")
    from services.excel_export import MOTIVATIONS_HEADERS
    ws.append(MOTIVATIONS_HEADERS)
    ws.append([99, "MOT_NEW", "new label", "Yes"])

    report = import_excel(db_session, _wb_to_bytes(wb), user.id)
    db_session.commit()

    assert db_session.query(models.Motivation).filter_by(code="MOT_NEW").first() is None
    assert any(e.value == "MOT_NEW" and "non esiste" in e.reason for e in report.errors)


# ============================================================================
# Database_model: replace + skip righe errate (domande svuotate)
# ============================================================================

def test_database_model_replace_with_invalid_question_skipped(db_session):
    """File con 2 righe valide + 1 con question_id inesistente.
    Le 2 risposte vecchie vengono cancellate. Della riga errata, la domanda
    resta non risposta (= visibile nel report)."""
    lang, user = _seed_full(db_session)

    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Database_model")
    from services.excel_export import DATABASE_MODEL_HEADERS
    ws.append(DATABASE_MODEL_HEADERS)
    # Layout colonne (11): Language, Parameter_Label, Question_ID,
    # Language_Answer, Language_Comments, Language_Examples,
    # Language_Example_Gloss, Language_Example_Translation, Language_References,
    # Motivations, Admin_Note
    # riga valida
    ws.append([
        "Italiano", "FGM", "FGM_01",
        "YES", "new comment",
        "Esempio nuovo 1\nEsempio nuovo 2",
        "g1\ng2", "t1\nt2", "r1\nr2",
        "", "",
    ])
    # riga valida (FGM_02 = no, no esempi)
    ws.append([
        "Italiano", "FGM", "FGM_02",
        "NO", "", "", "", "", "", "", "",
    ])
    # riga errata (FGM_99 inesistente)
    ws.append([
        "Italiano", "FGM", "FGM_99",
        "YES", "should fail", "ex", "g", "t", "r", "", "",
    ])

    report = import_excel(db_session, _wb_to_bytes(wb), user.id)
    db_session.commit()

    # 2 risposte presenti, 0 per FGM_99
    answers = db_session.query(models.Answer).filter_by(language_id="ITA").all()
    by_qid = {a.question_id: a for a in answers}
    assert "FGM_01" in by_qid and by_qid["FGM_01"].response_text == "yes"
    assert "FGM_02" in by_qid and by_qid["FGM_02"].response_text == "no"
    assert "FGM_99" not in by_qid  # SVUOTATA — la riga errata = non risposta

    # Nuovi esempi (2) sulla yes, vecchi 2 cancellati
    fgm01_ans = by_qid["FGM_01"]
    assert len(fgm01_ans.examples) == 2
    texts = sorted(ex.textarea for ex in fgm01_ans.examples)
    assert texts == ["Esempio nuovo 1", "Esempio nuovo 2"]
    # Comment aggiornato
    assert fgm01_ans.comments == "new comment"

    # Errore reportato
    assert any("FGM_99" in e.value for e in report.errors)
    db_summary = report.by_sheet["Database_model"]
    assert db_summary.inserted == 2
    assert db_summary.errors == 1


def test_database_model_invalid_answer_value_skipped(db_session):
    """Risposta non YES/NO/'' → riga saltata + errore."""
    lang, user = _seed_full(db_session)

    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Database_model")
    from services.excel_export import DATABASE_MODEL_HEADERS
    ws.append(DATABASE_MODEL_HEADERS)
    ws.append([
        "Italiano", "FGM", "FGM_01",
        "MAYBE", "", "", "", "", "", "", "",
    ])

    report = import_excel(db_session, _wb_to_bytes(wb), user.id)
    db_session.commit()

    assert db_session.query(models.Answer).filter_by(language_id="ITA", question_id="FGM_01").first() is None
    assert any("MAYBE" in (e.value or "") for e in report.errors)


def test_database_model_unknown_language(db_session):
    """Lingua non in DB → tutto il sheet fallisce con errore."""
    lang, user = _seed_full(db_session)

    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Database_model")
    from services.excel_export import DATABASE_MODEL_HEADERS
    ws.append(DATABASE_MODEL_HEADERS)
    ws.append([
        "Klingon", "FGM", "FGM_01",
        "YES", "", "", "", "", "", "", "",
    ])

    # le risposte ITA esistenti devono restare invariate
    answers_before = db_session.query(models.Answer).filter_by(language_id="ITA").count()

    report = import_excel(db_session, _wb_to_bytes(wb), user.id)
    db_session.commit()

    answers_after = db_session.query(models.Answer).filter_by(language_id="ITA").count()
    assert answers_after == answers_before  # ITA non toccato
    assert any("Klingon" in (e.value or "") for e in report.errors)


# ============================================================================
# Cascade errors
# ============================================================================

def test_cascade_qam_when_motivation_failed(db_session):
    """Motivation con code inesistente → fallisce. La QAM che la referenzia
    nel file deve dare errore CASCADE esplicito."""
    lang, user = _seed_full(db_session)

    wb = Workbook(); wb.remove(wb.active)

    ws_mot = wb.create_sheet("Motivations")
    from services.excel_export import MOTIVATIONS_HEADERS, QUESTION_ALLOWED_MOTIVATIONS_HEADERS
    ws_mot.append(MOTIVATIONS_HEADERS)
    # MOT_BAD non esiste in DB → fallisce
    ws_mot.append([1, "MOT_BAD", "Bad mot", "Yes"])

    ws_qam = wb.create_sheet("QuestionAllowedMotivations")
    ws_qam.append(QUESTION_ALLOWED_MOTIVATIONS_HEADERS)
    ws_qam.append(["FGM_02", "MOT_BAD"])

    report = import_excel(db_session, _wb_to_bytes(wb), user.id)
    db_session.commit()

    # Errore upstream sulla motivation
    mot_errors = [e for e in report.errors if e.sheet == "Motivations"]
    assert len(mot_errors) == 1
    # Errore cascade sulla QAM
    qam_errors = [e for e in report.errors if e.sheet == "QuestionAllowedMotivations"]
    assert len(qam_errors) == 1
    assert "errore upstream" in qam_errors[0].reason


def test_qam_replaces_links_for_questions_in_file(db_session):
    """Le QAM nel file rimpiazzano i link esistenti per le questions menzionate."""
    lang, user = _seed_full(db_session)

    # FGM_02 ha già link a MOT_X
    pre = db_session.query(models.QuestionAllowedMotivation).filter_by(
        question_id="FGM_02"
    ).count()
    assert pre == 1

    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("QuestionAllowedMotivations")
    from services.excel_export import QUESTION_ALLOWED_MOTIVATIONS_HEADERS
    ws.append(QUESTION_ALLOWED_MOTIVATIONS_HEADERS)
    # File NON contiene il link FGM_02 → MOT_X. Aggiunge solo FGM_01 → MOT_X
    ws.append(["FGM_01", "MOT_X"])

    report = import_excel(db_session, _wb_to_bytes(wb), user.id)
    db_session.commit()

    # FGM_02 NON è stato menzionato → link conservato
    fgm02_links = db_session.query(models.QuestionAllowedMotivation).filter_by(
        question_id="FGM_02"
    ).count()
    assert fgm02_links == 1

    # FGM_01 ora ha 1 link (creato dal file)
    fgm01_links = db_session.query(models.QuestionAllowedMotivation).filter_by(
        question_id="FGM_01"
    ).count()
    assert fgm01_links == 1


# ============================================================================
# Edge cases
# ============================================================================

def test_unreadable_file_returns_error(db_session):
    """File non valido → errore catastrofico nel report."""
    user = models.User(id=1, email="x@x.it", hashed_password="x", role="admin")
    db_session.add(user); db_session.commit()

    report = import_excel(db_session, b"not an xlsx", user.id)
    assert len(report.errors) >= 1
    assert "non leggibile" in report.errors[0].reason.lower()


def test_param_invalid_condition_skipped(db_session):
    """Parametro con condition syntax errata → riga saltata, parametro non aggiornato."""
    lang, user = _seed_full(db_session)

    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Parameters")
    from services.excel_export import PARAMETERS_HEADERS
    ws.append(PARAMETERS_HEADERS)
    ws.append([
        "FGM", 1, "Updated", "", "", "",
        "short", "long",
        "((( invalid syntax",  # condizione errata
        "", "Yes",
    ])

    p_before = db_session.query(models.ParameterDef).filter_by(id="FGM").one()
    name_before = p_before.name

    report = import_excel(db_session, _wb_to_bytes(wb), user.id)
    db_session.commit()

    p_after = db_session.query(models.ParameterDef).filter_by(id="FGM").one()
    assert p_after.name == name_before  # NON aggiornato per errore di sintassi
    assert any("Sintassi" in e.reason for e in report.errors)


# ============================================================================
# Round-trip lossless: backup → restore preserva motivazioni e admin notes
# ============================================================================

def test_roundtrip_preserves_motivations(db_session):
    """Lingua con motivazione su FGM_02 → export → wipe → import → motivazione
    presente sull'answer di FGM_02 (round-trip lossless)."""
    lang, user = _seed_full(db_session)
    # Aggiungo una motivazione all'answer FGM_02 (creandola: c'è solo FGM_01 in seed)
    ans2 = models.Answer(language_id="ITA", question_id="FGM_02",
                         response_text="no", comments="", status="approved")
    db_session.add(ans2)
    db_session.flush()
    mot = db_session.query(models.Motivation).filter_by(code="MOT_X").one()
    db_session.add(models.AnswerMotivation(answer_id=ans2.id, motivation_id=mot.id))
    db_session.commit()

    wb = build_language_workbook(db_session, lang, is_admin=True)
    file_bytes = _wb_to_bytes(wb)

    report = import_excel(db_session, file_bytes, user.id)
    db_session.commit()
    assert report.errors == [], f"Errori inattesi: {report.errors}"

    # Recupero answer FGM_02 dopo round-trip
    a = db_session.query(models.Answer).filter_by(
        language_id="ITA", question_id="FGM_02"
    ).one()
    mot_codes = [
        db_session.query(models.Motivation).get(am.motivation_id).code
        for am in a.answer_motivations
    ]
    assert "MOT_X" in mot_codes


def test_roundtrip_preserves_admin_note(db_session):
    """Lingua con admin_note su FGM → export → import → admin_note ripristinata."""
    lang, user = _seed_full(db_session)
    db_session.add(models.LanguageParameterStatus(
        language_id="ITA", parameter_id="FGM",
        admin_note="Nota admin di prova\ncon a capo",
        is_unsure=False,
    ))
    db_session.commit()

    wb = build_language_workbook(db_session, lang, is_admin=True)
    file_bytes = _wb_to_bytes(wb)

    report = import_excel(db_session, file_bytes, user.id)
    db_session.commit()
    assert report.errors == [], f"Errori inattesi: {report.errors}"

    s = db_session.query(models.LanguageParameterStatus).filter_by(
        language_id="ITA", parameter_id="FGM"
    ).one()
    assert s.admin_note == "Nota admin di prova\ncon a capo"
