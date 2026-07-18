"""
Test end-to-end del flusso backup → wipe → restore.

Verifica che il bundle ZIP prodotto da build_backup_zip_bytes sia restorable
tramite restore_backup_bundle: le lingue, parametri, domande, motivazioni,
glossario e i dati di compilazione devono tornare uguali dopo un wipe + restore.
"""
import pytest

import models
from services.excel_export import build_backup_zip_bytes, build_full_backup_zip_bytes
from services.backup_restore import restore_backup_bundle


def _seed_full(db_session):
    user = models.User(
        id=1, email="alice@test.it", hashed_password="x",
        name="Alice", surname="Smith", role="admin",
    )
    db_session.add(user)

    lang = models.Language(
        id="ITA", name_full="Italiano", position=1,
        family="Romance", top_level_family="Indo-European", grp="Italo-Western",
        latitude=42.5, longitude=12.0, historical_language=False,
        isocode="it", glottocode="ital1282",
        informant="Mario Rossi", supervisor="Cristina Guardiano",
        source="Various sources", location="Italia",
        status="validated",
    )
    db_session.add(lang)

    param = models.ParameterDef(
        id="FGM", position=1, name="Feature Geometry Marker",
        short_description="Test", long_description="",
        is_active=True,
    )
    db_session.add(param)

    q1 = models.Question(id="FGM_01", parameter_id="FGM", text="Q1?",
                         is_stop_question=False, is_active=True)
    q2 = models.Question(id="FGM_02", parameter_id="FGM", text="Q2?",
                         is_stop_question=False, is_active=True)
    db_session.add_all([q1, q2])

    mot = models.Motivation(code="MOT_X", label="Not applicable")
    db_session.add(mot)
    db_session.flush()

    db_session.add(models.QuestionAllowedMotivation(question_id="FGM_02", motivation_id=mot.id))

    ans1 = models.Answer(language_id="ITA", question_id="FGM_01",
                         response_text="yes", comments="ok", status="approved")
    ans2 = models.Answer(language_id="ITA", question_id="FGM_02",
                         response_text="no", comments="", status="approved")
    db_session.add_all([ans1, ans2])
    db_session.flush()

    db_session.add(models.AnswerMotivation(answer_id=ans2.id, motivation_id=mot.id))
    db_session.add(models.Example(
        answer_id=ans1.id, number="1", textarea="Esempio",
        gloss="g", translation="t", reference="r",
    ))
    db_session.add(models.LanguageParameterStatus(
        language_id="ITA", parameter_id="FGM",
        admin_note="Nota admin",
        is_unsure=False,
    ))
    db_session.add(models.Glossary(word="alpha", description="first letter"))
    db_session.commit()
    return user


def test_backup_restore_roundtrip(db_session):
    """Backup completo → wipe → restore: lo stato del DB dev'essere ripristinato."""
    user = _seed_full(db_session)

    # 1. Genera bundle
    languages = db_session.query(models.Language).all()
    zip_bytes = build_backup_zip_bytes(db_session, languages)

    # 2. Restore con wipe=True su una sessione "sporca": ripulisce e re-importa
    report = restore_backup_bundle(db_session, zip_bytes, user.id, wipe=True)
    db_session.commit()

    # Nessun errore atteso (o al massimo errori non bloccanti per le motivazioni
    # — ma in questo seed minimale tutto dovrebbe filare liscio)
    blocking = [e for e in report.errors if "Motivation" not in e.get("reason", "")]
    assert blocking == [], f"Errori bloccanti: {blocking}"

    # 3. Verifica DB ripristinato
    # Lingue
    lang = db_session.query(models.Language).filter_by(id="ITA").one()
    assert lang.name_full == "Italiano"
    assert lang.family == "Romance"
    assert lang.isocode == "it"
    assert float(lang.latitude) == 42.5

    # Schema (parametri / domande / motivazioni)
    p = db_session.query(models.ParameterDef).filter_by(id="FGM").one()
    assert p.name == "Feature Geometry Marker"
    assert db_session.query(models.Question).filter_by(parameter_id="FGM").count() == 2
    assert db_session.query(models.Motivation).filter_by(code="MOT_X").count() == 1

    # Risposte + esempi + motivazioni: ripristinate dal Database_model
    answers = db_session.query(models.Answer).filter_by(language_id="ITA").all()
    by_qid = {a.question_id: a for a in answers}
    assert by_qid["FGM_01"].response_text == "yes"
    assert by_qid["FGM_02"].response_text == "no"
    assert len(by_qid["FGM_01"].examples) == 1
    # Motivation MOT_X ripristinata sull'answer FGM_02
    mot_codes = [
        db_session.get(models.Motivation, am.motivation_id).code
        for am in by_qid["FGM_02"].answer_motivations
    ]
    assert "MOT_X" in mot_codes

    # Admin note ripristinata
    s = db_session.query(models.LanguageParameterStatus).filter_by(
        language_id="ITA", parameter_id="FGM"
    ).one()
    assert s.admin_note == "Nota admin"

    # Glossario ripristinato
    g = db_session.query(models.Glossary).filter_by(word="alpha").one()
    assert g.description == "first letter"

    # Files processati
    assert "schema.xlsx" in report.files_processed
    assert "languages_metadata.xlsx" in report.files_processed
    assert "glossary.xlsx" in report.files_processed
    assert "languages/ITA.xlsx" in report.files_processed
    assert "ITA" in report.languages_restored


def test_backup_restore_bad_zip(db_session):
    """Bundle non valido → errore esplicito, no crash."""
    user = _seed_full(db_session)

    report = restore_backup_bundle(db_session, b"not a zip", user.id, wipe=False)
    assert any("ZIP" in e["reason"] for e in report.errors)
    assert report.files_processed == []


def _seed_extras(db_session, user):
    """Aggiunge dati per le tabelle extras: site_content, submission,
    parameter_submission, archived_question (con figli)."""
    db_session.add(models.SiteContent(
        key="how_to_cite_body",
        content="<p>Cite this work as...</p>",
        page="HowToCite",
        updated_by_id=user.id,
    ))

    sub = models.Submission(
        language_id="ITA",
        submitted_by_id=user.id,
        note="Initial submission",
    )
    db_session.add(sub)
    db_session.flush()
    db_session.add_all([
        models.SubmissionAnswer(submission_id=sub.id, question_code="FGM_01",
                                response_text="yes", comments="ok"),
        models.SubmissionExample(submission_id=sub.id, question_code="FGM_01",
                                 textarea="Esempio sub", gloss="g", translation="t",
                                 is_test=True),
        models.SubmissionAnswerMotivation(submission_id=sub.id, question_code="FGM_02",
                                          motivation_code="MOT_X", motivation_label="Not applicable"),
        models.SubmissionParam(submission_id=sub.id, parameter_id="FGM",
                               value_orig="+", warning_orig=False,
                               value_eval="+", warning_eval=False),
    ])

    psub = models.ParameterSubmission(
        parameter_id="FGM", parameter_name="Feature Geometry Marker",
        submitted_by_id=user.id, note="Param snapshot",
        short_description="Test", long_description="",
        is_active=True, position=1, schema="", param_type="", level_of_comparison="",
    )
    db_session.add(psub)
    db_session.flush()
    psq = models.ParameterSubmissionQuestion(
        submission_id=psub.id, question_code="FGM_01",
        text="Q1?", is_stop_question=False, is_active=True,
    )
    db_session.add(psq)
    db_session.flush()
    db_session.add(models.ParameterSubmissionAllowedMotivation(
        question_id=psq.id, motivation_code="MOT_X", motivation_label="Not applicable",
    ))

    aq = models.ArchivedQuestion(
        original_question_id="OBS_01", parameter_id="OBS", parameter_name="Obsolete",
        text="Old question", is_stop_question=False, is_active=False,
        archived_by_id=user.id, archive_note="Reworded",
        answers_count=1, examples_count=1,
    )
    db_session.add(aq)
    db_session.flush()
    db_session.add(models.ArchivedQuestionMotivation(
        archived_question_id=aq.id, motivation_code="MOT_X", motivation_label="Not applicable",
    ))
    aa = models.ArchivedAnswer(
        archived_question_id=aq.id, language_id="ITA", language_name_full="Italiano",
        status="approved", response_text="yes", comments="legacy",
    )
    db_session.add(aa)
    db_session.flush()
    db_session.add(models.ArchivedExample(
        archived_answer_id=aa.id, number="1", textarea="Old example",
    ))
    db_session.add(models.ArchivedAnswerMotivation(
        archived_answer_id=aa.id, motivation_code="MOT_X", motivation_label="Not applicable",
    ))

    db_session.add(models.ParameterChangeLog(
        parameter_id="FGM", user_id=user.id, change_note="Refined definition",
    ))

    db_session.commit()


def test_create_language_submission_copies_is_test(db_session):
    """Lo snapshot di backup copia il flag is_test dall'Example originale."""
    from services.backup_service import create_language_submission
    user = _seed_full(db_session)
    ex = db_session.query(models.Example).first()
    ex.is_test = True
    db_session.commit()

    lang = db_session.query(models.Language).filter_by(id="ITA").one()
    sub, _ = create_language_submission(db_session, lang, user.id, note="snap")
    db_session.commit()

    sub_exs = db_session.query(models.SubmissionExample).filter_by(submission_id=sub.id).all()
    assert len(sub_exs) == 1
    assert sub_exs[0].is_test is True


def test_full_backup_restore_roundtrip(db_session):
    """Bundle full → wipe → restore: anche le tabelle extras tornano identiche."""
    user = _seed_full(db_session)
    _seed_extras(db_session, user)

    languages = db_session.query(models.Language).all()
    zip_bytes = build_full_backup_zip_bytes(db_session, languages)

    # Verifica che il bundle contenga la cartella extras/
    import zipfile, io
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()
    assert "extras/site_content.xlsx" in names
    assert "extras/submissions.xlsx" in names
    assert "extras/parameter_submissions.xlsx" in names
    assert "extras/archived_questions.xlsx" in names

    # Restore con wipe → estras vengono ripopolati
    report = restore_backup_bundle(db_session, zip_bytes, user.id, wipe=True)
    db_session.commit()

    blocking = [e for e in report.errors if "Motivation" not in e.get("reason", "")]
    assert blocking == [], f"Errori bloccanti: {blocking}"

    # SiteContent: upsert per key
    sc = db_session.query(models.SiteContent).filter_by(key="how_to_cite_body").one()
    assert sc.page == "HowToCite"
    assert "Cite this work" in sc.content

    # Submission + figlie
    subs = db_session.query(models.Submission).all()
    assert len(subs) == 1
    sub = subs[0]
    assert sub.language_id == "ITA"
    assert sub.note == "Initial submission"
    assert len(sub.answers) == 1
    assert sub.answers[0].response_text == "yes"
    assert len(sub.examples) == 1
    assert sub.examples[0].textarea == "Esempio sub"
    assert sub.examples[0].is_test is True  # il flag esempio-di-test sopravvive al round-trip zip
    assert len(sub.answer_motivations) == 1
    assert sub.answer_motivations[0].motivation_code == "MOT_X"
    assert len(sub.params) == 1
    assert sub.params[0].value_orig == "+"

    # ParameterSubmission + figlie
    psubs = db_session.query(models.ParameterSubmission).all()
    assert len(psubs) == 1
    psub = psubs[0]
    assert psub.parameter_id == "FGM"
    assert psub.note == "Param snapshot"
    assert len(psub.questions) == 1
    psq = psub.questions[0]
    assert psq.question_code == "FGM_01"
    assert len(psq.allowed_motivations) == 1
    assert psq.allowed_motivations[0].motivation_code == "MOT_X"

    # ArchivedQuestion + figli/nipoti
    aqs = db_session.query(models.ArchivedQuestion).all()
    assert len(aqs) == 1
    aq = aqs[0]
    assert aq.original_question_id == "OBS_01"
    assert aq.archive_note == "Reworded"
    assert len(aq.allowed_motivations) == 1
    assert len(aq.answers) == 1
    aa = aq.answers[0]
    assert aa.language_id == "ITA"
    assert aa.response_text == "yes"
    assert len(aa.examples) == 1
    assert aa.examples[0].textarea == "Old example"
    assert len(aa.answer_motivations) == 1


def test_full_backup_restore_no_wipe_skips_snapshots(db_session):
    """Senza wipe: site_content viene comunque upsertato (chiave naturale),
    ma submissions/parameter_submissions/archived_questions sono saltati per
    evitare duplicati su PK auto-increment."""
    user = _seed_full(db_session)
    _seed_extras(db_session, user)

    languages = db_session.query(models.Language).all()
    zip_bytes = build_full_backup_zip_bytes(db_session, languages)

    # Conta quanti record ci sono PRIMA del restore
    n_subs_before = db_session.query(models.Submission).count()
    n_psubs_before = db_session.query(models.ParameterSubmission).count()
    n_aqs_before = db_session.query(models.ArchivedQuestion).count()

    report = restore_backup_bundle(db_session, zip_bytes, user.id, wipe=False)
    db_session.commit()

    # site_content: upsertato comunque
    assert "extras/site_content.xlsx" in report.files_processed
    sc = db_session.query(models.SiteContent).filter_by(key="how_to_cite_body").one()
    assert "Cite this work" in sc.content

    # snapshot tables: saltate
    assert "extras/submissions.xlsx" in report.files_skipped
    assert "extras/parameter_submissions.xlsx" in report.files_skipped
    assert "extras/archived_questions.xlsx" in report.files_skipped

    # E i conteggi non sono cambiati
    assert db_session.query(models.Submission).count() == n_subs_before
    assert db_session.query(models.ParameterSubmission).count() == n_psubs_before
    assert db_session.query(models.ArchivedQuestion).count() == n_aqs_before


def test_full_backup_restores_change_logs_and_flags(db_session):
    """parameter_change_logs e i flag unsure/needs_review sopravvivono al
    wipe+restore (prima venivano persi: il wipe li cancellava e il bundle
    non li conteneva)."""
    user = _seed_full(db_session)
    _seed_extras(db_session, user)

    # Flag accesi sulla riga di status creata da _seed_full
    status = db_session.query(models.LanguageParameterStatus).filter_by(
        language_id="ITA", parameter_id="FGM").one()
    status.is_unsure = True
    status.needs_review = True
    db_session.commit()

    languages = db_session.query(models.Language).all()
    zip_bytes = build_full_backup_zip_bytes(db_session, languages)

    import zipfile, io
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()
    assert "extras/parameter_change_logs.xlsx" in names
    assert "extras/parameter_flags.xlsx" in names

    # La riga di status caricata sopra resta nella identity map: il wipe la
    # cancella via SQL raw e il restore la ricrea con la stessa PK -> senza
    # expunge SQLAlchemy emette un SAWarning di identity conflict.
    user_id = user.id
    db_session.expunge_all()

    report = restore_backup_bundle(db_session, zip_bytes, user_id, wipe=True)
    db_session.commit()

    blocking = [e for e in report.errors if "Motivation" not in e.get("reason", "")]
    assert blocking == [], f"Errori bloccanti: {blocking}"

    # Change log ripristinato, autore risolto via email
    logs = db_session.query(models.ParameterChangeLog).all()
    assert len(logs) == 1
    assert logs[0].parameter_id == "FGM"
    assert logs[0].change_note == "Refined definition"
    assert logs[0].user_id == user.id

    # Flag ripristinati; l'admin note (dal file lingua) convive coi flag
    s = db_session.query(models.LanguageParameterStatus).filter_by(
        language_id="ITA", parameter_id="FGM").one()
    assert s.is_unsure is True
    assert s.needs_review is True
    assert s.admin_note == "Nota admin"


def test_flags_upsert_without_wipe_change_logs_skipped(db_session):
    """Senza wipe: i flag vengono comunque riallineati al bundle (upsert su
    chiave naturale), i change log invece sono saltati per evitare duplicati."""
    user = _seed_full(db_session)
    _seed_extras(db_session, user)

    status = db_session.query(models.LanguageParameterStatus).filter_by(
        language_id="ITA", parameter_id="FGM").one()
    status.is_unsure = True
    db_session.commit()

    languages = db_session.query(models.Language).all()
    zip_bytes = build_full_backup_zip_bytes(db_session, languages)

    # Dopo l'export i flag cambiano: il restore deve riportarli al bundle
    status.is_unsure = False
    status.needs_review = True
    db_session.commit()
    n_logs_before = db_session.query(models.ParameterChangeLog).count()

    report = restore_backup_bundle(db_session, zip_bytes, user.id, wipe=False)
    db_session.commit()

    assert "extras/parameter_flags.xlsx" in report.files_processed
    assert "extras/parameter_change_logs.xlsx" in report.files_skipped
    assert db_session.query(models.ParameterChangeLog).count() == n_logs_before

    s = db_session.query(models.LanguageParameterStatus).filter_by(
        language_id="ITA", parameter_id="FGM").one()
    assert s.is_unsure is True      # tornato com'era nel bundle
    assert s.needs_review is False  # nel bundle era spento


def test_full_backup_restores_aliases(db_session):
    """Gli alias storici (lingue/parametri/question) sopravvivono al
    wipe+restore e l'upsert per old_id non crea duplicati al secondo giro."""
    user = _seed_full(db_session)
    db_session.add_all([
        models.LanguageAlias(language_id="ITA", old_id="OIT"),
        models.ParameterAlias(parameter_id="FGM", old_id="FGX"),
        models.QuestionAlias(question_id="FGM_01", old_id="FGX_01"),
    ])
    db_session.commit()

    languages = db_session.query(models.Language).all()
    zip_bytes = build_full_backup_zip_bytes(db_session, languages)

    import zipfile, io
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        assert "extras/aliases.xlsx" in zf.namelist()

    user_id = user.id
    db_session.expunge_all()
    report = restore_backup_bundle(db_session, zip_bytes, user_id, wipe=True)
    db_session.commit()

    blocking = [e for e in report.errors if "Motivation" not in e.get("reason", "")]
    assert blocking == [], f"Errori bloccanti: {blocking}"

    la = db_session.query(models.LanguageAlias).one()
    assert (la.old_id, la.language_id) == ("OIT", "ITA")
    pa = db_session.query(models.ParameterAlias).one()
    assert (pa.old_id, pa.parameter_id) == ("FGX", "FGM")
    qa = db_session.query(models.QuestionAlias).one()
    assert (qa.old_id, qa.question_id) == ("FGX_01", "FGM_01")

    # Secondo restore senza wipe: upsert, niente duplicati
    report2 = restore_backup_bundle(db_session, zip_bytes, user_id, wipe=False)
    db_session.commit()
    assert "extras/aliases.xlsx" in report2.files_processed
    assert db_session.query(models.LanguageAlias).count() == 1
    assert db_session.query(models.ParameterAlias).count() == 1
    assert db_session.query(models.QuestionAlias).count() == 1


def test_full_backup_restores_legal_documents_consents_and_pdfs(db_session, tmp_path, monkeypatch):
    """Documenti legali, consensi e PDF viaggiano nel bundle full: dopo una
    perdita totale (righe cancellate + PDF sparito) il restore li ricrea,
    is_current viene rinormalizzato sull'ultima versione e il secondo
    restore non duplica nulla."""
    from datetime import datetime
    import services.excel_export as excel_export_mod
    import services.backup_restore as backup_restore_mod
    monkeypatch.setattr(excel_export_mod, "LEGAL_DOCUMENTS_DIR", str(tmp_path))
    monkeypatch.setattr(backup_restore_mod, "LEGAL_DOCUMENTS_DIR", str(tmp_path))

    user = _seed_full(db_session)
    doc_v1 = models.LegalDocument(
        type="terms_of_use", version="v1.0", file_path="tou_v1.pdf",
        sha256="a" * 64, published_at=datetime(2026, 1, 1), is_current=False,
        vexatious_clauses=["7", "8"], note="prima versione",
    )
    doc_v2 = models.LegalDocument(
        type="terms_of_use", version="v1.1", file_path="tou_v11.pdf",
        sha256="b" * 64, published_at=datetime(2026, 6, 1), is_current=True,
        vexatious_clauses=["7", "8", "9.2"],
    )
    db_session.add_all([doc_v1, doc_v2])
    db_session.flush()
    db_session.add(models.Consent(
        user_id=user.id, legal_document_id=doc_v2.id,
        accepted_at=datetime(2026, 6, 2, 10, 30), ip_address="1.2.3.4",
        user_agent="TestBrowser/1.0", method="first_login_modal",
        vexatious_clauses_approved=True,
    ))
    db_session.commit()
    (tmp_path / "tou_v11.pdf").write_bytes(b"%PDF-1.4 fake")

    languages = db_session.query(models.Language).all()
    zip_bytes = build_full_backup_zip_bytes(db_session, languages)

    import zipfile, io
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()
    assert "extras/legal_documents.xlsx" in names
    assert "extras/legal_pdfs/tou_v11.pdf" in names
    # tou_v1.pdf non esiste su disco: nel bundle c'e' solo il metadato
    assert "extras/legal_pdfs/tou_v1.pdf" not in names

    # Perdita totale: righe cancellate e PDF sparito
    db_session.query(models.Consent).delete()
    db_session.query(models.LegalDocument).delete()
    db_session.commit()
    (tmp_path / "tou_v11.pdf").unlink()

    user_id = user.id
    db_session.expunge_all()
    report = restore_backup_bundle(db_session, zip_bytes, user_id, wipe=False)
    db_session.commit()

    blocking = [e for e in report.errors if "Motivation" not in e.get("reason", "")]
    assert blocking == [], f"Errori bloccanti: {blocking}"

    docs = {d.version: d for d in db_session.query(models.LegalDocument).all()}
    assert set(docs) == {"v1.0", "v1.1"}
    assert docs["v1.1"].is_current is True
    assert docs["v1.0"].is_current is False
    assert docs["v1.0"].note == "prima versione"
    assert docs["v1.1"].vexatious_clauses == ["7", "8", "9.2"]
    assert docs["v1.1"].sha256 == "b" * 64

    consents = db_session.query(models.Consent).all()
    assert len(consents) == 1
    assert consents[0].user_id == user_id
    assert consents[0].legal_document_id == docs["v1.1"].id
    assert consents[0].vexatious_clauses_approved is True
    assert consents[0].ip_address == "1.2.3.4"

    # PDF rimaterializzato
    assert (tmp_path / "tou_v11.pdf").read_bytes() == b"%PDF-1.4 fake"

    # Secondo restore: upsert, niente duplicati
    report2 = restore_backup_bundle(db_session, zip_bytes, user_id, wipe=False)
    db_session.commit()
    assert db_session.query(models.LegalDocument).count() == 2
    assert db_session.query(models.Consent).count() == 1


def test_full_backup_restores_entity_versions(db_session):
    """La History (entity_versions) viaggia come jsonl nel bundle full:
    insert-if-missing idempotente, snapshot JSON preservato fedelmente
    anche oltre il limite di 32k caratteri di una cella Excel."""
    from datetime import datetime
    user = _seed_full(db_session)

    big_snapshot = {"name": "Feature Geometry Marker", "long_description": "x" * 40000}
    db_session.add_all([
        models.EntityVersion(
            entity_type="parameter", entity_id="FGM", snapshot=big_snapshot,
            operation="update", source="manual", note="test note",
            user_id=user.id, created_at=datetime(2026, 3, 1, 12, 0),
        ),
        models.EntityVersion(
            entity_type="language", entity_id="ITA", snapshot={"name_full": "Italiano"},
            operation="create", source="excel_import",
            created_at=datetime(2026, 2, 1, 9, 30),
        ),
    ])
    db_session.commit()

    languages = db_session.query(models.Language).all()
    zip_bytes = build_full_backup_zip_bytes(db_session, languages)

    import zipfile, io
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        assert "extras/entity_versions.jsonl" in zf.namelist()

    # Perdita delle righe (la tabella non e' nel wipe: simulo cancellazione)
    db_session.query(models.EntityVersion).delete()
    db_session.commit()

    user_id = user.id
    db_session.expunge_all()
    report = restore_backup_bundle(db_session, zip_bytes, user_id, wipe=False)
    db_session.commit()

    blocking = [e for e in report.errors if "Motivation" not in e.get("reason", "")]
    assert blocking == [], f"Errori bloccanti: {blocking}"

    # NB: il restore stesso genera altre righe di History (l'import versiona
    # le entità che tocca), quindi si asserisce sulle due righe seminate, non
    # sull'intera tabella.
    fgm = db_session.query(models.EntityVersion).filter_by(
        entity_id="FGM", created_at=datetime(2026, 3, 1, 12, 0)).one()
    assert fgm.snapshot == big_snapshot
    assert fgm.note == "test note"
    assert fgm.user_id == user_id

    ita = db_session.query(models.EntityVersion).filter_by(
        entity_id="ITA", created_at=datetime(2026, 2, 1, 9, 30)).one()
    assert ita.operation == "create"
    assert ita.source == "excel_import"
    assert ita.user_id is None

    # Secondo restore: dedupe, le righe seminate non vengono duplicate
    restore_backup_bundle(db_session, zip_bytes, user_id, wipe=False)
    db_session.commit()
    assert db_session.query(models.EntityVersion).filter_by(
        entity_id="FGM", created_at=datetime(2026, 3, 1, 12, 0)).count() == 1
    assert db_session.query(models.EntityVersion).filter_by(
        entity_id="ITA", created_at=datetime(2026, 2, 1, 9, 30)).count() == 1


def test_full_backup_restores_users_without_password(db_session):
    """Gli utenti viaggiano nel bundle SENZA hash password: su un DB privo di
    quell'utente il restore lo ricrea (password inutilizzabile, ruolo e stato
    preservati) e rimette l'assegnazione lingua->utente; la password degli
    utenti gia' esistenti non viene mai toccata."""
    user = _seed_full(db_session)
    bob = models.User(
        id=2, email="bob@test.it", hashed_password="bob-secret-hash",
        name="Bob", surname="Jones", role="user", is_active=False,
    )
    db_session.add(bob)
    lang = db_session.query(models.Language).filter_by(id="ITA").one()
    lang.assigned_user_id = 2
    db_session.commit()

    languages = db_session.query(models.Language).all()
    zip_bytes = build_full_backup_zip_bytes(db_session, languages)

    import zipfile, io
    from openpyxl import load_workbook
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        assert "extras/users.xlsx" in zf.namelist()
        # L'hash password NON deve stare nel file
        wb = load_workbook(io.BytesIO(zf.read("extras/users.xlsx")))
        all_cells = " ".join(
            str(c.value) for row in wb["Users"].iter_rows() for c in row if c.value
        )
        assert "bob-secret-hash" not in all_cells
        assert "bob@test.it" in all_cells

    # Bob sparisce (simula restore su DB che non lo ha)
    lang.assigned_user_id = None
    db_session.commit()
    db_session.query(models.User).filter_by(email="bob@test.it").delete()
    db_session.commit()

    user_id = user.id
    db_session.expunge_all()
    report = restore_backup_bundle(db_session, zip_bytes, user_id, wipe=True)
    db_session.commit()

    blocking = [e for e in report.errors if "Motivation" not in e.get("reason", "")]
    assert blocking == [], f"Errori bloccanti: {blocking}"

    bob2 = db_session.query(models.User).filter_by(email="bob@test.it").one()
    assert bob2.role == "user"
    assert bob2.is_active is False
    assert bob2.name == "Bob"
    # Password inutilizzabile, NON quella originale
    assert bob2.hashed_password != "bob-secret-hash"
    assert bob2.hashed_password.startswith("$2")  # hash bcrypt valido

    # Alice esisteva gia': password intatta
    alice = db_session.query(models.User).filter_by(email="alice@test.it").one()
    assert alice.hashed_password == "x"

    # Assegnazione lingua ripristinata sul nuovo id di Bob
    lang = db_session.query(models.Language).filter_by(id="ITA").one()
    assert lang.assigned_user_id == bob2.id


def test_standard_backup_compat_with_extras_aware_restore(db_session):
    """Il bundle standard (senza extras/) resta restorable con la nuova
    versione del restore — retrocompatibilità."""
    user = _seed_full(db_session)

    languages = db_session.query(models.Language).all()
    zip_bytes = build_backup_zip_bytes(db_session, languages)

    report = restore_backup_bundle(db_session, zip_bytes, user.id, wipe=True)
    db_session.commit()

    # Nessun file extras/* dovrebbe apparire
    assert not any(p.startswith("extras/") for p in report.files_processed)
    assert "ITA" in report.languages_restored
