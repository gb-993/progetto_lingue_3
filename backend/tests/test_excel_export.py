"""
Test per l'export Excel.

Due livelli:
  1. Test "header" — verifica che gli header dei 3 sheet "vecchi" (Database_model,
     Examples, Answers) siano IDENTICI a quelli del vecchio progetto Django.
     I valori del vecchio progetto sono stati copiati letteralmente da
     core/views/languages.py::_build_language_workbook (vedi _OLD_HEADERS).

  2. Test "round-trip" — crea un mini-DB con dati realistici, genera il workbook,
     lo riapre con openpyxl e verifica:
     - Sheet names + ordine
     - Header riga 1 di ogni sheet
     - Numero righe coerenti con i dati
     - Multi-line concatenation degli esempi nel sheet Database_model
     - Round-trip: salva su BytesIO, riapri, contenuto identico.
"""
import io
import zipfile
from datetime import datetime

import pytest
from openpyxl import Workbook, load_workbook

import models
from services.excel_export import (
    build_language_workbook,
    build_language_list_workbook,
    build_schema_workbook,
    build_glossary_workbook,
    build_backup_zip_bytes,
    DATABASE_MODEL_HEADERS,
    EXAMPLES_HEADERS,
    ANSWERS_HEADERS,
    LANGUAGE_LIST_HEADERS,
    MOTIVATIONS_HEADERS,
    PARAMETERS_HEADERS,
    QUESTIONS_HEADERS,
    QUESTION_ALLOWED_MOTIVATIONS_HEADERS,
    GLOSSARY_HEADERS,
)


# ============================================================================
# 1) HEADER TEST — confronto bit-per-bit con il vecchio progetto Django
# ============================================================================

# Valori COPIATI letteralmente dal vecchio _build_language_workbook in
# core/views/languages.py (vedi messaggio dell'utente). Qualsiasi modifica
# accidentale all'ordine o ai nomi delle colonne fa fallire questo test.

# Database_model è stato ristrutturato (2026-05): le 3 colonne ridondanti
# rispetto allo schema globale (Question, Question_Examples_YES,
# Question_Intructions_Comments) sono state rimosse, e in coda sono state
# aggiunte 2 colonne per backup lossless: Motivations, Admin_Note.
_EXPECTED_DATABASE_MODEL_HEADERS = [
    "Language",
    "Parameter_Label",
    "Question_ID",
    "Language_Answer",
    "Language_Comments",
    "Language_Examples",
    "Language_Example_Transliteration",
    "Language_Example_Gloss",
    "Language_Example_Translation",
    "Language_References",
    "Motivations",
    "Admin_Note",
]

_OLD_EXAMPLES_HEADERS = [
    "Language ID", "Question ID", "Example #",
    "Example text", "Transliteration", "Gloss", "English translation", "Reference",
]

_OLD_ANSWERS_HEADERS = [
    "Language ID", "Parameter Label", "Question ID", "Question",
    "Question status", "Answer", "Parameter value", "Motivation", "Comments",
]


def test_database_model_headers():
    assert DATABASE_MODEL_HEADERS == _EXPECTED_DATABASE_MODEL_HEADERS, (
        "Modifica all'ordine/contenuto delle colonne di 'Database_model'. "
        "Questo foglio è la fonte canonica per re-import / backup-restore. "
        "Verifica services/excel_export.py e services/excel_import.py "
        "(_import_compilation) restino allineati."
    )


def test_examples_headers_match_legacy():
    assert EXAMPLES_HEADERS == _OLD_EXAMPLES_HEADERS


def test_answers_headers_match_legacy():
    assert ANSWERS_HEADERS == _OLD_ANSWERS_HEADERS


def test_database_model_headers_count():
    assert len(DATABASE_MODEL_HEADERS) == 12


def test_examples_headers_count():
    assert len(EXAMPLES_HEADERS) == 8


def test_answers_headers_count():
    assert len(ANSWERS_HEADERS) == 9


# ============================================================================
# 2) ROUND-TRIP TEST con mini-DB
# ============================================================================

def _seed_basic(db_session):
    """Popola il DB con: 1 lingua, 1 parametro attivo, 2 domande,
    2 risposte (yes + no), 3 esempi sulla risposta yes, 1 motivazione."""
    user = models.User(
        email="alice@test.it", hashed_password="x", name="Alice", surname="Smith", role="user"
    )
    db_session.add(user); db_session.flush()

    lang = models.Language(
        id="ITA", name_full="Italiano", position=1,
        family="Romance", top_level_family="Indo-European", grp="Italo-Western",
        latitude=42.5, longitude=12.0, historical_language=False,
        isocode="it", glottocode="ital1282",
        informant="Mario Rossi", supervisor="Cristina Guardiano",
        source="Various sources", location="Italia",
        assigned_user_id=user.id,
    )
    db_session.add(lang)

    param = models.ParameterDef(
        id="FGM", position=1, name="Feature Geometry Marker",
        short_description="Test param", long_description="A longer description",
        implicational_condition="+ABC | -DEF",
        description_of_the_implicational_condition="Holds when ABC=+ or DEF=-",
        is_active=True, schema="Nominal", param_type="Binary", level_of_comparison="Macro",
    )
    db_session.add(param)

    q1 = models.Question(
        id="FGM_01", parameter_id="FGM", text="Does it have FGM marker?",
        instruction="Look for it.", instruction_yes="Provide examples.",
        instruction_no="Explain motivation.",
        example_yes="e.g. il libro = the book",
        help_info="More background info here.",
        is_stop_question=False, is_active=True,
    )
    q2 = models.Question(
        id="FGM_02", parameter_id="FGM", text="Is it productive?",
        is_stop_question=False, is_active=True,
    )
    db_session.add_all([q1, q2])

    mot = models.Motivation(code="MOT_X", label="Not applicable")
    db_session.add(mot); db_session.flush()

    qam = models.QuestionAllowedMotivation(question_id="FGM_02", motivation_id=mot.id)
    db_session.add(qam)

    ans1 = models.Answer(
        language_id="ITA", question_id="FGM_01",
        response_text="yes", comments="Some comment", status="approved",
    )
    ans2 = models.Answer(
        language_id="ITA", question_id="FGM_02",
        response_text="no", comments="", status="approved",
    )
    db_session.add_all([ans1, ans2]); db_session.flush()

    db_session.add(models.AnswerMotivation(answer_id=ans2.id, motivation_id=mot.id))

    for i, txt in enumerate(["Esempio uno", "Esempio due", "Esempio tre"], start=1):
        db_session.add(models.Example(
            answer_id=ans1.id, number=str(i),
            textarea=txt,
            transliteration=f"trans-{i}", gloss=f"gloss-{i}",
            translation=f"translation-{i}", reference=f"ref-{i}",
        ))

    db_session.commit()
    return lang


def _read_workbook_from_memory(wb: Workbook) -> Workbook:
    """Salva e riapre il workbook (verifica round-trip serializzazione)."""
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return load_workbook(buf, data_only=True)


def test_language_workbook_admin_has_four_sheets(db_session):
    """Da 2026-05 i fogli schema (Motivations/Parameters/Questions/QAM) NON
    sono più replicati in ogni per-lingua xlsx (vivono in schema.xlsx separato).
    L'admin riceve 4 sheet: Database_model + Answers + Examples + Admin Notes."""
    lang = _seed_basic(db_session)
    wb = build_language_workbook(db_session, lang, is_admin=True)
    wb2 = _read_workbook_from_memory(wb)
    assert wb2.sheetnames == [
        "Database_model", "Answers", "Examples", "Admin Notes",
    ]


def test_language_workbook_user_has_only_examples(db_session):
    lang = _seed_basic(db_session)
    wb = build_language_workbook(db_session, lang, is_admin=False)
    wb2 = _read_workbook_from_memory(wb)
    assert wb2.sheetnames == ["Examples"]


def test_database_model_sheet_headers_and_count(db_session):
    lang = _seed_basic(db_session)
    wb = build_language_workbook(db_session, lang, is_admin=True)
    wb2 = _read_workbook_from_memory(wb)
    ws = wb2["Database_model"]
    headers = [c.value for c in ws[1]]
    assert headers == _EXPECTED_DATABASE_MODEL_HEADERS

    # 2 question rows (FGM_01 e FGM_02), entrambe attive
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)]
    assert len(data_rows) == 2


def test_database_model_sheet_examples_concatenation(db_session):
    lang = _seed_basic(db_session)
    wb = build_language_workbook(db_session, lang, is_admin=True)
    wb2 = _read_workbook_from_memory(wb)
    ws = wb2["Database_model"]
    # Indici colonna by name, robusto a future ristrutturazioni
    h = {name: i for i, name in enumerate(_EXPECTED_DATABASE_MODEL_HEADERS)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[h["Question_ID"]] == "FGM_01":
            assert row[h["Language_Examples"]] == "Esempio uno\nEsempio due\nEsempio tre"
            assert row[h["Language_Example_Transliteration"]] == "trans-1\ntrans-2\ntrans-3"
            assert row[h["Language_Example_Gloss"]] == "gloss-1\ngloss-2\ngloss-3"
            assert row[h["Language_Example_Translation"]] == "translation-1\ntranslation-2\ntranslation-3"
            assert row[h["Language_References"]] == "ref-1\nref-2\nref-3"
            break
    else:
        pytest.fail("Riga FGM_01 non trovata in Database_model")


def test_examples_sheet_one_row_per_example(db_session):
    lang = _seed_basic(db_session)
    wb = build_language_workbook(db_session, lang, is_admin=True)
    wb2 = _read_workbook_from_memory(wb)
    ws = wb2["Examples"]
    headers = [c.value for c in ws[1]]
    assert headers == _OLD_EXAMPLES_HEADERS
    # 3 esempi totali
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True)]
    assert len(rows) == 3
    # primo esempio: ITA, FGM_01, "1", "Esempio uno", ...
    assert rows[0][0] == "ITA"
    assert rows[0][1] == "FGM_01"
    assert rows[0][2] == "1"
    assert rows[0][3] == "Esempio uno"


def test_answers_sheet_headers_and_motivations(db_session):
    lang = _seed_basic(db_session)
    wb = build_language_workbook(db_session, lang, is_admin=True)
    wb2 = _read_workbook_from_memory(wb)
    ws = wb2["Answers"]
    headers = [c.value for c in ws[1]]
    assert headers == _OLD_ANSWERS_HEADERS
    # FGM_02 deve avere "Not applicable" in Motivation
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True)]
    fgm02 = next(r for r in rows if r[2] == "FGM_02")
    assert fgm02[5] == "no"  # Answer
    assert fgm02[7] == "Not applicable"  # Motivation


def test_schema_sheets_have_correct_headers(db_session):
    _seed_basic(db_session)
    wb = build_schema_workbook(db_session)
    wb2 = _read_workbook_from_memory(wb)
    assert wb2.sheetnames == [
        "Motivations", "Parameters", "Questions", "QuestionAllowedMotivations",
    ]
    assert [c.value for c in wb2["Motivations"][1]] == MOTIVATIONS_HEADERS
    assert [c.value for c in wb2["Parameters"][1]] == PARAMETERS_HEADERS
    assert [c.value for c in wb2["Questions"][1]] == QUESTIONS_HEADERS
    assert [c.value for c in wb2["QuestionAllowedMotivations"][1]] == QUESTION_ALLOWED_MOTIVATIONS_HEADERS


def test_schema_workbook_data_rows(db_session):
    _seed_basic(db_session)
    wb = build_schema_workbook(db_session)
    wb2 = _read_workbook_from_memory(wb)

    # Motivations: 1 entry
    rows = list(wb2["Motivations"].iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][1] == "MOT_X"

    # Parameters: 1 entry
    rows = list(wb2["Parameters"].iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][0] == "FGM"

    # Questions: 2 entries
    rows = list(wb2["Questions"].iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2

    # QAM: 1 entry (FGM_02 -> MOT_X)
    rows = list(wb2["QuestionAllowedMotivations"].iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0] == ("FGM_02", "MOT_X")


def test_language_list_workbook(db_session):
    _seed_basic(db_session)
    languages = db_session.query(models.Language).all()
    wb = build_language_list_workbook(db_session, languages)
    wb2 = _read_workbook_from_memory(wb)
    assert wb2.sheetnames == ["Languages"]
    headers = [c.value for c in wb2["Languages"][1]]
    assert headers == LANGUAGE_LIST_HEADERS
    rows = list(wb2["Languages"].iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    # Verifica un paio di celle chiave
    assert rows[0][0] == "Italiano"  # Name
    assert rows[0][1] == "ITA"       # ID
    assert rows[0][5] == "it"        # ISO code
    assert rows[0][12] == "No"       # Historical (mapped to "Yes"/"No")
    assert rows[0][14] == "pending"  # Status


def test_language_list_user_metadata_export_works_with_zero_languages(db_session):
    """Edge case: nessuna lingua → workbook con solo header."""
    wb = build_language_list_workbook(db_session, [])
    wb2 = _read_workbook_from_memory(wb)
    ws = wb2["Languages"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows == []
    assert [c.value for c in ws[1]] == LANGUAGE_LIST_HEADERS


# ============================================================================
# 3) GLOSSARY WORKBOOK
# ============================================================================

def test_glossary_workbook_empty_db(db_session):
    """DB senza glossario → workbook con solo header."""
    wb = build_glossary_workbook(db_session)
    wb2 = _read_workbook_from_memory(wb)
    assert wb2.sheetnames == ["Glossary"]
    ws = wb2["Glossary"]
    assert [c.value for c in ws[1]] == GLOSSARY_HEADERS
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows == []


def test_glossary_workbook_with_entries(db_session):
    db_session.add_all([
        models.Glossary(word="alpha", description="first letter"),
        models.Glossary(word="beta", description="second letter"),
    ])
    db_session.commit()

    wb = build_glossary_workbook(db_session)
    wb2 = _read_workbook_from_memory(wb)
    ws = wb2["Glossary"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # Ordinati alfabeticamente
    assert rows == [("alpha", "first letter"), ("beta", "second letter")]


# ============================================================================
# 4) BACKUP ZIP — struttura del bundle completo
# ============================================================================

def test_backup_zip_structure(db_session):
    """Il backup zip deve contenere: schema.xlsx, languages_metadata.xlsx,
    glossary.xlsx, e una entry languages/<ID>.xlsx per ogni lingua."""
    _seed_basic(db_session)
    db_session.add(models.Glossary(word="hub", description="central node"))
    db_session.commit()

    languages = db_session.query(models.Language).all()
    data = build_backup_zip_bytes(db_session, languages)

    with zipfile.ZipFile(io.BytesIO(data), "r") as zf:
        names = set(zf.namelist())
        assert "schema.xlsx" in names
        assert "languages_metadata.xlsx" in names
        assert "glossary.xlsx" in names
        assert "languages/ITA.xlsx" in names

        # Schema dentro lo zip ha i 4 sheet attesi
        with zf.open("schema.xlsx") as f:
            schema_wb = load_workbook(io.BytesIO(f.read()), data_only=True)
            assert schema_wb.sheetnames == [
                "Motivations", "Parameters", "Questions", "QuestionAllowedMotivations",
            ]

        # Per-lingua dentro lo zip ha solo i 4 sheet (no più schema replicato)
        with zf.open("languages/ITA.xlsx") as f:
            lang_wb = load_workbook(io.BytesIO(f.read()), data_only=True)
            assert lang_wb.sheetnames == [
                "Database_model", "Answers", "Examples", "Admin Notes",
            ]


def test_backup_zip_progress_callback(db_session):
    """on_language deve essere chiamato (idx, total, lang) per ciascuna lingua."""
    _seed_basic(db_session)
    languages = db_session.query(models.Language).all()

    calls = []
    build_backup_zip_bytes(
        db_session, languages,
        on_language=lambda idx, total, lang: calls.append((idx, total, lang.id)),
    )
    assert calls == [(1, 1, "ITA")]
