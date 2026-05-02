"""
Router di import Excel (admin only).

Endpoint:
  POST /api/admin/import/excel              -> upload xlsx, ritorna ImportReport JSON
  POST /api/admin/import/error-report/xlsx  -> genera xlsx scaricabile con i soli errori
"""
from __future__ import annotations
from typing import List
from datetime import datetime
from time_utils import utc_now
import io

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font

import models
from dependencies import get_db, require_admin
from services.excel_import import import_excel


router = APIRouter(prefix="/api/admin/import", tags=["Import"])


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@router.post("/excel")
def post_import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    """
    Riceve un file .xlsx e lo importa secondo la strategia configurata.
    Ritorna sempre 200 con il report (anche se ci sono errori — il client
    decide cosa fare). Solo i casi catastrofici (file non leggibile,
    errori di sistema) ritornano 4xx/5xx.
    """
    fname = (file.filename or "").lower()
    if not fname.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400,
                            detail="Unsupported file format. Upload a .xlsx file")

    try:
        contents = file.file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read the file: {e}")

    if len(contents) == 0:
        raise HTTPException(status_code=400, detail="Empty file")
    if len(contents) > 50 * 1024 * 1024:  # 50 MB
        raise HTTPException(status_code=413, detail="File too large (max 50 MB)")

    report = import_excel(db, contents, current_user.id)
    return report.to_dict()


# ============================================================================
# Error report download
# ============================================================================

class ImportErrorRow(BaseModel):
    sheet: str
    row: int
    column: str | None = None
    value: str | None = None
    reason: str = ""


class ImportErrorPayload(BaseModel):
    errors: List[ImportErrorRow]
    target_language_id: str | None = None
    target_language_name: str | None = None


@router.post("/error-report/xlsx")
def post_error_report_xlsx(
    payload: ImportErrorPayload,
    current_user: models.User = Depends(require_admin),
):
    """Genera un xlsx scaricabile con un sheet 'Errors' contenente l'elenco completo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Errors"
    headers = ["Sheet", "Row", "Column", "Value", "Reason"]
    ws.append(headers)
    bold_white = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = bold_white

    for e in payload.errors:
        ws.append([
            e.sheet,
            e.row,
            e.column or "",
            (e.value or "")[:300],
            e.reason,
        ])

    # Larghezze indicative
    widths = [22, 8, 22, 40, 60]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if ws.max_row >= 2:
        ws.freeze_panes = "A2"

    # Sheet "Summary"
    if payload.target_language_name:
        ws_s = wb.create_sheet("Summary")
        ws_s.append(["Target language", payload.target_language_name, payload.target_language_id])
        ws_s.append(["Total errors", len(payload.errors)])
        ws_s.append(["Generated at", utc_now().isoformat()])
        for cell in ws_s["A"]:
            cell.font = Font(bold=True)
        ws_s.column_dimensions["A"].width = 22
        ws_s.column_dimensions["B"].width = 50

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    ts = utc_now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        buf, media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="PCM_import_errors_{ts}.xlsx"'},
    )
